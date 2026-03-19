import streamlit as st
import pandas as pd
import io
import openpyxl

st.set_page_config(page_title="Input Configuration", layout="wide")

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────
ALLOWED_UNITS = ["g", "kg", "t", "mL", "L", "m³", "kWh", "MWh", "GWh", "MMBtu"]

PRODUCT_TYPES    = ["Basic Chemical", "Specialty chemical", "Consumer product", "Pharmaceutical"]
TRL_OPTIONS      = ["Industrial (8 or 9)", "Pilot (5 to 7)", "Bench (3 or 4)", "Theoretical (1 or 2)"]
INFO_OPTIONS     = ["High", "Medium", "Low"]
SEVERITY_OPTIONS = ["High", "Medium", "Low"]
MAT_OPTIONS      = ["Solids", "Fluids and solids", "Fluids"]
SIZE_OPTIONS     = ["Large", "Medium", "Small"]

# Unscheduled equipment factors keyed by (TRL, info_availability)
UNSCHED_FACTORS = {
    ("Industrial (8 or 9)", "High"):   0.05,
    ("Industrial (8 or 9)", "Medium"): 0.10,
    ("Industrial (8 or 9)", "Low"):    0.15,
    ("Pilot (5 to 7)",       "Medium"): 0.15,
    ("Pilot (5 to 7)",       "Low"):    0.20,
    ("Bench (3 or 4)",       "Low"):    0.25,
    ("Theoretical (1 or 2)", "Low"):    0.30,
}

# Lang cost factors keyed by field name; tuple = (with_utility, without_utility)
LANG_FACTORS = {
    "Spare Parts":            (0.083, 0.051),
    "Equipment Setting":      (0.019, 0.019),
    "Unscheduled Equipment":  (0.110, 0.107),
    "Piping":                 (0.131, 0.368),
    "Civil":                  (0.041, 0.191),
    "Steel":                  (0.017, 0.272),
    "Instrumentals":          (0.033, 0.342),
    "Electrical":             (0.041, 0.335),
    "Insulation":             (0.015, 0.082),
    "Paint":                  (0.002, 0.040),
    "Field Office Staff":     (0.037, 0.172),
    "Construction Indirects": (0.077, 0.377),
    "Freight":                (0.052, 0.091),
    "Taxes and Permits":      (0.081, 0.142),
    "Engineering and HO":     (0.065, 0.684),
    "GA Overheads":           (0.049, 0.104),
    "Contract Fee":           (0.044, 0.161),
}

# Project contingency factors keyed by (TRL, process_severity) → fraction
PROJECT_CONTINGENCY: dict[tuple[str, str], float] = {
    ("Industrial (8 or 9)", "Low"):    0.15,
    ("Industrial (8 or 9)", "Medium"): 0.20,
    ("Industrial (8 or 9)", "High"):   0.25,
    ("Pilot (5 to 7)",       "Low"):    0.20,
    ("Pilot (5 to 7)",       "Medium"): 0.25,
    ("Pilot (5 to 7)",       "High"):   0.30,
    ("Bench (3 or 4)",       "Low"):    0.25,
    ("Bench (3 or 4)",       "Medium"): 0.30,
    ("Bench (3 or 4)",       "High"):   0.35,
    ("Theoretical (1 or 2)", "Low"):    0.30,
    ("Theoretical (1 or 2)", "Medium"): 0.35,
    ("Theoretical (1 or 2)", "High"):   0.40,
}

# Laboratory charges (% of OLC) keyed by product type
LAB_CHARGES: dict[str, float] = {
    "Basic Chemical":    0.10,
    "Specialty chemical":0.15,
    "Consumer product":  0.20,
    "Pharmaceutical":    0.25,
}

# Office labor (% of OLC) keyed by product type
OFFICE_LABOR: dict[str, float] = {
    "Basic Chemical":    0.10,
    "Specialty chemical":0.175,
    "Consumer product":  0.25,
    "Pharmaceutical":    0.175,
}

MIN_SALARY = 273.0   # USD/month — regulatory floor

# Maintenance and repairs (% of CAPEX) keyed by (material_type, product_type)
MAINTENANCE_REPAIRS: dict[tuple[str,str], float] = {
    ("Solids",           "Basic Chemical"):     0.02,
    ("Solids",           "Specialty chemical"): 0.03,
    ("Solids",           "Consumer product"):   0.04,
    ("Solids",           "Pharmaceutical"):     0.02,
    ("Fluids and solids","Basic Chemical"):     0.015,
    ("Fluids and solids","Specialty chemical"): 0.025,
    ("Fluids and solids","Consumer product"):   0.035,
    ("Fluids and solids","Pharmaceutical"):     0.015,
    ("Fluids",           "Basic Chemical"):     0.01,
    ("Fluids",           "Specialty chemical"): 0.02,
    ("Fluids",           "Consumer product"):   0.03,
    ("Fluids",           "Pharmaceutical"):     0.01,
}

# Operating supplies (% of Maintenance) keyed by process_severity
OPERATING_SUPPLIES: dict[str, float] = {
    "High":   0.0020,
    "Medium": 0.0015,
    "Low":    0.0010,
}

# Administrative overhead (% of OLC) keyed by product_type
ADMIN_OVERHEAD: dict[str, float] = {
    "Basic Chemical":    0.50,
    "Specialty chemical":0.60,
    "Consumer product":  0.70,
    "Pharmaceutical":    0.60,
}

# Manufacturing overhead (% of CAPEX) keyed by process_severity
MFG_OVERHEAD: dict[str, float] = {
    "High":   0.0070,
    "Medium": 0.0060,
    "Low":    0.0050,
}

# Taxes and insurance (% of CAPEX) keyed by process_severity
TAXES_INSURANCE: dict[str, float] = {
    "High":   0.050,
    "Medium": 0.032,
    "Low":    0.014,
}

# Patents and royalties (% of OPEX) keyed by (TRL, product_type); None = not applicable
PATENTS_ROYALTIES: dict[tuple[str,str], float | None] = {
    ("Industrial (8 or 9)", "Basic Chemical"):     0.010,
    ("Industrial (8 or 9)", "Specialty chemical"): 0.020,
    ("Industrial (8 or 9)", "Consumer product"):   0.040,
    ("Industrial (8 or 9)", "Pharmaceutical"):     0.060,
    ("Pilot (5 to 7)",       "Basic Chemical"):     None,
    ("Pilot (5 to 7)",       "Specialty chemical"): 0.010,
    ("Pilot (5 to 7)",       "Consumer product"):   0.020,
    ("Pilot (5 to 7)",       "Pharmaceutical"):     0.030,
    ("Bench (3 or 4)",       "Basic Chemical"):     None,
    ("Bench (3 or 4)",       "Specialty chemical"): 0.010,
    ("Bench (3 or 4)",       "Consumer product"):   0.020,
    ("Bench (3 or 4)",       "Pharmaceutical"):     0.030,
    ("Theoretical (1 or 2)", "Basic Chemical"):     None,
    ("Theoretical (1 or 2)", "Specialty chemical"): 0.010,
    ("Theoretical (1 or 2)", "Consumer product"):   0.020,
    ("Theoretical (1 or 2)", "Pharmaceutical"):     0.030,
}

# Distribution and selling (% of OPEX) keyed by product_type
DIST_SELLING: dict[str, float] = {
    "Basic Chemical":    0.08,
    "Specialty chemical":0.02,
    "Consumer product":  0.20,
    "Pharmaceutical":    0.14,
}

# Research and development (% of OPEX) keyed by (TRL, product_type)
R_AND_D: dict[tuple[str,str], float] = {
    ("Industrial (8 or 9)", "Basic Chemical"):     0.020,
    ("Industrial (8 or 9)", "Specialty chemical"): 0.030,
    ("Industrial (8 or 9)", "Consumer product"):   0.020,
    ("Industrial (8 or 9)", "Pharmaceutical"):     0.120,
    ("Pilot (5 to 7)",       "Basic Chemical"):     0.030,
    ("Pilot (5 to 7)",       "Specialty chemical"): 0.050,
    ("Pilot (5 to 7)",       "Consumer product"):   0.025,
    ("Pilot (5 to 7)",       "Pharmaceutical"):     0.170,
    ("Bench (3 or 4)",       "Basic Chemical"):     0.030,
    ("Bench (3 or 4)",       "Specialty chemical"): 0.050,
    ("Bench (3 or 4)",       "Consumer product"):   0.025,
    ("Bench (3 or 4)",       "Pharmaceutical"):     0.170,
    ("Theoretical (1 or 2)", "Basic Chemical"):     0.030,
    ("Theoretical (1 or 2)", "Specialty chemical"): 0.050,
    ("Theoretical (1 or 2)", "Consumer product"):   0.025,
    ("Theoretical (1 or 2)", "Pharmaceutical"):     0.170,
}

# TIC accuracy — Lower bound (%) keyed by (TRL, info_availability); None = N/A
TIC_LOWER: dict[tuple[str, str], float | None] = {
    ("Industrial (8 or 9)", "High"):    -15.0,
    ("Industrial (8 or 9)", "Medium"):  -20.0,
    ("Industrial (8 or 9)", "Low"):     -25.0,
    ("Pilot (5 to 7)",       "High"):    None,
    ("Pilot (5 to 7)",       "Medium"): -25.0,
    ("Pilot (5 to 7)",       "Low"):    -30.0,
    ("Bench (3 or 4)",       "High"):    None,
    ("Bench (3 or 4)",       "Medium"):  None,
    ("Bench (3 or 4)",       "Low"):    -40.0,
    ("Theoretical (1 or 2)", "High"):    None,
    ("Theoretical (1 or 2)", "Medium"):  None,
    ("Theoretical (1 or 2)", "Low"):    -50.0,
}

# TIC accuracy — Upper bound (%) keyed by (TRL, info_availability); None = N/A
TIC_UPPER: dict[tuple[str, str], float | None] = {
    ("Industrial (8 or 9)", "High"):    20.0,
    ("Industrial (8 or 9)", "Medium"):  30.0,
    ("Industrial (8 or 9)", "Low"):     40.0,
    ("Pilot (5 to 7)",       "High"):    None,
    ("Pilot (5 to 7)",       "Medium"):  40.0,
    ("Pilot (5 to 7)",       "Low"):     50.0,
    ("Bench (3 or 4)",       "High"):    None,
    ("Bench (3 or 4)",       "Medium"):  None,
    ("Bench (3 or 4)",       "Low"):     70.0,
    ("Theoretical (1 or 2)", "High"):    None,
    ("Theoretical (1 or 2)", "Medium"):  None,
    ("Theoretical (1 or 2)", "Low"):    100.0,
}

# Corporate tax rate by country
TAXES_BY_COUNTRY: dict[str, float] = {
    "Brazil": 0.340, "United States": 0.258, "Albania": 0.150, "Andorra": 0.100,
    "Angola": 0.250, "Anguilla": 0.000, "Argentina": 0.300, "Armenia": 0.180,
    "Aruba": 0.250, "Australia": 0.300, "Austria": 0.250, "Bahamas": 0.000,
    "Bahrain": 0.000, "Barbados": 0.055, "Belgium": 0.250, "Belize": 0.000,
    "Bermuda": 0.000, "Bosnia and Herzegovina": 0.100, "Botswana": 0.220,
    "British Virgin Islands": 0.000, "Brunei Darussalam": 0.185, "Bulgaria": 0.100,
    "Burkina Faso": 0.275, "Canada": 0.262, "Cayman Islands": 0.000, "Chile": 0.100,
    "China (People's Republic of)": 0.250, "Colombia": 0.310, "Costa Rica": 0.300,
    "Côte d'Ivoire": 0.250, "Croatia": 0.180, "Curacao": 0.220, "Czech Republic": 0.190,
    "Democratic Republic of the Congo": 0.300, "Denmark": 0.220, "Dominica": 0.250,
    "Egypt": 0.225, "Estonia": 0.200, "Eswatini": 0.275, "Faeroe Islands": 0.180,
    "Finland": 0.200, "France": 0.284, "Gabon": 0.300, "Georgia": 0.150,
    "Germany": 0.299, "Greece": 0.240, "Greenland": 0.265, "Grenada": 0.280,
    "Guernsey": 0.000, "Hong Kong, China": 0.165, "Hungary": 0.090, "Iceland": 0.200,
    "India": 0.252, "Indonesia": 0.220, "Ireland": 0.125, "Isle of Man": 0.000,
    "Israel": 0.230, "Italy": 0.278, "Jamaica": 0.250, "Japan": 0.297,
    "Jersey": 0.000, "Kenya": 0.300, "Korea": 0.275, "Latvia": 0.200,
    "Liberia": 0.250, "Liechtenstein": 0.125, "Lithuania": 0.150, "Luxembourg": 0.249,
    "Macau, China": 0.120, "Malaysia": 0.240, "Maldives": 0.150, "Malta": 0.350,
    "Mauritius": 0.150, "Mexico": 0.300, "Monaco": 0.265, "Montserrat": 0.300,
    "Namibia": 0.320, "Netherlands": 0.250, "New Zealand": 0.280, "Nigeria": 0.300,
    "North Macedonia": 0.100, "Norway": 0.220, "Oman": 0.150, "Panama": 0.250,
    "Paraguay": 0.100, "Peru": 0.295, "Poland": 0.190, "Portugal": 0.315,
    "Romania": 0.160, "Russia": 0.200, "Saint Lucia": 0.300,
    "Saint Vincent and the Grenadines": 0.300, "San Marino": 0.170,
    "Saudi Arabia": 0.200, "Senegal": 0.300, "Serbia": 0.150, "Seychelles": 0.300,
    "Singapore": 0.170, "Slovak Republic": 0.210, "Slovenia": 0.190,
    "South Africa": 0.280, "Spain": 0.250, "Sweden": 0.206, "Switzerland": 0.197,
    "Thailand": 0.200, "Türkiye": 0.250, "Turks and Caicos Islands": 0.000,
    "United Arab Emirates": 0.000, "United Kingdom": 0.190, "Uruguay": 0.250,
    "Vietnam": 0.200,
}
COUNTRY_LIST = sorted(TAXES_BY_COUNTRY.keys())

# CAPEX distribution by EPC duration
# Rows = execution year (1st … 10th), columns = total EPC duration in years (1 … 10)
CAPEX_DISTRIBUTION = pd.DataFrame({
    "1":  [1.00000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000,  0.000000],
    "2":  [0.60000, 0.40000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000,  0.000000],
    "3":  [0.30000, 0.50000, 0.20000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000,  0.000000],
    "4":  [0.40000, 0.30000, 0.20000, 0.10000, 0.00000, 0.00000, 0.00000, 0.00000, 0.00000,  0.000000],
    "5":  [0.30000, 0.20000, 0.20000, 0.20000, 0.10000, 0.00000, 0.00000, 0.00000, 0.00000,  0.000000],
    "6":  [0.30000, 0.20000, 0.20000, 0.20000, 0.05000, 0.05000, 0.00000, 0.00000, 0.00000,  0.000000],
    "7":  [0.30000, 0.20000, 0.20000, 0.20000, 0.05000, 0.02500, 0.02500, 0.00000, 0.00000,  0.000000],
    "8":  [0.30000, 0.20000, 0.20000, 0.20000, 0.05000, 0.02500, 0.01250, 0.01250, 0.00000,  0.000000],
    "9":  [0.30000, 0.20000, 0.20000, 0.20000, 0.05000, 0.02500, 0.01250, 0.00625, 0.00625,  0.000000],
    "10": [0.30000, 0.20000, 0.20000, 0.20000, 0.05000, 0.02500, 0.01250, 0.00625, 0.003125, 0.003125],
}, index=["1st","2nd","3rd","4th","5th","6th","7th","8th","9th","10th"])
CAPEX_DISTRIBUTION.index.name = "Execution Year"

PLANT_COST_INDEX: dict[int, float] = {
    2000: 102.44, 2001: 102.32, 2002: 102.09, 2003: 106.35,
    2004: 119.36, 2005: 128.89, 2006: 135.32, 2007: 140.98,
    2008: 157.68, 2009: 138.86, 2010: 146.42, 2011: 156.66,
    2012: 155.24, 2013: 152.74, 2014: 155.32, 2015: 144.33,
    2016: 139.48, 2017: 145.48, 2018: 155.62, 2019: 156.33,
    2020: 150.65, 2021: 181.39, 2022: 214.60, 2023: 206.48,
    2024: 203.90,
}
PCI_YEARS = sorted(PLANT_COST_INDEX.keys())

# Rates-and-prices unit mapping: rate_unit → price_unit
RATE_TO_PRICE_UNIT: dict[str, str] = {
    "g/h":      "USD/g",
    "kg/h":     "USD/kg",
    "t/h":      "USD/t",
    "mL/h":     "USD/mL",
    "L/h":      "USD/L",
    "m³/h":     "USD/m³",
    "kW":       "USD/kWh",
    "MW":       "USD/MWh",
    "GW":       "USD/GWh",
    "MMBtu/h":  "USD/MMBtu",
    "g/y":      "USD/g",
    "kg/y":     "USD/kg",
    "t/y":      "USD/t",
    "mL/y":     "USD/mL",
    "L/y":      "USD/L",
    "m³/y":     "USD/m³",
    "MMBtu/y":  "USD/MMBtu",
}

# Default session-state values — single source of truth
DEFAULTS = {
    "sn_input":        "",
    "mp_input":        "",
    "pu_input":        "kg",
    "pc_input":        None,
    "eq_cost_src":     "Manual Input",
    "oth_cost_src":    "Manual Input",
    "lang_utility":    False,
    "dm_prod_type":    "Basic Chemical",
    "dm_trl":          "Industrial (8 or 9)",
    "dm_info_avail":   "Medium",
    "dm_severity":     "Medium",
    "dm_mat_type":     "Fluids",
    "dm_plant_size":   "Medium",
    # CAPEX
    "equip_acq":         0.0,
    "spare_parts":       0.0,
    "equipment_setting": 0.0,
    "piping":            0.0,
    "civil":             0.0,
    "steel":             0.0,
    "instrumentals":     0.0,
    "electrical":        0.0,
    "insulation":        0.0,
    "paint":             0.0,
    "isbl_contrib":      100.0,
    "field_office_staff":     0.0,
    "construction_indirects": 0.0,
    "freight":           0.0,
    "taxes_and_permits": 0.0,
    "engineering_and_ho":0.0,
    "ga_overheads":      0.0,
    "contract_fee":      0.0,
    # Lang override support
    "allow_override":       False,   # master switch for manual editing of Lang fields
    "lang_seeded_acq":      None,    # equip_acq value used for last Lang seed
    # Aspen PEA import tracking
    "pea_last_fingerprint": "",
    "pea_last_parsed":      None,
    "pea_fmt":              "",
    # CAPEX calculations
    "databank_year":   2022,
    "analysis_year":   PCI_YEARS[-1],
    "proj_location":   "Brazil",
    "location_factor": 0.97,
    # Additional information
    "working_hours":   8000.0,
    "scaling_factor":  0.6,
    # Fixed costs — Labor
    "n_operators":          2,
    "operator_salary":      1247.75,
    "n_supervisors":        1,
    "supervisor_salary":    1660.155,
    "salary_charges":       2.2,
    "plant_daily_hours":    24.0,
    "weekly_op_days":       7.0,
    "worker_hours_shift":   8.0,
    "worker_shifts_week":   5.0,
    "worker_vacation_weeks":4.0,
    # lab/office overrides (None = use table value)
    "lab_charges_override":    None,
    "office_labor_override":   None,
    "labor_working_hrs_override": None,
    # Fixed costs — Supply & maintenance overrides
    "maint_repair_override":   None,
    "op_supplies_override":    None,
    # Fixed costs — Additional fixed overrides
    "admin_overhead_override": None,
    "mfg_overhead_override":   None,
    "taxes_ins_override":      None,
    "patents_roy_override":    None,
    # Fixed costs — Indirect fixed overrides
    "admin_costs_override":    None,
    "mfg_costs_override":      None,
    "dist_selling_override":   None,
    "r_and_d_override":        None,
    # Working capital
    "wc_method":               "Percentage",
    "wc_pct":                  5.0,
    "wc_equiv_cash_days":      30.0,
    "wc_raw_mat_days":         15.0,
    "wc_accounts_rec_days":    30.0,
    "wc_accrued_payroll_days": 30.0,
    "wc_accounts_pay_days":    30.0,
    # Startup costs
    "startup_method":          "Multiple Factors",
    "startup_single_pct":      8.0,
    "startup_op_training_days":150.0,
    "startup_commerc_pct":     5.0,
    "startup_inefficiency_pct":4.0,
    # Total investment summary
    "additional_costs":        0.0,
    # TIC accuracy overrides
    "tic_lower_override":      None,
    "tic_upper_override":      None,
    # Financial assumptions — Land
    "land_option":             "Buy",
    "land_buy_pct_override":   None,
    "land_rent_pct_override":  None,
    # Financial assumptions — Depreciation
    "depreciation_method":     "Straight Line",
    "depreciation_years":      10,
    "residual_value_pct":      20.0,
    # Financial assumptions — Taxes
    "tax_country":             "Brazil",
    "tax_rate_override":       None,
    # Financial assumptions — Financial leverage
    "financing_type":          "None",
    "debt_ratio_pct":          50.0,
    "amortization_years":      13,
    "grace_period_years":      5,
    # Financial assumptions — MARR
    "central_bank_rate":       5.45,
    "credit_spread":           2.94,
    "unlevered_beta":          1.0,
    "market_return":           8.63,
    "risk_free_rate":          1.94,
    "country_risk_premium":    3.63,
    "us_cpi":                  2.46,
    "country_cpi":             4.65,
    "marr_override":           None,
    # Project lifetime and CAPEX distribution
    "epc_years":               3,
    "project_lifetime":        20,
    # Percentage of total capacity
    "capacity_first_year":     100.0,
    "capacity_intermediate":   100.0,
    "capacity_last_year":      100.0,
    # Percentage of fixed costs per year
    "fixed_costs_first_year":  100.0,
    "fixed_costs_intermediate":100.0,
    "fixed_costs_last_year":   100.0,
    # Market assumptions — annual growth rates
    "growth_main_price":       0.0,
    "growth_byproduct_price":  0.0,
    "growth_raw_materials":    0.0,
    "growth_chem_utilities":   0.0,
    "growth_fixed_costs":      0.0,
    # Other premises
    "main_product_price":      None,
}

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def fmt_curr(val: float) -> str:
    return f"${val:,.2f}"


def _result_row(label: str, value: float, key: str, formula: str = ""):
    """Render a labelled result row. If formula is provided, appends a ? tooltip."""
    c1, c2 = st.columns([3, 2])
    with c1:
        if formula:
            st.markdown(
                f'**{label}** <span title="{formula}" '
                f'style="cursor:help;color:#888;font-size:0.85rem;">ⓘ</span>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(f"**{label}**")
    with c2:
        st.text_input(key, value=fmt_curr(value), disabled=True,
                      label_visibility="collapsed")


def parse_aspen_pea(ipewb_bytes: bytes | None, reports_bytes: bytes | None,
                    import_equip: bool, import_other: bool) -> dict:
    """
    Parse IPEWB and/or Reports Excel files from Aspen PEA.
    Auto-detects legacy format (Equipment sheet) vs. new format (Equipment Summary sheet).
    Returns a dict of session-state keys → values for all fields that could be extracted.
    Returns an empty dict if parsing fails.
    """
    result = {}

    # ── Parse IPEWB ───────────────────────────────────────────────────────
    if ipewb_bytes and import_equip:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(ipewb_bytes), data_only=True)
            sheets = wb.sheetnames

            # Auto-detect format
            if "Equipment Summary" in sheets:
                fmt = "new"
            elif "Equipment" in sheets:
                fmt = "legacy"
            else:
                fmt = None

            if fmt == "new":
                # Equipment acquisition: sum col H of Equipment Summary
                ws_eq = wb["Equipment Summary"]
                equip_acq = sum(
                    r[7] for r in ws_eq.iter_rows(min_row=1, values_only=True)
                    if isinstance(r[7], (int, float))
                )
                result["equip_acq"] = float(equip_acq)

                ws_ps = wb["Project Summary"]
                ops = ws_ps.cell(row=146, column=3).value
                sups = ws_ps.cell(row=155, column=3).value
                epc_weeks = ws_ps.cell(row=49, column=3).value
                if epc_weeks:
                    result["epc_years"] = max(1, round(epc_weeks * 7 / 365))
                if import_other:
                    eq_set = ws_ps.cell(row=102, column=3).value
                    if isinstance(eq_set, (int, float)):
                        result["equipment_setting"] = float(eq_set)

            elif fmt == "legacy":
                # Equipment acquisition: sum col E of Equipment
                ws_eq = wb["Equipment"]
                equip_acq = sum(
                    r[4] for r in ws_eq.iter_rows(min_row=1, values_only=True)
                    if isinstance(r[4], (int, float))
                )
                result["equip_acq"] = float(equip_acq)

                ws_ps = wb["Project Summary"]
                ops  = ws_ps.cell(row=225, column=3).value
                sups = ws_ps.cell(row=234, column=3).value
                epc_weeks = ws_ps.cell(row=62, column=3).value
                if epc_weeks:
                    result["epc_years"] = max(1, round(epc_weeks * 7 / 365))
                if import_other:
                    eq_set = ws_ps.cell(row=173, column=3).value
                    if isinstance(eq_set, (int, float)):
                        result["equipment_setting"] = float(eq_set)

            if fmt in ("new", "legacy"):
                if isinstance(ops, (int, float)):
                    result["n_operators"] = max(1, int(ops))
                if isinstance(sups, (int, float)):
                    result["n_supervisors"] = max(1, int(sups))

        except Exception as e:
            result["_ipewb_error"] = str(e)

    # ── Parse Reports ─────────────────────────────────────────────────────
    if reports_bytes and import_other:
        try:
            wb_r = openpyxl.load_workbook(io.BytesIO(reports_bytes), data_only=True)
            ws = wb_r["Proj Cost Sumry"]

            def _v(row, col):
                v = ws.cell(row=row, column=col).value
                return float(v) if isinstance(v, (int, float)) else 0.0

            # Spare parts = Reports E8 − IPEWB equipment acquisition
            reports_total_equip = _v(8, 5)  # col E, row 8
            ipewb_equip_acq = result.get("equip_acq", 0.0)
            result["spare_parts"] = max(0.0, reports_total_equip - ipewb_equip_acq)

            # Installation costs — col F
            result["piping"]           = _v(9,  6)
            result["civil"]            = _v(10, 6)
            result["steel"]            = _v(11, 6)
            result["instrumentals"]    = _v(12, 6)
            result["electrical"]       = _v(13, 6)
            result["insulation"]       = _v(14, 6)
            result["paint"]            = _v(15, 6)

            # Indirect field costs — col M
            # Keys must match lang_or_manual ss_key = field.lower().replace(" ", "_")
            result["field_office_staff"]       = _v(7,  13)
            result["construction_indirects"]   = _v(18, 13)

            # Non-field costs — col R
            result["freight"]              = _v(6,  18)
            result["taxes_and_permits"]    = _v(10, 18)
            result["engineering_and_ho"]   = _v(16, 18) + _v(19, 18)
            result["ga_overheads"]         = _v(20, 18)
            result["contract_fee"]         = _v(21, 18)

        except Exception as e:
            result["_reports_error"] = str(e)

    return result


def reset_state(keys: dict = DEFAULTS):
    """Write every key in *keys* back to session_state, and wipe Lang field overrides."""
    for k, v in keys.items():
        st.session_state[k] = v
    for k in list(st.session_state.keys()):
        if k.startswith("val_") or k.startswith("w_"):
            del st.session_state[k]
    st.session_state.table_key = st.session_state.get("table_key", 0) + 1


def _reseed_lang_fields(equip_acq: float):
    """Overwrite all lf_ keys with freshly computed Lang values for equip_acq."""
    idx = 0 if st.session_state.lang_utility else 1
    for field, factors in LANG_FACTORS.items():
        key = f"lf_{field.lower().replace(' ', '_')}"
        st.session_state[key] = equip_acq * factors[idx]
    st.session_state.lang_seeded_acq = equip_acq


def lang_val(field: str, equip_acq: float) -> float:
    """Return the Lang-factored value for *field* given Equipment Acquisition cost."""
    idx = 0 if st.session_state.lang_utility else 1
    return equip_acq * LANG_FACTORS[field][idx]


def get_pci(year: int) -> float | None:
    return PLANT_COST_INDEX.get(year)


def pci_escalate(base_cost: float, base_year: int, target_year: int) -> float | None:
    pci_base   = PLANT_COST_INDEX.get(base_year)
    pci_target = PLANT_COST_INDEX.get(target_year)
    if pci_base and pci_target:
        return base_cost * (pci_target / pci_base)
    return None


def price_unit_for(rate_unit: str) -> str:
    return RATE_TO_PRICE_UNIT.get(rate_unit, "")


LANG_FIELDS = list(LANG_FACTORS.keys())  # canonical field names in order

def _seed_override_values():
    """
    Called when the user ticks 'Allow manual override'.
    Reads the current equip_acq and writes a val_ key for every
    Lang field.  These are plain state variables — NOT widget keys —
    so Streamlit never overwrites them on widget init.
    """
    acq = st.session_state.get("equip_acq", 0.0)
    idx = 0 if st.session_state.get("lang_utility", False) else 1
    for field, factors in LANG_FACTORS.items():
        vkey = f"val_{field.lower().replace(' ', '_')}"
        st.session_state[vkey] = acq * factors[idx]


def lang_or_manual(field: str, label: str, equip_acq: float, step: float = 100.0) -> float:
    """
    Lang Factors mode — override OFF:
        Read-only text_input showing the live computed Lang value.

    Lang Factors mode — override ON:
        Editable number_input whose *initial* value was seeded by
        _seed_override_values() via the checkbox callback.
        Stored in val_<field> (plain state, not a widget key), read
        back via a widget key w_<field> so Streamlit never zeros it.
        Label is green (✓ Lang) or amber (⚠ overridden).

    Manual mode:
        Plain number_input bound to session_state via ss_key.
    """
    ss_key  = field.lower().replace(" ", "_")
    val_key = f"val_{ss_key}"   # plain storage key — written by callback
    w_key   = f"w_{ss_key}"     # widget key — only used to read user input

    if st.session_state.oth_cost_src == "Lang Factors":
        lang_computed = lang_val(field, equip_acq)
        allow = st.session_state.get("allow_override", False)

        if not allow:
            st.text_input(f"{label}", value=fmt_curr(lang_computed), disabled=True)
            return lang_computed

        # Override ON — use val_key as the source of truth for the current value.
        # val_key was already populated by _seed_override_values() callback.
        stored = st.session_state.get(val_key, lang_computed)
        is_overridden = abs(stored - lang_computed) > 0.005

        color = "#e67e00" if is_overridden else "#2e7d32"
        hint  = " ⚠ overridden" if is_overridden else " ✓ Lang"
        st.markdown(
            f'<p style="margin-bottom:0px; font-size:0.85rem; color:{color};">'
            f'<b>{label} ($)</b>{hint}</p>',
            unsafe_allow_html=True,
        )

        # Render number_input with value= drawn from val_key.
        # on_change writes the widget's new value back into val_key.
        def _sync(vk=val_key, wk=w_key):
            st.session_state[vk] = st.session_state[wk]

        result = st.number_input(
            label, min_value=0.0, step=step,
            value=stored,
            key=w_key,
            label_visibility="collapsed",
            on_change=_sync,
        )
        return result

    # Plain manual mode
    return st.number_input(f"{label} ($)", min_value=0.0, step=step, key=ss_key)


# ─────────────────────────────────────────────
# SESSION STATE INIT
# ─────────────────────────────────────────────
if "scenarios" not in st.session_state:
    st.session_state.scenarios = {}
if "table_key" not in st.session_state:
    st.session_state.table_key = 0

for key, default in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = default

# Deferred clear (must run before widgets are drawn)
if st.session_state.pop("clear_on_next_run", False):
    reset_state()

if st.session_state.get("success_msg"):
    st.success(st.session_state.pop("success_msg"))


# ─────────────────────────────────────────────
# CALLBACK
# ─────────────────────────────────────────────
def load_scenario_data():
    sn = st.session_state.sn_input
    data = st.session_state.scenarios.get(sn, {})

    mapping = {
        "mp_input":        ("Product Name",            ""),
        "pu_input":        ("Unit",                    "kg"),
        "pc_input":        ("Capacity",                None),
        "eq_cost_src":     ("Equipment Costs Source",  "Manual Input"),
        "oth_cost_src":    ("Other Costs Source",      "Manual Input"),
        "lang_utility":    ("Contains Utility Systems",False),
        "dm_prod_type":    ("Product Type",            "Basic Chemical"),
        "dm_trl":          ("TRL",                     "Industrial (8 or 9)"),
        "dm_info_avail":   ("Info Availability",       "Medium"),
        "dm_severity":     ("Process Severity",        "Medium"),
        "dm_mat_type":     ("Material Handled",        "Fluids"),
        "dm_plant_size":   ("Plant Size",              "Medium"),
        "equip_acq":       ("Equipment Acquisition",   0.0),
        "spare_parts":     ("Spare Parts",             0.0),
        "equipment_setting": ("Equipment Setting",     0.0),
        "piping":          ("Piping",                  0.0),
        "civil":           ("Civil",                   0.0),
        "steel":           ("Steel",                   0.0),
        "instrumentals":   ("Instrumentals",           0.0),
        "electrical":      ("Electrical",              0.0),
        "insulation":      ("Insulation",              0.0),
        "paint":           ("Paint",                   0.0),
        "isbl_contrib":    ("ISBL Contribution (%)",   100.0),
        "field_office_staff":      ("Field Office Staff",      0.0),
        "construction_indirects":  ("Construction Indirects",  0.0),
        "freight":         ("Freight",                 0.0),
        "taxes_and_permits": ("Taxes and Permits",     0.0),
        "engineering_and_ho": ("Engineering and HO",  0.0),
        "ga_overheads":    ("GA Overheads",            0.0),
        "contract_fee":    ("Contract Fee",            0.0),
        "allow_override":  ("Allow Override",          False),
        "databank_year":   ("Databank Year",           2022),
        "analysis_year":   ("Year of Analysis",        PCI_YEARS[-1]),
        "proj_location":   ("Project Location",        "Brazil"),
        "location_factor": ("Location Factor",         0.97),
        "working_hours":   ("Working Hours per Year",  8000.0),
        "scaling_factor":  ("Scaling Factor",          0.6),
        # Fixed costs — Labor
        "n_operators":               ("Num Operators",            2),
        "operator_salary":           ("Operator Salary",          1247.75),
        "n_supervisors":             ("Num Supervisors",          1),
        "supervisor_salary":         ("Supervisor Salary",        1660.155),
        "salary_charges":            ("Salary Charges",           2.2),
        "plant_daily_hours":         ("Plant Daily Hours",        24.0),
        "weekly_op_days":            ("Weekly Op Days",           7.0),
        "worker_hours_shift":        ("Worker Hours per Shift",   8.0),
        "worker_shifts_week":        ("Worker Shifts per Week",   5.0),
        "worker_vacation_weeks":     ("Worker Vacation Weeks",    4.0),
        "lab_charges_override":      ("Lab Charges Override",     None),
        "office_labor_override":     ("Office Labor Override",    None),
        "labor_working_hrs_override":("Labor Working Hrs Override", None),
        "maint_repair_override":     ("Maint Repair Override",    None),
        "op_supplies_override":      ("Op Supplies Override",     None),
        "admin_overhead_override":   ("Admin Overhead Override",  None),
        "mfg_overhead_override":     ("Mfg Overhead Override",    None),
        "taxes_ins_override":        ("Taxes Ins Override",       None),
        "patents_roy_override":      ("Patents Roy Override",     None),
        "admin_costs_override":      ("Admin Costs Override",     None),
        "mfg_costs_override":        ("Mfg Costs Override",       None),
        "dist_selling_override":     ("Dist Selling Override",    None),
        "r_and_d_override":          ("R And D Override",         None),
        # Working capital
        "wc_method":                 ("WC Method",                "Percentage"),
        "wc_pct":                    ("WC Pct",                   5.0),
        "wc_equiv_cash_days":        ("WC Equiv Cash Days",       30.0),
        "wc_raw_mat_days":           ("WC Raw Mat Days",          15.0),
        "wc_accounts_rec_days":      ("WC Accounts Rec Days",     30.0),
        "wc_accrued_payroll_days":   ("WC Accrued Payroll Days",  30.0),
        "wc_accounts_pay_days":      ("WC Accounts Pay Days",     30.0),
        # Startup costs
        "startup_method":            ("Startup Method",           "Multiple Factors"),
        "startup_single_pct":        ("Startup Single Pct",       8.0),
        "startup_op_training_days":  ("Startup Op Training Days", 150.0),
        "startup_commerc_pct":       ("Startup Commerc Pct",      5.0),
        "startup_inefficiency_pct":  ("Startup Inefficiency Pct", 4.0),
        # Total investment
        "additional_costs":          ("Additional Costs",         0.0),
        # TIC overrides
        "tic_lower_override":        ("TIC Lower Override",       None),
        "tic_upper_override":        ("TIC Upper Override",       None),
        # Financial assumptions — Land
        "land_option":               ("Land Option",              "Buy"),
        "land_buy_pct_override":     ("Land Buy Pct Override",    None),
        "land_rent_pct_override":    ("Land Rent Pct Override",   None),
        # Financial assumptions — Depreciation
        "depreciation_method":       ("Depreciation Method",      "Straight Line"),
        "depreciation_years":        ("Depreciation Years",       10),
        "residual_value_pct":        ("Residual Value Pct",       20.0),
        # Financial assumptions — Taxes
        "tax_country":               ("Tax Country",              "Brazil"),
        "tax_rate_override":         ("Tax Rate Override",        None),
        # Financial leverage
        "financing_type":            ("Financing Type",           "None"),
        "debt_ratio_pct":            ("Debt Ratio Pct",           50.0),
        "amortization_years":        ("Amortization Years",       13),
        "grace_period_years":        ("Grace Period Years",       5),
        # MARR
        "central_bank_rate":         ("Central Bank Rate",        5.45),
        "credit_spread":             ("Credit Spread",            2.94),
        "unlevered_beta":            ("Unlevered Beta",           1.0),
        "market_return":             ("Market Return",            8.63),
        "risk_free_rate":            ("Risk Free Rate",           1.94),
        "country_risk_premium":      ("Country Risk Premium",     3.63),
        "us_cpi":                    ("US CPI",                   2.46),
        "country_cpi":               ("Country CPI",              4.65),
        "marr_override":             ("MARR Override",            None),
        # Project lifetime
        "epc_years":                 ("EPC Years",                3),
        "project_lifetime":          ("Project Lifetime",         20),
        # Capacity
        "capacity_first_year":       ("Capacity First Year",      100.0),
        "capacity_intermediate":     ("Capacity Intermediate",    100.0),
        "capacity_last_year":        ("Capacity Last Year",       100.0),
        # Fixed costs
        "fixed_costs_first_year":    ("Fixed Costs First Year",   100.0),
        "fixed_costs_intermediate":  ("Fixed Costs Intermediate", 100.0),
        "fixed_costs_last_year":     ("Fixed Costs Last Year",    100.0),
        # Market assumptions
        "growth_main_price":         ("Growth Main Price",        0.0),
        "growth_byproduct_price":    ("Growth Byproduct Price",   0.0),
        "growth_raw_materials":      ("Growth Raw Materials",     0.0),
        "growth_chem_utilities":     ("Growth Chem Utilities",    0.0),
        "growth_fixed_costs":        ("Growth Fixed Costs",       0.0),
        # Other premises — always reset to None so user sets fresh price per scenario
        "main_product_price":        ("_reset_", None),
    }

    for ss_key, (data_key, default) in mapping.items():
        if data_key == "_reset_":
            st.session_state[ss_key] = None   # always reset, never load from saved data
        else:
            st.session_state[ss_key] = data.get(data_key, default)

    st.session_state.table_key += 1
    st.session_state.lang_seeded_acq = None  # force re-seed on next render


# ─────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────
st.title("Scenario Configuration & Inputs")
st.markdown("Define the parameters for your scenario. "
            "Type an existing Scenario Name to load it, or type a new name for a blank slate.")

# 1. Scenario Name
st.header("1. Scenario Name")
scenario_name = st.text_input("Scenario Name", key="sn_input", on_change=load_scenario_data)
st.divider()

# 2. Basic Information
st.header("2. Basic Information")
col1, col2 = st.columns(2)
with col1:
    st.text_input("Main Product Name", key="mp_input")
with col2:
    st.selectbox("Main product unit", ALLOWED_UNITS, key="pu_input")
    st.number_input(f"Main product capacity ({st.session_state.pu_input}/year)",
                    min_value=0.0, step=1.0, key="pc_input")
st.divider()

# 3. Process Variables
st.header("3. Process Variables")
saved_vars = st.session_state.scenarios.get(scenario_name, {}).get("Process Variables", [])
df_vars = pd.DataFrame(saved_vars) if saved_vars else pd.DataFrame(columns=["Variable Name", "Unit", "Value"])

edited_process_vars = st.data_editor(
    df_vars, key=f"editor_{st.session_state.table_key}", num_rows="dynamic",
    use_container_width=True, hide_index=True,
    column_config={
        "Variable Name": st.column_config.TextColumn(required=True),
        "Unit":          st.column_config.TextColumn(required=True),
        "Value":         st.column_config.NumberColumn(required=True),
    },
)
st.divider()

# 4. Investment Costs Sources
st.header("4. Investment Costs Sources")
col_eq, col_oth = st.columns(2)
with col_eq:
    st.selectbox("Equipment costs source", ["Manual Input", "Aspen PEA"], key="eq_cost_src")
with col_oth:
    st.selectbox("Other Costs source", ["Manual Input", "Aspen PEA", "Lang Factors"], key="oth_cost_src")
    if st.session_state.oth_cost_src == "Lang Factors":
        st.checkbox("Contains utility systems?", key="lang_utility")

_eq_is_pea  = st.session_state.eq_cost_src  == "Aspen PEA"
_oth_is_pea = st.session_state.oth_cost_src == "Aspen PEA"

if _eq_is_pea or _oth_is_pea:
    st.markdown("---")
    st.markdown("##### Aspen PEA File Import")
    st.caption(
        "Upload the IPEWB (.xlsx) and/or Reports (.xlsm) files for this scenario. "
        "The tool auto-detects the Aspen PEA version. "
        "Imported values become the defaults for this scenario and can still be edited below."
    )
    col_up1, col_up2 = st.columns(2)
    with col_up1:
        ipewb_file = st.file_uploader(
            "IPEWB file (.xlsx)", type=["xlsx"],
            key="ipewb_uploader",
            help="Required when Equipment costs source = Aspen PEA.",
        ) if _eq_is_pea else None
    with col_up2:
        reports_file = st.file_uploader(
            "Reports file (.xlsm / .xlsx)", type=["xlsm", "xlsx"],
            key="reports_uploader",
            help="Required when Other Costs source = Aspen PEA.",
        ) if _oth_is_pea else None

    # Build a fingerprint of the currently uploaded files (name + size)
    # Only re-parse when this changes — prevents infinite rerun loops
    _ipewb_fp   = f"{ipewb_file.name}:{ipewb_file.size}"   if ipewb_file   else ""
    _reports_fp = f"{reports_file.name}:{reports_file.size}" if reports_file else ""
    _current_fp = f"{_ipewb_fp}|{_reports_fp}"
    _last_fp    = st.session_state.get("pea_last_fingerprint", "")

    if _current_fp != _last_fp and (_ipewb_fp or _reports_fp):
        _ipewb_bytes   = ipewb_file.read()   if ipewb_file   else None
        _reports_bytes = reports_file.read() if reports_file else None

        with st.spinner("Reading Aspen PEA files…"):
            _parsed = parse_aspen_pea(
                _ipewb_bytes, _reports_bytes,
                import_equip=_eq_is_pea,
                import_other=_oth_is_pea,
            )

        _errs = {k: v for k, v in _parsed.items() if k.startswith("_") and k.endswith("_error")}
        if _errs:
            for e in _errs.values():
                st.error(f"Import error: {e}")
        else:
            # Write values into session state — widgets below will pick them up this run
            _skip = {"fmt", "_ipewb_error", "_reports_error"}
            for k, v in _parsed.items():
                if k not in _skip:
                    st.session_state[k] = v
            # Store fingerprint so we don't reparse on next natural rerun
            st.session_state["pea_last_fingerprint"] = _current_fp
            # Store summary for display
            st.session_state["pea_last_parsed"] = {k: v for k, v in _parsed.items() if k not in _skip}
            st.session_state["pea_fmt"] = _parsed.get("fmt", "unknown")

    # Always show the summary table if we have parsed results for this upload
    if st.session_state.get("pea_last_parsed") and _current_fp == st.session_state.get("pea_last_fingerprint", ""):
        _p = st.session_state["pea_last_parsed"]
        _fmt_label = {"legacy": "Legacy (Equipment sheet)", "new": "New (Equipment Summary sheet)"}.get(
            st.session_state.get("pea_fmt", ""), "Unknown")

        st.success(f"✓ Aspen PEA import complete — format detected: **{_fmt_label}**")

        _label_map = {
            "equip_acq":             "Equipment Acquisition ($)",
            "equipment_setting":     "Equipment Setting ($)",
            "spare_parts":           "Spare Parts ($)",
            "n_operators":           "Operators per shift",
            "n_supervisors":         "Supervisors per shift",
            "epc_years":             "EPC time (years)",
            "piping":                "Piping ($)",
            "civil":                 "Civil ($)",
            "steel":                 "Steel ($)",
            "instrumentals":         "Instrumentals ($)",
            "electrical":            "Electrical ($)",
            "insulation":            "Insulation ($)",
            "paint":                 "Paint ($)",
            "field_office_staff":    "Field Office Staff ($)",
            "construction_indirects":"Construction Indirects ($)",
            "freight":               "Freight ($)",
            "taxes_and_permits":     "Taxes and Permits ($)",
            "engineering_and_ho":    "Engineering and HO ($)",
            "ga_overheads":          "GA Overheads ($)",
            "contract_fee":          "Contract Fee ($)",
        }
        # Keys treated as integer counts, not monetary
        _int_keys = {"n_operators", "n_supervisors", "epc_years"}
        _rows = []
        for k, v in _p.items():
            lbl = _label_map.get(k, k)
            if k in _int_keys:
                disp = str(int(v))
            elif isinstance(v, float):
                disp = f"${v:,.2f}"
            else:
                disp = str(v)
            _rows.append({"Field": lbl, "Imported Value": disp})
        st.dataframe(pd.DataFrame(_rows), use_container_width=False, hide_index=True)

st.divider()

# 5. Decision Making Assistant
st.header("5. Decision Making Assistant")

# Override keys derived from DMA selections — cleared when any DMA choice changes
# so that _overridable_number re-seeds from the new reference value
_DMA_DERIVED_OVERRIDES = [
    "op_supplies_override", "maint_repair_override",
    "admin_overhead_override", "mfg_overhead_override", "taxes_ins_override",
    "patents_roy_override", "lab_charges_override", "office_labor_override",
    "admin_costs_override", "mfg_costs_override",
    "dist_selling_override", "r_and_d_override",
    "tic_lower_override", "tic_upper_override",
]

def _on_dma_change():
    """Clear all reference-table override values and their widget keys when DMA changes."""
    for ok in _DMA_DERIVED_OVERRIDES:
        st.session_state[ok] = None
        st.session_state.pop(f"w_{ok}", None)

col_dm1, col_dm2, col_dm3 = st.columns(3)
with col_dm1:
    st.selectbox("Type of main product",        PRODUCT_TYPES,    key="dm_prod_type", on_change=_on_dma_change)
    st.selectbox("Availability of information", INFO_OPTIONS,     key="dm_info_avail", on_change=_on_dma_change)
with col_dm2:
    st.selectbox("TRL",                         TRL_OPTIONS,      key="dm_trl",       on_change=_on_dma_change)
    st.selectbox("Process severity",            SEVERITY_OPTIONS, key="dm_severity",  on_change=_on_dma_change)
with col_dm3:
    st.selectbox("Type of material handled",    MAT_OPTIONS,      key="dm_mat_type",  on_change=_on_dma_change)
    st.selectbox("Plant size",                  SIZE_OPTIONS,     key="dm_plant_size")
st.divider()

# 7. Project CAPEX
st.header("7. Project CAPEX")

is_lang = st.session_state.oth_cost_src == "Lang Factors"

if is_lang:
    st.checkbox(
        "Allow manual override of Lang factor fields?",
        key="allow_override",
        on_change=_seed_override_values,
        help="When checked, all Lang-computed fields become editable. "
             "Fields that differ from their Lang value are highlighted in amber.",
    )
else:
    st.session_state.allow_override = False

# 7.1 Equipment Costs
st.subheader("Equipment Costs")
col1, col2, col3 = st.columns(3)
with col1:
    equip_acq = st.number_input("Equipment Acquisition ($)", min_value=0.0, step=1000.0, key="equip_acq")
with col2:
    spare_parts = lang_or_manual("Spare Parts", "Spare Parts", equip_acq)
with col3:
    equip_setting = lang_or_manual("Equipment Setting", "Equipment Setting", equip_acq)

base_equip_costs  = equip_acq + spare_parts + equip_setting
unsched_pct       = UNSCHED_FACTORS.get((st.session_state.dm_trl, st.session_state.dm_info_avail), 0.0)
total_equip_costs = base_equip_costs * (1 + unsched_pct)

st.markdown("##### Total Equipment Assembly")
col4, col5, col6 = st.columns(3)
with col4:
    st.text_input("Base Equipment Costs",      value=fmt_curr(base_equip_costs),      disabled=True)
with col5:
    st.text_input("Unscheduled Equipment (%)", value=f"{unsched_pct * 100:.2f}%",     disabled=True)
with col6:
    st.text_input("Total Equipment Costs",     value=fmt_curr(total_equip_costs),     disabled=True)
st.markdown("---")

# 7.2 Installation Costs
st.subheader("Installation Costs")
inst_fields = ["Piping", "Civil", "Steel", "Instrumentals", "Electrical", "Insulation"]
inst_cols   = st.columns(3)
inst_values = []
for i, field in enumerate(inst_fields):
    with inst_cols[i % 3]:
        inst_values.append(lang_or_manual(field, field, equip_acq))

col_p, col_t = st.columns(2)
with col_p:
    paint = lang_or_manual("Paint", "Paint", equip_acq)
total_inst_costs = sum(inst_values) + paint
with col_t:
    st.text_input("Total Installation Costs", value=fmt_curr(total_inst_costs), disabled=True)
st.markdown("---")

# 7.3 Direct Field Costs
st.subheader("Direct Field Costs")
total_direct_field_costs = total_equip_costs + total_inst_costs
col1, col2, col3 = st.columns(3)
with col1:
    st.text_input("Total Direct Field Costs", value=fmt_curr(total_direct_field_costs), disabled=True)
with col2:
    isbl_contrib = st.number_input("ISBL Contribution (%)", min_value=0.0, max_value=100.0,
                                   step=1.0, key="isbl_contrib")
with col3:
    st.text_input("OSBL Contribution (%)", value=f"{100.0 - isbl_contrib:.2f}%", disabled=True)
st.markdown("---")

# 7.4 Indirect Field Costs
st.subheader("Indirect Field Costs")
col1, col2, col3 = st.columns(3)
with col1:
    field_office = lang_or_manual("Field Office Staff", "Field Office Staff", equip_acq)
with col2:
    const_indirects = lang_or_manual("Construction Indirects", "Construction Indirects", equip_acq)
total_indirect_field_costs = field_office + const_indirects
with col3:
    st.text_input("Total Indirect Field Costs", value=fmt_curr(total_indirect_field_costs), disabled=True)
st.markdown("---")

# 7.5 Non-Field Costs
st.subheader("Non-Field Costs")
nf_fields = ["Freight", "Taxes and Permits", "Engineering and HO", "GA Overheads", "Contract Fee"]
nf_cols   = st.columns(3)
nf_values = []
for i, field in enumerate(nf_fields):
    with nf_cols[i % 3]:
        nf_values.append(lang_or_manual(field, field, equip_acq))

total_non_field_costs = sum(nf_values)
# Fill remaining cell in last row with the total
with nf_cols[(len(nf_fields)) % 3]:
    st.text_input("Total Non-Field Costs", value=fmt_curr(total_non_field_costs), disabled=True)
st.markdown("---")

# 7.6 CAPEX Calculations
st.subheader("Capex Calculations")
project_costs_isbl_osbl = total_direct_field_costs + total_indirect_field_costs + total_non_field_costs
st.metric("Project Costs for ISBL + OSBL", fmt_curr(project_costs_isbl_osbl))

st.markdown("---")
col_cx1, col_cx2, col_cx3 = st.columns(3)

# --- Project Contingency (auto-looked-up) ---
with col_cx1:
    contingency_pct = PROJECT_CONTINGENCY.get(
        (st.session_state.dm_trl, st.session_state.dm_severity), 0.0
    )
    st.text_input(
        "Project Contingency (auto)",
        value=f"{contingency_pct * 100:.1f}%  "
              f"[TRL: {st.session_state.dm_trl} / Severity: {st.session_state.dm_severity}]",
        disabled=True,
        help="Looked up from the Project Contingency reference table using TRL and Process Severity."
    )

# --- Databank Year ---
with col_cx2:
    databank_year_input = st.number_input(
        "Databank Year", min_value=PCI_YEARS[0], max_value=PCI_YEARS[-1],
        step=1, key="databank_year",
        help="Reference year for the cost data (defaults to 2022)."
    )
    pci_databank = PLANT_COST_INDEX.get(int(databank_year_input))
    st.text_input(
        "PCI (Databank Year)",
        value=f"{pci_databank:.2f}" if pci_databank else "—",
        disabled=True,
    )

# --- Year of Analysis ---
with col_cx3:
    analysis_year_input = st.number_input(
        "Year of Analysis", min_value=1900, max_value=2100,
        step=1, key="analysis_year",
        help=f"Target year for cost update. Latest available index year: {PCI_YEARS[-1]}."
    )
    pci_analysis = PLANT_COST_INDEX.get(int(analysis_year_input))
    if pci_analysis is None:
        # Year not in database — warn and fall back to most recent year
        st.warning(
            "Index not yet available in the database. "
            f"Value will default to most recent year ({PCI_YEARS[-1]})."
        )
        pci_analysis = PLANT_COST_INDEX[PCI_YEARS[-1]]
        effective_analysis_year = PCI_YEARS[-1]
    else:
        effective_analysis_year = int(analysis_year_input)
    st.text_input("PCI (Year of Analysis)", value=f"{pci_analysis:.2f}", disabled=True)

# --- Time Update Factor + Location ---
st.markdown("---")
col_cx4, col_cx5, col_cx6, col_cx7 = st.columns(4)

time_update_factor = pci_analysis / pci_databank if pci_databank else 0.0

with col_cx4:
    st.text_input(
        "Time Update Factor",
        value=f"{time_update_factor:.4f}",
        disabled=True,
        help="PCI(Year of Analysis) ÷ PCI(Databank Year)"
    )

with col_cx5:
    st.text_input("Project Location", key="proj_location",
                  help="Enter the target country for this project.")

with col_cx6:
    location_factor = st.number_input(
        "Location Factor", min_value=0.0, step=0.01,
        key="location_factor",
        help="Multiplier that accounts for regional cost differences (e.g. 0.97 for Brazil)."
    )

with col_cx7:
    project_capex = (
        project_costs_isbl_osbl * (1 + contingency_pct)
        * time_update_factor
        * location_factor
    )
    st.metric("Project CAPEX", fmt_curr(project_capex))

st.divider()

# ─────────────────────────────────────────────
# Additional Information
# ─────────────────────────────────────────────
st.header("Additional Information")
col_ai1, col_ai2 = st.columns(2)
with col_ai1:
    working_hours = st.number_input(
        "Working Hours per Year (h/y)",
        min_value=1.0, max_value=8760.0,
        step=10.0, key="working_hours",
        help="Operating hours per year. Maximum is 8 760 h/y (continuous operation)."
    )
with col_ai2:
    scaling_factor = st.number_input(
        "Scaling Factor",
        min_value=0.001, step=0.01,
        key="scaling_factor",
        help="Six-tenths rule exponent or other capacity scaling factor. Must be > 0."
    )
st.divider()

# ─────────────────────────────────────────────
# Variable Costs
# ─────────────────────────────────────────────
st.header("8. Variable Costs")

RATE_UNITS  = list(RATE_TO_PRICE_UNIT.keys())
PRICE_UNITS = list(dict.fromkeys(RATE_TO_PRICE_UNIT.values()))  # unique, order-preserved

def _cost_table(section_key: str, editor_key: str) -> tuple[pd.DataFrame, float]:
    """
    Render a Name | Rate | Rate Unit | Price | Price Unit data_editor.
    Returns (active_rows_df, total_cost) where total = Σ Price × Rate × working_hours.
    section_key : key used to load saved rows from st.session_state.scenarios
    editor_key  : unique key for the st.data_editor widget
    """
    existing = (
        st.session_state.scenarios
        .get(st.session_state.sn_input, {})
        .get(section_key, [])
    )
    _blank = {"Name": None, "Rate": 0.0, "Rate Unit": RATE_UNITS[0],
              "Price": 0.0, "Price Unit": PRICE_UNITS[0]}
    if existing:
        df = pd.DataFrame(existing)[["Name", "Rate", "Rate Unit", "Price", "Price Unit"]]
    else:
        df = pd.DataFrame([_blank])

    edited = st.data_editor(
        df,
        key=editor_key,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Name":       st.column_config.TextColumn("Name",       width="large"),
            "Rate":       st.column_config.NumberColumn(
                              "Rate", min_value=0.0, step=0.000001,
                              format="%.6f",                        width="small"),
            "Rate Unit":  st.column_config.SelectboxColumn(
                              "Rate Unit", options=RATE_UNITS,      width="small"),
            "Price":      st.column_config.NumberColumn(
                              "Price", min_value=0.0, step=0.000001,
                              format="%.6f",                        width="small"),
            "Price Unit": st.column_config.SelectboxColumn(
                              "Price Unit", options=PRICE_UNITS,    width="small"),
        },
    )

    edited = edited.copy()
    edited["Price Unit"] = edited["Rate Unit"].map(RATE_TO_PRICE_UNIT).fillna(PRICE_UNITS[0])

    active = edited[edited["Name"].notna() & (edited["Name"].str.strip() != "")].copy()
    active["Line Cost"] = (
        pd.to_numeric(active["Price"], errors="coerce").fillna(0.0)
        * pd.to_numeric(active["Rate"],  errors="coerce").fillna(0.0)
        * working_hours
    )
    # Show line cost as formatted monetary string in a display-only column
    display_df = active.copy()
    display_df["Cost (USD/year)"] = display_df["Line Cost"].apply(lambda v: f"${v:,.2f}")
    if not display_df.empty:
        st.dataframe(
            display_df[["Name", "Rate", "Rate Unit", "Price", "Price Unit", "Cost (USD/year)"]],
            use_container_width=True, hide_index=True,
        )
    total = active["Line Cost"].sum()
    return active, total


def _total_row(label: str, value: float, key: str):
    st.markdown("---")
    c1, c2 = st.columns([3, 2])
    with c1:
        st.markdown(f"**{label}**")
    with c2:
        st.text_input(key, value=fmt_curr(value), disabled=True, label_visibility="collapsed")


# ── Raw Materials ──────────────────────────────
st.subheader("Raw Materials")
rm_active, total_raw_material_cost = _cost_table(
    "Raw Materials", f"rm_editor_{st.session_state.table_key}"
)
_total_row("Total raw materials cost", total_raw_material_cost, "total_rm")

st.divider()

# ── Chemical Inputs and Utilities ─────────────
st.subheader("Chemical Inputs and Utilities")
cu_active, total_chemical_utilities = _cost_table(
    "Chemical Inputs and Utilities", f"cu_editor_{st.session_state.table_key}"
)
_total_row("Total chemical inputs and utilities", total_chemical_utilities, "total_cu")

st.divider()

# ── Credits and Byproducts ────────────────────
st.subheader("Credits and Byproducts")
cb_active, total_revenue = _cost_table(
    "Credits and Byproducts", f"cb_editor_{st.session_state.table_key}"
)
_total_row("Total Revenue", total_revenue, "total_cb")

st.divider()

# Reference tables (PLANT_COST_INDEX, RATE_TO_PRICE_UNIT) are defined as
# constants above and used programmatically — no UI section needed.

# ─────────────────────────────────────────────
# 9. Fixed Costs
# ─────────────────────────────────────────────
st.header("9. Fixed Costs")
st.subheader("Labor Costs")

def _overridable_number(label: str, ref_val: float, override_key: str,
                        step: float = 0.01, min_value: float | None = None) -> float:
    """
    Show a number_input pre-seeded from ref_val.
    Only highlights amber (⚠ overridden) when user has changed the value.
    No color when value matches the reference.
    min_value defaults to None (no lower bound) to support negative values.
    """
    stored = st.session_state.get(override_key)
    current = stored if stored is not None else ref_val
    is_overridden = stored is not None and abs(stored - ref_val) > 1e-9

    if is_overridden:
        st.markdown(
            f'<p style="margin-bottom:0px;font-size:0.85rem;color:#e67e00;">'
            f'<b>{label}</b> ⚠ overridden</p>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<p style="margin-bottom:0px;font-size:0.85rem;">'
            f'<b>{label}</b></p>',
            unsafe_allow_html=True,
        )

    def _on_change(ok=override_key, rk=ref_val, wk=f"w_{override_key}"):
        v = st.session_state.get(wk, rk)
        st.session_state[ok] = v if abs(v - rk) > 1e-9 else None

    return st.number_input(
        label, min_value=min_value, step=step,
        value=float(current),
        key=f"w_{override_key}",
        label_visibility="collapsed",
        on_change=_on_change,
    )

# ── Row 1: Operators & Salary ──────────────────
col1, col2, col3, col4 = st.columns(4)
with col1:
    n_operators = st.number_input(
        "Number of operators per shift",
        min_value=1, step=1, key="n_operators",
        help="Must be a positive integer."
    )
with col2:
    operator_salary = st.number_input(
        "Operator salary (USD/month per operator)",
        min_value=MIN_SALARY, step=10.0, key="operator_salary",
        help=f"Cannot be lower than {fmt_curr(MIN_SALARY)}."
    )
    if operator_salary < MIN_SALARY:
        st.error(f"Operator salary cannot be lower than {fmt_curr(MIN_SALARY)}.")

with col3:
    n_supervisors = st.number_input(
        "Number of supervisors per shift",
        min_value=1, step=1, key="n_supervisors",
        help="Must be a positive integer and less than number of operators."
    )
    if n_supervisors >= n_operators:
        st.error("Number of supervisors must be less than number of operators.")

with col4:
    supervisor_salary = st.number_input(
        "Supervisor salary (USD/month per supervisor)",
        min_value=MIN_SALARY, step=10.0, key="supervisor_salary",
        help=f"Cannot be lower than {fmt_curr(MIN_SALARY)} and must exceed operator salary."
    )
    if supervisor_salary < MIN_SALARY:
        st.error(f"Supervisor salary cannot be lower than {fmt_curr(MIN_SALARY)}.")
    if supervisor_salary <= operator_salary:
        st.error("Supervisor salary must be greater than operator salary.")

# ── Row 2: Salary charges ──────────────────────
col1, col2 = st.columns([1, 3])
with col1:
    salary_charges = st.number_input(
        "Salary charges (multiplier)",
        min_value=1.0, step=0.05, key="salary_charges",
        help="Must be ≥ 1. Typical value: 2.2."
    )

st.markdown("---")

# ── Schedule inputs ────────────────────────────
st.markdown("##### Schedule")
col1, col2, col3 = st.columns(3)

# Plant working hours — overridable (tracks working_hours from Additional Info)
labor_wh_ref = st.session_state.get("working_hours", 8000.0)
with col1:
    labor_wh = _overridable_number(
        "Plant working hours per year", labor_wh_ref,
        "labor_working_hrs_override", step=10.0
    )
with col2:
    plant_daily_hours = st.number_input(
        "Plant daily operation hours (h/day)",
        min_value=0.1, max_value=24.0, step=0.5, key="plant_daily_hours"
    )
with col3:
    weekly_op_days = st.number_input(
        "Weekly operation days (days/week)",
        min_value=1.0, max_value=7.0, step=1.0, key="weekly_op_days"
    )

import math
operating_weeks = labor_wh / plant_daily_hours / weekly_op_days

col4, col5, col6 = st.columns(3)
with col4:
    st.text_input("Operating weeks per year",
                  value=f"{operating_weeks:.2f} weeks", disabled=True)
with col5:
    worker_hours_shift = st.number_input(
        "Worker hours per shift (h/shift)",
        min_value=0.1, step=0.5, key="worker_hours_shift"
    )
with col6:
    worker_shifts_week = st.number_input(
        "Worker shifts per week (shifts/week)",
        min_value=0.1, step=0.5, key="worker_shifts_week"
    )

col7, col8, col9 = st.columns(3)
with col7:
    worker_vacation_weeks = st.number_input(
        "Worker vacation weeks per year (weeks/year)",
        min_value=0.0, step=1.0, key="worker_vacation_weeks"
    )

worker_weeks_per_year = math.floor(operating_weeks - worker_vacation_weeks)

with col8:
    st.text_input("Worker weeks per year",
                  value=f"{worker_weeks_per_year} weeks", disabled=True)

# Operating team factor = (labor_wh / worker_hours_shift) / (worker_shifts_week * worker_weeks_per_year)
denom = worker_shifts_week * worker_weeks_per_year
op_team_factor = math.ceil((labor_wh / worker_hours_shift) / denom) if denom > 0 else 0

with col9:
    st.text_input("Operating team factor",
                  value=f"{op_team_factor}", disabled=True,
                  help="ceil( (Plant working hours / Worker hours per shift) / (Worker shifts per week × Worker weeks per year) )")

st.markdown("---")

# ── OLC ───────────────────────────────────────
olc = (
    (n_operators * operator_salary + n_supervisors * supervisor_salary)
    * salary_charges * op_team_factor * 12.0
)
st.markdown("##### Operating Labor Costs")
col1, col2, col3 = st.columns(3)
with col1:
    st.text_input("OLC – Operating Labor Costs (USD/year)",
                  value=fmt_curr(olc), disabled=True)

# ── Lab charges (overridable, shown as %) ──────
lab_ref_pct = LAB_CHARGES.get(st.session_state.dm_prod_type, 0.10) * 100.0
with col2:
    lab_pct_input = _overridable_number(
        f"Laboratory charges (% of OLC)  [ref: {lab_ref_pct:.2f}% — {st.session_state.dm_prod_type}]",
        lab_ref_pct, "lab_charges_override", step=0.1
    )
    lab_pct_input = max(0.0, lab_pct_input)
lab_pct = lab_pct_input / 100.0   # convert back to fraction for formula

# ── Office labor (overridable, shown as %) ─────
office_ref_pct = OFFICE_LABOR.get(st.session_state.dm_prod_type, 0.10) * 100.0
with col3:
    office_pct_input = _overridable_number(
        f"Office labor (% of OLC)  [ref: {office_ref_pct:.2f}% — {st.session_state.dm_prod_type}]",
        office_ref_pct, "office_labor_override", step=0.1
    )
    office_pct_input = max(0.0, office_pct_input)
office_pct = office_pct_input / 100.0   # convert back to fraction for formula

total_labor_costs = olc * (1.0 + lab_pct + office_pct)

st.markdown("---")
col_lc1, col_lc2 = st.columns([3, 2])
with col_lc1:
    st.markdown("**Total labor costs (USD/year)**")
with col_lc2:
    st.text_input("total_labor", value=fmt_curr(total_labor_costs),
                  disabled=True, label_visibility="collapsed")

st.divider()

# ── Supply and Maintenance ─────────────────────
st.subheader("Supply and Maintenance Costs")

maint_ref_pct = MAINTENANCE_REPAIRS.get(
    (st.session_state.dm_mat_type, st.session_state.dm_prod_type), 0.01
) * 100.0
op_sup_ref_pct = OPERATING_SUPPLIES.get(st.session_state.dm_severity, 0.15) * 100.0

col1, col2, col3 = st.columns(3)
with col1:
    maint_pct_input = _overridable_number(
        f"Maintenance and repairs (% of CAPEX)  [ref: {maint_ref_pct:.2f}% — {st.session_state.dm_mat_type} / {st.session_state.dm_prod_type}]",
        maint_ref_pct, "maint_repair_override", step=0.1
    )
maint_pct = maint_pct_input / 100.0

with col2:
    op_sup_pct_input = _overridable_number(
        f"Operating supplies (% of Maintenance)  [ref: {op_sup_ref_pct:.2f}% — {st.session_state.dm_severity}]",
        op_sup_ref_pct, "op_supplies_override", step=0.1
    )
op_sup_pct = op_sup_pct_input / 100.0

supply_maint_costs = (maint_pct + maint_pct * op_sup_pct) * project_capex

with col3:
    st.text_input("Supply and maintenance costs (USD/year)",
                  value=fmt_curr(supply_maint_costs), disabled=True)

st.divider()

# ── Additional Fixed Costs ─────────────────────
st.subheader("Additional Fixed Costs")

admin_ov_ref_pct  = ADMIN_OVERHEAD.get(st.session_state.dm_prod_type, 0.50) * 100.0 * (1 + office_pct)
mfg_ov_ref_pct    = MFG_OVERHEAD.get(st.session_state.dm_severity, 0.006) * 100.0
taxes_ins_ref_pct = TAXES_INSURANCE.get(st.session_state.dm_severity, 0.032) * 100.0
patents_ref       = PATENTS_ROYALTIES.get((st.session_state.dm_trl, st.session_state.dm_prod_type))
patents_ref_pct   = (patents_ref * 100.0) if patents_ref is not None else 0.0
patents_na        = patents_ref is None

col1, col2, col3, col4 = st.columns(4)
with col1:
    admin_ov_pct_input = _overridable_number(
        f"Administrative overhead (% of OLC)  [ref: {admin_ov_ref_pct:.2f}% — {st.session_state.dm_prod_type}]",
        admin_ov_ref_pct, "admin_overhead_override", step=0.1
    )
admin_ov_pct = admin_ov_pct_input / 100.0

with col2:
    mfg_ov_pct_input = _overridable_number(
        f"Manufacturing overhead (% of CAPEX)  [ref: {mfg_ov_ref_pct:.2f}% — {st.session_state.dm_severity}]",
        mfg_ov_ref_pct, "mfg_overhead_override", step=0.01
    )
mfg_ov_pct = mfg_ov_pct_input / 100.0

with col3:
    taxes_ins_pct_input = _overridable_number(
        f"Taxes and insurance (% of CAPEX)  [ref: {taxes_ins_ref_pct:.2f}% — {st.session_state.dm_severity}]",
        taxes_ins_ref_pct, "taxes_ins_override", step=0.01
    )
taxes_ins_pct = taxes_ins_pct_input / 100.0

with col4:
    if patents_na:
        st.markdown(
            '<p style="margin-bottom:0px;font-size:0.85rem;">'
            '<b>Patents and royalties (% of OPEX)</b></p>',
            unsafe_allow_html=True
        )
        st.info(f"N/A for {st.session_state.dm_trl} / {st.session_state.dm_prod_type}")
        patents_pct = 0.0
    else:
        patents_pct_input = _overridable_number(
            f"Patents and royalties (% of OPEX)  [ref: {patents_ref_pct:.2f}% — {st.session_state.dm_trl} / {st.session_state.dm_prod_type}]",
            patents_ref_pct, "patents_roy_override", step=0.1
        )
        patents_pct = patents_pct_input / 100.0

afc_pre_patents = (admin_ov_pct * olc) + (mfg_ov_pct * project_capex) + (taxes_ins_pct * project_capex)

st.divider()

# ── Indirect Fixed Costs ───────────────────────
st.subheader("Indirect Fixed Costs")

admin_costs_ref_pct = (1.0 + office_pct) * 0.15 * 100.0
mfg_costs_ref_pct   = maint_pct * 0.15 * 100.0
dist_sell_ref_pct   = DIST_SELLING.get(st.session_state.dm_prod_type, 0.08) * 100.0
r_d_ref_pct         = R_AND_D.get((st.session_state.dm_trl, st.session_state.dm_prod_type), 0.02) * 100.0

col1, col2, col3, col4 = st.columns(4)
with col1:
    admin_costs_pct_input = _overridable_number(
        f"Administrative costs (% of OLC)  [ref: {admin_costs_ref_pct:.2f}%]",
        admin_costs_ref_pct, "admin_costs_override", step=0.1
    )
admin_costs_pct = admin_costs_pct_input / 100.0

with col2:
    mfg_costs_pct_input = _overridable_number(
        f"Manufacturing costs (% of CAPEX)  [ref: {mfg_costs_ref_pct:.2f}%]",
        mfg_costs_ref_pct, "mfg_costs_override", step=0.01
    )
mfg_costs_pct = mfg_costs_pct_input / 100.0

with col3:
    dist_sell_pct_input = _overridable_number(
        f"Distribution and selling (% of OPEX)  [ref: {dist_sell_ref_pct:.2f}% — {st.session_state.dm_prod_type}]",
        dist_sell_ref_pct, "dist_selling_override", step=0.1
    )
dist_sell_pct = dist_sell_pct_input / 100.0

with col4:
    r_d_pct_input = _overridable_number(
        f"Research and development (% of OPEX)  [ref: {r_d_ref_pct:.2f}% — {st.session_state.dm_trl} / {st.session_state.dm_prod_type}]",
        r_d_ref_pct, "r_and_d_override", step=0.1
    )
r_d_pct = r_d_pct_input / 100.0

# ── Resolve OPEX & all costs analytically ────────────────────────────────
# AFC = admin_ov*OLC + (mfg_ov + taxes_ins)*CAPEX + patents*OPEX
#
# OPEX = [ TVC + labor + supply_maint
#           + (admin_ov + admin_costs)*OLC
#           + (mfg_ov + taxes_ins + mfg_costs)*CAPEX ]
#         / (1 - patents - dist_sell - r_d)

direct_var_costs = total_raw_material_cost + total_chemical_utilities - total_revenue

_olc_coeff   = admin_ov_pct + admin_costs_pct
_capex_coeff = mfg_ov_pct + taxes_ins_pct + mfg_costs_pct

# OPEX base (shared numerator — used by dist_sell, r_d, and patents)
_numerator   = (direct_var_costs
                + total_labor_costs
                + supply_maint_costs
                + _olc_coeff   * olc
                + _capex_coeff * project_capex)
_denominator = 1.0 - patents_pct - dist_sell_pct - r_d_pct
opex = _numerator / _denominator if _denominator > 0 else 0.0

# AFC = admin_ov×OLC + (mfg_ov + taxes_ins)×CAPEX + patents×OPEX
afc = (admin_ov_pct * olc
       + (mfg_ov_pct + taxes_ins_pct) * project_capex
       + patents_pct * opex)

# Indirect = adm_costs×OLC + mfg_costs×CAPEX + (dist_sell + r_d)×OPEX
# Note: patents is already included in AFC, NOT repeated here
_opex_base = _numerator / _denominator if _denominator > 0 else 0.0
indirect_fixed_costs = (admin_costs_pct * olc
                        + mfg_costs_pct  * project_capex
                        + (dist_sell_pct + r_d_pct) * _opex_base)

direct_fixed_costs = total_labor_costs + supply_maint_costs + afc
total_fixed_costs  = direct_fixed_costs + indirect_fixed_costs

# Total OPEX = the analytically solved value (algebraically equal to TVC + DFC + IFC)
# Using opex directly avoids any floating-point divergence
total_opex = opex

# ── Step-by-step Indirect Fixed Costs breakdown ──────────────────────────

st.markdown("---")
_result_row("Additional fixed costs (USD/year)", afc, "afc_display",
    "AFC = admin_overhead×OLC + (mfg_overhead + taxes_insurance)×CAPEX + patents×OPEX")

st.markdown("---")
_result_row("Direct fixed costs (USD/year)", direct_fixed_costs, "dfc_display",
    "Direct fixed costs = Total labor costs + Supply & maintenance + AFC")

st.markdown("---")
_result_row("Indirect fixed costs (USD/year)", indirect_fixed_costs, "ifc_display",
    "Indirect fixed costs = adm_costs×OLC + mfg_costs×CAPEX + (dist_sell + R&D)×OPEX")

st.markdown("---")
_result_row("Total fixed costs (USD/year)", total_fixed_costs, "tfc_display",
    "Total fixed costs = Direct fixed costs + Indirect fixed costs")

st.divider()

_result_row("Total OPEX (USD/year)", total_opex, "total_opex_display",
    "OPEX = (TVC + Labor + Supply + (admin_ov+adm_costs)×OLC + (mfg_ov+taxes+mfg_costs)×CAPEX) / (1 − patents − dist_sell − R&D)")

st.divider()

# ─────────────────────────────────────────────
# 10. Working Capital
# ─────────────────────────────────────────────
st.header("10. Working Capital")

wc_method = st.radio(
    "Working capital method",
    options=["Percentage", "Operating Cycle"],
    index=0 if st.session_state.wc_method == "Percentage" else 1,
    horizontal=True,
    key="wc_method",
)
st.markdown("---")

if wc_method == "Percentage":
    col1, col2 = st.columns(2)
    with col1:
        wc_pct_input = st.number_input(
            "Working capital percentage (% of CAPEX)",
            min_value=0.0, step=0.5,
            value=st.session_state.wc_pct,
            key="wc_pct",
        )
    working_capital = project_capex * (wc_pct_input / 100.0)
    with col2:
        st.markdown("&nbsp;", unsafe_allow_html=True)
        _result_row("Working capital (USD)", working_capital, "wc_display",
            "Working capital = CAPEX × Working capital %")

else:  # Operating Cycle
    # ── Current Assets inputs ──────────────────
    st.markdown("##### Current Assets")
    col1, col2, col3 = st.columns(3)
    with col1:
        equiv_cash_days = st.number_input(
            "Equivalent cash (days of OPEX)",
            min_value=0.0, step=1.0,
            value=st.session_state.wc_equiv_cash_days,
            key="wc_equiv_cash_days",
        )
    with col2:
        raw_mat_days = st.number_input(
            "Raw material inventory (days of variable costs)",
            min_value=0.0, step=1.0,
            value=st.session_state.wc_raw_mat_days,
            key="wc_raw_mat_days",
        )
    with col3:
        acc_rec_days = st.number_input(
            "Accounts receivable (days of OPEX and simple ROI)",
            min_value=0.0, step=1.0,
            value=st.session_state.wc_accounts_rec_days,
            key="wc_accounts_rec_days",
        )

    # PMT(0.15, 10, -CAPEX) — annual payment on a 10-year loan at 15%
    # PMT = r×PV / (1 − (1+r)^−n)
    _r, _n = 0.15, 10
    simple_roi = (_r * project_capex) / (1.0 - (1.0 + _r) ** (-_n)) if project_capex > 0 else 0.0

    current_assets = (
        (equiv_cash_days * total_opex
         + direct_var_costs * raw_mat_days
         + acc_rec_days * (total_opex + simple_roi))
        * (24.0 / working_hours)
    )

    st.markdown("---")
    _result_row("Current assets (USD)", current_assets, "current_assets_display",
        "Current assets = (Equiv_cash×OPEX + TVC×Raw_mat_days + Acc_rec×(OPEX + PMT(15%,10,-CAPEX))) × (24 / working_hours)")

    st.markdown("---")

    # ── Current Liabilities inputs ─────────────
    st.markdown("##### Current Liabilities")
    col1, col2 = st.columns(2)
    with col1:
        acc_payroll_days = st.number_input(
            "Accrued payroll (days of OLC)",
            min_value=0.0, step=1.0,
            value=st.session_state.wc_accrued_payroll_days,
            key="wc_accrued_payroll_days",
        )
    with col2:
        acc_pay_days = st.number_input(
            "Accounts payable (days of OPEX)",
            min_value=0.0, step=1.0,
            value=st.session_state.wc_accounts_pay_days,
            key="wc_accounts_pay_days",
        )

    current_liabilities = (
        (acc_payroll_days * olc + acc_pay_days * total_opex)
        * (24.0 / working_hours)
    )

    st.markdown("---")
    _result_row("Current liabilities (USD)", current_liabilities, "current_liab_display",
        "Current liabilities = (Accrued_payroll×OLC + Acc_payable×OPEX) × (24 / working_hours)")

    working_capital = current_assets - current_liabilities

    st.markdown("---")
    _result_row("Working capital (USD)", working_capital, "wc_display",
        "Working capital = Current assets − Current liabilities")

st.divider()

# Reference tables (PLANT_COST_INDEX, RATE_TO_PRICE_UNIT) are defined as
# constants above and used programmatically — no UI section needed.

# ─────────────────────────────────────────────
# 11. Startup Costs
# ─────────────────────────────────────────────
st.header("11. Startup Costs")

startup_method = st.radio(
    "Startup costs method",
    options=["Single Factor", "Multiple Factors"],
    index=0 if st.session_state.startup_method == "Single Factor" else 1,
    horizontal=True,
    key="startup_method",
)
st.markdown("---")

if startup_method == "Single Factor":
    col1, col2 = st.columns(2)
    with col1:
        startup_single_pct = st.number_input(
            "Startup cost single factor (% of CAPEX)",
            min_value=0.0, step=0.5,
            value=st.session_state.startup_single_pct,
            key="startup_single_pct",
        )
    startup_costs = project_capex * (startup_single_pct / 100.0)
    with col2:
        st.markdown("&nbsp;", unsafe_allow_html=True)
        _result_row("Startup costs (USD)", startup_costs, "startup_display",
            "Startup costs = CAPEX × Startup single factor %")

else:  # Multiple Factors
    col1, col2, col3 = st.columns(3)
    with col1:
        startup_op_training_days = st.number_input(
            "Operational training (days of OLC)",
            min_value=0.0, step=1.0,
            value=st.session_state.startup_op_training_days,
            key="startup_op_training_days",
        )
    with col2:
        startup_commerc_pct = st.number_input(
            "Commercialization costs (% of direct field costs)",
            min_value=0.0, step=0.5,
            value=st.session_state.startup_commerc_pct,
            key="startup_commerc_pct",
        )
    with col3:
        startup_inefficiency_pct = st.number_input(
            "Startup inefficiency (% of OPEX)",
            min_value=0.0, step=0.5,
            value=st.session_state.startup_inefficiency_pct,
            key="startup_inefficiency_pct",
        )

    startup_costs = (
        startup_op_training_days * olc * (24.0 / working_hours)
        + total_direct_field_costs * (startup_commerc_pct / 100.0)
        + (startup_inefficiency_pct / 100.0) * total_opex
    )
    st.markdown("---")
    _result_row("Startup costs (USD)", startup_costs, "startup_costs_display",
        "Startup costs = Op_training×OLC×(24/working_hours) + Commerc%×Direct_field_costs + Inefficiency%×Total_OPEX")

st.divider()

# ─────────────────────────────────────────────
# Total Investment Summary
# ─────────────────────────────────────────────
st.subheader("Total Investment Summary")

col1, col2 = st.columns([3, 2])
with col1:
    additional_costs = st.number_input(
        "Additional costs (USD)",
        min_value=0.0, step=1000.0,
        key="additional_costs",
        help="Any additional investment cost not captured above.",
    )

total_investment = project_capex + working_capital + startup_costs + additional_costs

st.markdown("---")
_result_row("Project CAPEX (USD)", project_capex, "ti_capex",
    "Project CAPEX = Project costs × (1 + contingency) × time_update_factor × location_factor")
st.markdown("---")
_result_row("Working capital (USD)", working_capital, "ti_wc",
    "Working capital — see section 10")
st.markdown("---")
_result_row("Startup costs (USD)", startup_costs, "ti_startup",
    "Startup costs — see section 11")
st.markdown("---")
_result_row("Additional costs (USD)", additional_costs, "ti_additional",
    "User-defined additional investment costs")
st.markdown("---")
_result_row("Total investment costs (USD)", total_investment, "ti_total",
    "TIC = Project CAPEX + Working capital + Startup costs + Additional costs")

st.divider()

# ─────────────────────────────────────────────
# Total Investment Cost (TIC) Accuracy
# ─────────────────────────────────────────────
st.subheader("Total Investment Cost (TIC) Accuracy")
st.caption(f"Reference: TRL = {st.session_state.dm_trl} | "
           f"Information availability = {st.session_state.dm_info_avail}")

_tic_key = (st.session_state.dm_trl, st.session_state.dm_info_avail)
tic_lower_ref = TIC_LOWER.get(_tic_key)
tic_upper_ref = TIC_UPPER.get(_tic_key)

col1, col2 = st.columns(2)
with col1:
    if tic_lower_ref is None:
        st.markdown('<p style="margin-bottom:0px;font-size:0.85rem;"><b>Lower bound (%)</b></p>',
                    unsafe_allow_html=True)
        st.info(f"N/A for {st.session_state.dm_trl} / {st.session_state.dm_info_avail}")
        tic_lower_input = None
    else:
        tic_lower_input = _overridable_number(
            f"Lower bound (%)  [ref: {tic_lower_ref:.1f}%]",
            tic_lower_ref, "tic_lower_override", step=0.5, min_value=None
        )

with col2:
    if tic_upper_ref is None:
        st.markdown('<p style="margin-bottom:0px;font-size:0.85rem;"><b>Upper bound (%)</b></p>',
                    unsafe_allow_html=True)
        st.info(f"N/A for {st.session_state.dm_trl} / {st.session_state.dm_info_avail}")
        tic_upper_input = None
    else:
        tic_upper_input = _overridable_number(
            f"Upper bound (%)  [ref: {tic_upper_ref:.1f}%]",
            tic_upper_ref, "tic_upper_override", step=0.5, min_value=None
        )

# TIC range results
st.markdown("---")
col1, col2 = st.columns(2)
with col1:
    if tic_lower_input is not None:
        tic_lower_value = total_investment * (1.0 + tic_lower_input / 100.0)
        _result_row("TIC lower bound (USD)", tic_lower_value, "tic_lower_usd",
            f"TIC lower = Total investment × (1 + {tic_lower_input:.1f}%)")
    else:
        st.info("TIC lower bound: N/A")
with col2:
    if tic_upper_input is not None:
        tic_upper_value = total_investment * (1.0 + tic_upper_input / 100.0)
        _result_row("TIC upper bound (USD)", tic_upper_value, "tic_upper_usd",
            f"TIC upper = Total investment × (1 + {tic_upper_input:.1f}%)")
    else:
        st.info("TIC upper bound: N/A")

st.divider()

# ─────────────────────────────────────────────
# 12. Financial Assumptions
# ─────────────────────────────────────────────
st.header("12. Financial Assumptions")

# ── Land ──────────────────────────────────────
st.subheader("Land")

land_option = st.radio(
    "Land option",
    options=["Buy", "Rent"],
    index=0 if st.session_state.land_option == "Buy" else 1,
    horizontal=True,
    key="land_option",
)
st.markdown("---")

col1, col2 = st.columns(2)
if land_option == "Buy":
    with col1:
        land_buy_pct = _overridable_number(
            f"Land purchase factor (% of ISBL+OSBL)  [ref: 2.00%]",
            2.0, "land_buy_pct_override", step=0.1
        )
    land_rent_pct = 0.0
    with col2:
        land_cost_usd = project_costs_isbl_osbl * (land_buy_pct / 100.0)
        _result_row("Land cost (USD)", land_cost_usd, "land_cost_display",
            "Land cost = (ISBL+OSBL costs) × Land purchase factor %")
else:  # Rent
    with col1:
        land_rent_pct = _overridable_number(
            f"Land rent factor (% of ISBL+OSBL per year)  [ref: 0.20%]",
            0.2, "land_rent_pct_override", step=0.01
        )
    land_buy_pct = 0.0
    with col2:
        land_rent_annual_usd = project_costs_isbl_osbl * (land_rent_pct / 100.0)
        _result_row("Annual land rent (USD/year)", land_rent_annual_usd, "land_rent_display",
            "Annual land rent = (ISBL+OSBL costs) × Land rent factor %")

st.divider()

# ── Depreciation ──────────────────────────────
st.subheader("Depreciation")

depreciation_method = st.radio(
    "Depreciation method",
    options=["Straight Line", "MACRS"],
    index=0 if st.session_state.depreciation_method == "Straight Line" else 1,
    horizontal=True,
    key="depreciation_method",
)
st.markdown("---")

col1, col2 = st.columns(2)
with col1:
    depreciation_years = st.number_input(
        "Depreciation period (years)",
        min_value=1, step=1,
        key="depreciation_years",
        help="Number of years over which CAPEX is depreciated. Must be a positive integer.",
    )
with col2:
    residual_value_pct = st.number_input(
        "Residual value (% of CAPEX)",
        min_value=0.0, max_value=99.9, step=1.0,
        key="residual_value_pct",
        help="Salvage/residual value at end of depreciation period, as % of Project CAPEX.",
    )

residual_value_usd = project_capex * (residual_value_pct / 100.0)
depreciable_base   = project_capex - residual_value_usd

col3, col4 = st.columns(2)
with col3:
    _result_row("Residual value (USD)", residual_value_usd, "residual_value_display",
        "Residual value = Project CAPEX × Residual value %")
with col4:
    _result_row("Depreciable base (USD)", depreciable_base, "depreciable_base_display",
        "Depreciable base = Project CAPEX − Residual value")

st.divider()

# ── Taxes ─────────────────────────────────────
st.subheader("Taxes")

def _on_country_change():
    """Clear the tax rate override whenever the country selection changes."""
    st.session_state.tax_rate_override = None
    # Also clear the widget key so _overridable_number re-seeds from new ref
    st.session_state.pop("w_tax_rate_override", None)

col1, col2 = st.columns(2)
with col1:
    tax_country = st.selectbox(
        "Country",
        options=COUNTRY_LIST,
        index=COUNTRY_LIST.index(st.session_state.get("tax_country", "Brazil")),
        key="tax_country",
        on_change=_on_country_change,
    )
tax_ref_pct = TAXES_BY_COUNTRY.get(tax_country, 0.34) * 100.0
with col2:
    tax_rate_input = _overridable_number(
        f"Corporate tax rate (%)  [ref: {tax_ref_pct:.1f}% — {tax_country}]",
        tax_ref_pct, "tax_rate_override", step=0.1
    )
tax_rate = tax_rate_input / 100.0

st.divider()

# ── Financial Leverage ────────────────────────
st.subheader("Financial Leverage")

financing_type = st.radio(
    "Financing type",
    options=["None", "Straight Line"],
    index=0 if st.session_state.financing_type == "None" else 1,
    horizontal=True,
    key="financing_type",
)
st.markdown("---")

is_leveraged = financing_type == "Straight Line"

col1, col2, col3 = st.columns(3)
with col1:
    debt_ratio_pct = st.number_input(
        "Debt ratio (% of CAPEX)",
        min_value=0.0, max_value=99.9, step=1.0,
        key="debt_ratio_pct",
        disabled=not is_leveraged,
        help="Proportion of CAPEX financed by debt.",
    )
with col2:
    amortization_years = st.number_input(
        "Amortization period (years)",
        min_value=1, step=1,
        key="amortization_years",
        disabled=not is_leveraged,
    )
with col3:
    grace_period_years = st.number_input(
        "Grace period (years)",
        min_value=0, step=1,
        key="grace_period_years",
        disabled=not is_leveraged,
        help="Number of years before principal repayment begins.",
    )

debt_ratio = (debt_ratio_pct / 100.0) if is_leveraged else 0.0

st.divider()

# ── MARR ──────────────────────────────────────
st.subheader("Minimum Acceptable Rate of Return (MARR)")

col1, col2 = st.columns(2)
with col1:
    central_bank_rate = st.number_input(
        "Central bank interest rate (%)",
        min_value=0.0, step=0.01,
        key="central_bank_rate",
    )
with col2:
    credit_spread = st.number_input(
        "Credit spread (%)",
        min_value=0.0, step=0.01,
        key="credit_spread",
    )

# COD = (central_bank_rate + credit_spread) × (1 − tax_rate)
cod = (central_bank_rate + credit_spread) / 100.0 * (1.0 - tax_rate)
st.markdown("---")
col1, col2 = st.columns(2)
with col1:
    _result_row("Cost of Debt — COD (%)", cod * 100.0, "cod_display",
        "COD = (Central bank rate + Credit spread) × (1 − Corporate tax rate)")
st.markdown("---")

col1, col2 = st.columns(2)
with col1:
    unlevered_beta = st.number_input(
        "Unlevered / Asset beta",
        min_value=0.0, step=0.01,
        key="unlevered_beta",
    )

# Levered beta = unlevered_beta × (1 + (1 − tax) × debt_ratio / (1 − debt_ratio))
# debt_ratio is 0 when financing_type is None
_leverage_factor = (1.0 - tax_rate) * debt_ratio / (1.0 - debt_ratio) if debt_ratio < 1.0 else 0.0
levered_beta = unlevered_beta * (1.0 + _leverage_factor)

with col2:
    _result_row("Levered / Equity beta", levered_beta, "levered_beta_display",
        "Levered beta = Unlevered beta × (1 + (1−tax) × debt_ratio / (1−debt_ratio))"
        "  |  debt_ratio = 0 when financing type is None")

st.markdown("---")
col1, col2, col3, col4 = st.columns(4)
with col1:
    market_return = st.number_input(
        "Market return — rM (%)",
        min_value=0.0, step=0.01,
        key="market_return",
    )
with col2:
    risk_free_rate = st.number_input(
        "Risk-free rate of return (%)",
        min_value=0.0, step=0.01,
        key="risk_free_rate",
    )

# Market risk premium = levered_beta × (rM − risk_free_rate)
market_risk_premium = levered_beta * (market_return - risk_free_rate) / 100.0

with col3:
    _result_row("Market risk premium — PRM (%)", market_risk_premium * 100.0,
        "mrp_display",
        "PRM = Levered beta × (Market return − Risk-free rate)")
with col4:
    country_risk_premium = st.number_input(
        "Country risk premium — PRP (%)",
        min_value=0.0, step=0.01,
        key="country_risk_premium",
    )

st.markdown("---")
col1, col2 = st.columns(2)
with col1:
    us_cpi = st.number_input(
        "U.S. Consumer Price Index (%)",
        min_value=0.0, step=0.01,
        key="us_cpi",
    )
with col2:
    country_cpi = st.number_input(
        "Country Consumer Price Index (%)",
        min_value=0.0, step=0.01,
        key="country_cpi",
    )

# COE = ((1+rfr)*(1+country_cpi)/(1+us_cpi) − 1) + market_risk_premium + country_risk_premium
_rfr   = risk_free_rate  / 100.0
_ccpi  = country_cpi     / 100.0
_uscpi = us_cpi          / 100.0
coe = ((1.0 + _rfr) * (1.0 + _ccpi) / (1.0 + _uscpi) - 1.0) + market_risk_premium + country_risk_premium / 100.0

st.markdown("---")
_result_row("Cost of Equity — COE (%)", coe * 100.0, "coe_display",
    "COE = ((1+rfr)×(1+country_CPI)/(1+US_CPI) − 1) + Market risk premium + Country risk premium")

# MARR — formula default, user-editable
_marr_formula = (debt_ratio * cod + (1.0 - debt_ratio) * coe) if is_leveraged else coe
_marr_ref_pct = _marr_formula * 100.0

st.markdown("---")
col1, col2 = st.columns(2)
with col1:
    marr_input = _overridable_number(
        f"MARR — Minimum Acceptable Rate of Return (%)",
        _marr_ref_pct, "marr_override", step=0.01, min_value=None
    )
marr = marr_input / 100.0

with col2:
    st.markdown(
        '<p style="font-size:0.82rem;color:#555;margin-top:28px;">'
        + ("COE  (financing type: None)" if not is_leveraged
           else f"debt_ratio×COD + (1−debt_ratio)×COE  |  {debt_ratio_pct:.1f}%×{cod*100:.2f}% + {100-debt_ratio_pct:.1f}%×{coe*100:.2f}%")
        + "</p>", unsafe_allow_html=True
    )

st.divider()

# ─────────────────────────────────────────────
# 13. Project Lifetime and CAPEX Distribution
# ─────────────────────────────────────────────
st.header("13. Project Lifetime and CAPEX Distribution")

col1, col2, col3 = st.columns(3)
with col1:
    epc_years = st.number_input(
        "EPC time (years)",
        min_value=1, max_value=10, step=1,
        key="epc_years",
        help="Engineering, Procurement and Construction duration. Must be between 1 and 10.",
    )
with col2:
    project_lifetime = st.number_input(
        "Project lifetime (years)",
        min_value=1, max_value=40, step=1,
        key="project_lifetime",
    )
with col3:
    total_project_lifetime = epc_years + project_lifetime
    _result_row("Total project lifetime (years)", float(total_project_lifetime),
        "total_lifetime_display",
        "Total project lifetime = EPC time + Project lifetime")

st.markdown("---")

# CAPEX distribution table — read from reference, display as editable-looking table
_capex_col = str(epc_years)
_capex_dist = CAPEX_DISTRIBUTION[_capex_col]   # Series indexed 1st…10th

_year_labels = ["1st year (%)","2nd year (%)","3rd year (%)","4th year (%)",
                "5th year (%)","6th year (%)","7th year (%)","8th year (%)",
                "9th year (%)","10th year (%)"]
_capex_df = pd.DataFrame({
    "Year":               _year_labels,
    "CAPEX Distribution": [f"{v*100:.4f}%" for v in _capex_dist.values],
})
st.dataframe(_capex_df, use_container_width=False, hide_index=True,
             column_config={"Year": st.column_config.TextColumn(width="medium"),
                            "CAPEX Distribution": st.column_config.TextColumn(width="small")})

st.divider()

# ── Percentage of Total Capacity ──────────────
st.subheader("Percentage of Total Capacity per Year")
col1, col2, col3 = st.columns(3)
with col1:
    capacity_first_year = st.number_input(
        "First year (%)", min_value=0.0, max_value=100.0, step=1.0,
        key="capacity_first_year",
    )
with col2:
    capacity_intermediate = st.number_input(
        "Intermediate years (%)", min_value=0.0, max_value=100.0, step=1.0,
        key="capacity_intermediate",
    )
with col3:
    capacity_last_year = st.number_input(
        "Last year (%)", min_value=0.0, max_value=100.0, step=1.0,
        key="capacity_last_year",
    )

st.divider()

# ── Percentage of Fixed Costs per Year ─────────
st.subheader("Percentage of Fixed Costs per Year")
col1, col2, col3 = st.columns(3)
with col1:
    fixed_costs_first_year = st.number_input(
        "First year (%)", min_value=0.0, max_value=100.0, step=1.0,
        key="fixed_costs_first_year",
    )
with col2:
    fixed_costs_intermediate = st.number_input(
        "Intermediate years (%)", min_value=0.0, max_value=100.0, step=1.0,
        key="fixed_costs_intermediate",
    )
with col3:
    fixed_costs_last_year = st.number_input(
        "Last year (%)", min_value=0.0, max_value=100.0, step=1.0,
        key="fixed_costs_last_year",
    )

st.divider()

# ── Market Assumptions ─────────────────────────
st.subheader("Market Assumptions — Annual Growth Rates")
st.caption("Enter the expected annual % change for each cost/price category.")

col1, col2, col3, col4, col5 = st.columns(5)
with col1:
    growth_main_price = st.number_input(
        "Main product price (% / year)",
        step=0.1, key="growth_main_price",
    )
with col2:
    growth_byproduct_price = st.number_input(
        "Byproduct prices (% / year)",
        step=0.1, key="growth_byproduct_price",
    )
with col3:
    growth_raw_materials = st.number_input(
        "Raw materials costs (% / year)",
        step=0.1, key="growth_raw_materials",
    )
with col4:
    growth_chem_utilities = st.number_input(
        "Chemical inputs & utilities costs (% / year)",
        step=0.1, key="growth_chem_utilities",
    )
with col5:
    growth_fixed_costs = st.number_input(
        "Fixed costs (% / year)",
        step=0.1, key="growth_fixed_costs",
    )

st.divider()

# ── Other Premises ─────────────────────────────
st.subheader("Other Premises")
col1, col2 = st.columns(2)
with col1:
    main_product_price = st.number_input(
        f"Main product selling price (USD / {st.session_state.pu_input})",
        min_value=0.0, step=0.01,
        value=st.session_state.main_product_price if st.session_state.main_product_price is not None else 0.0,
        key="main_product_price",
        help="Selling price per unit of the main product. Reset to 0 when a new scenario is loaded.",
    )

st.divider()

# ─────────────────────────────────────────────
# SAVE
if st.button("Save / Update Scenario", type="primary"):
    if not st.session_state.sn_input:
        st.error("Please provide a Scenario Name before saving.")
    elif st.session_state.pc_input is None:
        st.error("Please provide a Main Product Capacity before saving.")
    else:
        util_bool = st.session_state.lang_utility if st.session_state.oth_cost_src == "Lang Factors" else False

        # Unpack nf_values list to named keys
        freight, taxes_permits, eng_ho, ga_overheads, contract_fee = nf_values
        piping, civil, steel, instrumentals, electrical, insulation = inst_values

        st.session_state.scenarios[st.session_state.sn_input] = {
            "Product Name":              st.session_state.mp_input,
            "Unit":                      st.session_state.pu_input,
            "Capacity":                  st.session_state.pc_input,
            "Capacity Label":            f"{st.session_state.pc_input} {st.session_state.pu_input}/year",
            "Process Variables":         edited_process_vars.dropna(how="all").to_dict(orient="records"),
            "Equipment Costs Source":    st.session_state.eq_cost_src,
            "Other Costs Source":        st.session_state.oth_cost_src,
            "Contains Utility Systems":  util_bool,
            "Product Type":              st.session_state.dm_prod_type,
            "TRL":                       st.session_state.dm_trl,
            "Info Availability":         st.session_state.dm_info_avail,
            "Process Severity":          st.session_state.dm_severity,
            "Material Handled":          st.session_state.dm_mat_type,
            "Plant Size":                st.session_state.dm_plant_size,
            # CAPEX
            "Equipment Acquisition":     equip_acq,
            "Spare Parts":               spare_parts,
            "Equipment Setting":         equip_setting,
            "Base Equipment Costs":      base_equip_costs,
            "Unscheduled Equip Pct":     unsched_pct,
            "Total Equipment Costs":     total_equip_costs,
            "Piping":                    piping,
            "Civil":                     civil,
            "Steel":                     steel,
            "Instrumentals":             instrumentals,
            "Electrical":                electrical,
            "Insulation":                insulation,
            "Paint":                     paint,
            "Total Installation Costs":  total_inst_costs,
            "Total Direct Field Costs":  total_direct_field_costs,
            "ISBL Contribution (%)":     isbl_contrib,
            "OSBL Contribution (%)":     100.0 - isbl_contrib,
            "Field Office Staff":        field_office,
            "Construction Indirects":    const_indirects,
            "Total Indirect Field Costs":total_indirect_field_costs,
            "Freight":                   freight,
            "Taxes and Permits":         taxes_permits,
            "Engineering and HO":        eng_ho,
            "GA Overheads":              ga_overheads,
            "Contract Fee":              contract_fee,
            "Total Non-Field Costs":     total_non_field_costs,
            "Project Costs ISBL+OSBL":   project_costs_isbl_osbl,
            # CAPEX calculations
            "Allow Override":            st.session_state.allow_override,
            "Contingency Pct":           contingency_pct,
            "Databank Year":             int(databank_year_input),
            "Year of Analysis":          effective_analysis_year,
            "PCI Databank":              pci_databank,
            "PCI Analysis":              pci_analysis,
            "Time Update Factor":        time_update_factor,
            "Project Location":          st.session_state.proj_location,
            "Location Factor":           location_factor,
            "Project CAPEX":             project_capex,
            # Additional information
            "Working Hours per Year":    working_hours,
            "Scaling Factor":            scaling_factor,
            # Variable costs
            "Raw Materials":                    rm_active.drop(columns=["Line Cost"]).to_dict(orient="records"),
            "Total Raw Material Cost":          total_raw_material_cost,
            "Chemical Inputs and Utilities":    cu_active.drop(columns=["Line Cost"]).to_dict(orient="records"),
            "Total Chemical Inputs Utilities":  total_chemical_utilities,
            "Credits and Byproducts":           cb_active.drop(columns=["Line Cost"]).to_dict(orient="records"),
            "Total Revenue":                    total_revenue,
            # Fixed costs — Labor
            "Num Operators":             n_operators,
            "Operator Salary":           operator_salary,
            "Num Supervisors":           n_supervisors,
            "Supervisor Salary":         supervisor_salary,
            "Salary Charges":            salary_charges,
            "Plant Daily Hours":         plant_daily_hours,
            "Weekly Op Days":            weekly_op_days,
            "Operating Weeks":           operating_weeks,
            "Worker Hours per Shift":    worker_hours_shift,
            "Worker Shifts per Week":    worker_shifts_week,
            "Worker Vacation Weeks":     worker_vacation_weeks,
            "Worker Weeks per Year":     worker_weeks_per_year,
            "Operating Team Factor":     op_team_factor,
            "OLC":                       olc,
            "Lab Charges Override":      st.session_state.get("lab_charges_override"),
            "Office Labor Override":     st.session_state.get("office_labor_override"),
            "Labor Working Hrs Override":st.session_state.get("labor_working_hrs_override"),
            "Lab Charges Pct":           lab_pct,
            "Office Labor Pct":          office_pct,
            "Total Labor Costs":         total_labor_costs,
            # Fixed costs — Supply & Maintenance
            "Maint Repair Override":     st.session_state.get("maint_repair_override"),
            "Op Supplies Override":      st.session_state.get("op_supplies_override"),
            "Maint Pct":                 maint_pct,
            "Op Sup Pct":                op_sup_pct,
            "Supply Maint Costs":        supply_maint_costs,
            # Fixed costs — Additional
            "Admin Overhead Override":   st.session_state.get("admin_overhead_override"),
            "Mfg Overhead Override":     st.session_state.get("mfg_overhead_override"),
            "Taxes Ins Override":        st.session_state.get("taxes_ins_override"),
            "Patents Roy Override":      st.session_state.get("patents_roy_override"),
            "Admin Ov Pct":              admin_ov_pct,
            "Mfg Ov Pct":               mfg_ov_pct,
            "Taxes Ins Pct":             taxes_ins_pct,
            "Patents Pct":               patents_pct,
            "AFC Pre Patents":           afc,
            "Direct Fixed Costs":        direct_fixed_costs,
            # Fixed costs — Indirect
            "Admin Costs Override":      st.session_state.get("admin_costs_override"),
            "Mfg Costs Override":        st.session_state.get("mfg_costs_override"),
            "Dist Selling Override":     st.session_state.get("dist_selling_override"),
            "R And D Override":          st.session_state.get("r_and_d_override"),
            "Admin Costs Pct":           admin_costs_pct,
            "Mfg Costs Pct":             mfg_costs_pct,
            "Dist Sell Pct":             dist_sell_pct,
            "R D Pct":                   r_d_pct,
            "OPEX":                      opex,
            "Indirect Fixed Costs":      indirect_fixed_costs,
            "Total Fixed Costs":         total_fixed_costs,
            "Total OPEX":                total_opex,
            # Working capital
            "WC Method":                 wc_method,
            "WC Pct":                    st.session_state.get("wc_pct", 5.0),
            "WC Equiv Cash Days":        st.session_state.get("wc_equiv_cash_days", 30.0),
            "WC Raw Mat Days":           st.session_state.get("wc_raw_mat_days", 15.0),
            "WC Accounts Rec Days":      st.session_state.get("wc_accounts_rec_days", 30.0),
            "WC Accrued Payroll Days":   st.session_state.get("wc_accrued_payroll_days", 30.0),
            "WC Accounts Pay Days":      st.session_state.get("wc_accounts_pay_days", 30.0),
            "Working Capital":           working_capital,
            # Startup costs
            "Startup Method":            startup_method,
            "Startup Single Pct":        st.session_state.get("startup_single_pct", 8.0),
            "Startup Op Training Days":  st.session_state.get("startup_op_training_days", 150.0),
            "Startup Commerc Pct":       st.session_state.get("startup_commerc_pct", 5.0),
            "Startup Inefficiency Pct":  st.session_state.get("startup_inefficiency_pct", 4.0),
            "Startup Costs":             startup_costs,
            # Total investment
            "Additional Costs":          additional_costs,
            "Total Investment":          total_investment,
            # TIC accuracy
            "TIC Lower Override":        st.session_state.get("tic_lower_override"),
            "TIC Upper Override":        st.session_state.get("tic_upper_override"),
            "TIC Lower Pct":             tic_lower_input,
            "TIC Upper Pct":             tic_upper_input,
            # Financial assumptions — Land
            "Land Option":               land_option,
            "Land Buy Pct":              land_buy_pct,
            "Land Rent Pct":             land_rent_pct,
            "Land Buy Pct Override":     st.session_state.get("land_buy_pct_override"),
            "Land Rent Pct Override":    st.session_state.get("land_rent_pct_override"),
            # Financial assumptions — Depreciation
            "Depreciation Method":       depreciation_method,
            "Depreciation Years":        depreciation_years,
            "Residual Value Pct":        residual_value_pct,
            "Residual Value USD":        residual_value_usd,
            "Depreciable Base":          depreciable_base,
            # Financial assumptions — Taxes
            "Tax Country":               tax_country,
            "Tax Rate Override":         st.session_state.get("tax_rate_override"),
            "Tax Rate":                  tax_rate,
            # Financial leverage
            "Financing Type":            financing_type,
            "Debt Ratio Pct":            debt_ratio_pct,
            "Debt Ratio":                debt_ratio,
            "Amortization Years":        amortization_years,
            "Grace Period Years":        grace_period_years,
            # MARR
            "Central Bank Rate":         central_bank_rate,
            "Credit Spread":             credit_spread,
            "COD":                       cod,
            "Unlevered Beta":            unlevered_beta,
            "Levered Beta":              levered_beta,
            "Market Return":             market_return,
            "Risk Free Rate":            risk_free_rate,
            "Market Risk Premium":       market_risk_premium,
            "Country Risk Premium":      country_risk_premium,
            "US CPI":                    us_cpi,
            "Country CPI":               country_cpi,
            "COE":                       coe,
            "MARR Override":             st.session_state.get("marr_override"),
            "MARR":                      marr,
            # Project lifetime
            "EPC Years":                 epc_years,
            "Project Lifetime":          project_lifetime,
            "Total Project Lifetime":    total_project_lifetime,
            # Capacity
            "Capacity First Year":       capacity_first_year,
            "Capacity Intermediate":     capacity_intermediate,
            "Capacity Last Year":        capacity_last_year,
            # Fixed costs per year
            "Fixed Costs First Year":    fixed_costs_first_year,
            "Fixed Costs Intermediate":  fixed_costs_intermediate,
            "Fixed Costs Last Year":     fixed_costs_last_year,
            # Market assumptions
            "Growth Main Price":         growth_main_price,
            "Growth Byproduct Price":    growth_byproduct_price,
            "Growth Raw Materials":      growth_raw_materials,
            "Growth Chem Utilities":     growth_chem_utilities,
            "Growth Fixed Costs":        growth_fixed_costs,
            # Other premises
            "Main Product Price":        main_product_price,
        }

        st.session_state.success_msg      = f"Scenario '{st.session_state.sn_input}' successfully saved!"
        st.session_state.clear_on_next_run = True
        st.rerun()

# ─────────────────────────────────────────────
# SCENARIO COMPARISON TABLE
# ─────────────────────────────────────────────
st.header("Compiled Scenarios")
if st.session_state.scenarios:
    summary_rows = [
        {
            "Scenario Name":    name,
            "Product":          d["Product Name"],
            "Eq. Cost":         d["Equipment Costs Source"],
            "Total Equip. Cost":fmt_curr(d["Total Equipment Costs"]),
            "ISBL+OSBL Cost":   fmt_curr(d["Project Costs ISBL+OSBL"]),
            "Project CAPEX":    fmt_curr(d.get("Project CAPEX", 0.0)),
            "Raw Mat. Cost":    fmt_curr(d.get("Total Raw Material Cost", 0.0)),
            "TRL":              d["TRL"],
        }
        for name, d in st.session_state.scenarios.items()
    ]
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)

    if st.button("Clear All Data", type="secondary"):
        st.session_state.scenarios = {}
        st.rerun()
else:
    st.info("No scenarios defined yet.")
