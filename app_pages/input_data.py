import streamlit as st
import pandas as pd
import io
import math
import openpyxl
from utils.constants import (
    ALLOWED_UNITS, PRODUCT_TYPES, TRL_OPTIONS, INFO_OPTIONS, SEVERITY_OPTIONS,
    MAT_OPTIONS, SIZE_OPTIONS, UNSCHED_FACTORS, LANG_FACTORS, LANG_FIELDS,
    PROJECT_CONTINGENCY, LAB_CHARGES, OFFICE_LABOR, MIN_SALARY,
    MAINTENANCE_REPAIRS, OPERATING_SUPPLIES, ADMIN_OVERHEAD, MFG_OVERHEAD,
    TAXES_INSURANCE, PATENTS_ROYALTIES, DIST_SELLING, R_AND_D,
    TIC_LOWER, TIC_UPPER, TAXES_BY_COUNTRY, COUNTRY_LIST,
    PLANT_COST_INDEX, PCI_YEARS, CAPEX_DISTRIBUTION,
    RATE_TO_PRICE_UNIT, RATE_UNITS, PRICE_UNITS,
    DEFAULTS, fmt_curr, get_pci, pci_escalate, price_unit_for,
)
from utils.ui import inject_css

inject_css()

# ══════════════════════════════════════════════════════════════════════════════
# UNIT CONVERSION SYSTEM  (met-C-Bar base: kg, m³, kWh, mol)
# ══════════════════════════════════════════════════════════════════════════════

# Supported rate units: "qty_unit/time_unit" or bare power unit
RATE_UNITS_FULL = [
    # Mass / time
    "g/s","g/min","g/h","g/day","g/year",
    "kg/s","kg/min","kg/h","kg/day","kg/year",
    "t/s","t/min","t/h","t/day","t/year",
    "lb/s","lb/min","lb/h","lb/day","lb/year",
    "oz/s","oz/min","oz/h","oz/day","oz/year",
    # Volume / time
    "mL/s","mL/min","mL/h","mL/day","mL/year",
    "L/s","L/min","L/h","L/day","L/year",
    "m3/s","m3/min","m3/h","m3/day","m3/year",
    "gal/s","gal/min","gal/h","gal/day","gal/year",
    # Energy / time
    "J/s","J/min","J/h","J/day","J/year",
    "kJ/s","kJ/min","kJ/h","kJ/day","kJ/year",
    "MJ/s","MJ/min","MJ/h","MJ/day","MJ/year",
    "kWh/s","kWh/min","kWh/h","kWh/day","kWh/year",
    "MWh/s","MWh/min","MWh/h","MWh/day","MWh/year",
    "BTU/s","BTU/min","BTU/h","BTU/day","BTU/year",
    # Molar / time
    "mol/s","mol/min","mol/h","mol/day","mol/year",
    "kmol/s","kmol/min","kmol/h","kmol/day","kmol/year",
    # Power (treated as energy rate → kWh/year)
    "W","kW","MW",
]

# Main-product capacity units (flow-rate style)
CAPACITY_UNITS = [
    "kg/h","kg/day","kg/year",
    "t/h","t/day","t/year",
    "g/h","g/day","g/year",
    "lb/h","lb/day","lb/year",
    "L/h","L/day","L/year",
    "m3/h","m3/day","m3/year",
    "mol/h","mol/day","mol/year",
    "kmol/h","kmol/day","kmol/year",
]

# Quantity conversion factors → base unit
_QTY_TO_BASE = {
    # mass → kg
    "g": 1e-3, "kg": 1.0, "t": 1e3, "lb": 0.453592, "oz": 0.0283495,
    # volume → m3
    "mL": 1e-6, "L": 1e-3, "m3": 1.0, "gal": 0.00378541,
    # energy → kWh
    "J": 1/3_600_000, "kJ": 1/3_600, "MJ": 1/3.6,
    "kWh": 1.0, "MWh": 1_000.0, "BTU": 1/3412.14,
    # molar → mol
    "mol": 1.0, "kmol": 1_000.0,
    # power → kW  (then ×h gives kWh)
    "W": 1e-3, "kW": 1.0, "MW": 1_000.0,
}

# Time conversion factors → per hour
_TIME_TO_PER_HOUR = {
    "s": 3600.0, "min": 60.0, "h": 1.0, "day": 1/24, "year": None,  # year handled specially
}

# Dimension of each quantity unit
_QTY_TO_DIM = {
    "g":"mass","kg":"mass","t":"mass","lb":"mass","oz":"mass",
    "mL":"volume","L":"volume","m3":"volume","gal":"volume",
    "J":"energy","kJ":"energy","MJ":"energy","kWh":"energy","MWh":"energy","BTU":"energy",
    "mol":"molar","kmol":"molar",
    "W":"power","kW":"power","MW":"power",
}

# Valid price units per dimension
_DIM_TO_PRICE_UNITS: dict[str, list[str]] = {
    "mass":   ["$/g", "$/kg", "$/t", "$/lb", "$/oz"],
    "volume": ["$/mL", "$/L", "$/m3", "$/gal"],
    "energy": ["$/J", "$/kJ", "$/MJ", "$/kWh", "$/MWh", "$/BTU"],
    "molar":  ["$/mol", "$/kmol"],
    "power":  ["$/kWh", "$/MWh", "$/J", "$/kJ", "$/MJ", "$/BTU"],  # power priced as energy delivered
}

# Default price unit per dimension
_DIM_TO_DEFAULT_PRICE_UNIT: dict[str, str] = {
    "mass": "$/kg", "volume": "$/m3", "energy": "$/kWh",
    "molar": "$/mol", "power": "$/kWh",
}

# Price unit qty label (for conversion factor lookup)
_PRICE_UNIT_TO_QTY = {
    "$/g": "g", "$/kg": "kg", "$/t": "t", "$/lb": "lb", "$/oz": "oz",
    "$/mL": "mL", "$/L": "L", "$/m3": "m3", "$/gal": "gal",
    "$/J": "J", "$/kJ": "kJ", "$/MJ": "MJ", "$/kWh": "kWh",
    "$/MWh": "MWh", "$/BTU": "BTU",
    "$/mol": "mol", "$/kmol": "kmol",
}

# Flat list of all price units
PRICE_UNITS_FULL = [u for units in _DIM_TO_PRICE_UNITS.values() for u in units]


def parse_rate_unit(rate_unit: str) -> tuple[str | None, str | None, str | None]:
    """
    Parse a rate_unit string into (qty_unit, time_unit, dimension).
    Handles bare power units (W, kW, MW) as special case.
    Returns (None, None, None) if unparseable.
    """
    if rate_unit in ("W", "kW", "MW"):
        return rate_unit, None, "power"
    if "/" not in rate_unit:
        return None, None, None
    qty, time = rate_unit.split("/", 1)
    dim = _QTY_TO_DIM.get(qty)
    if dim is None or time not in _TIME_TO_PER_HOUR:
        return None, None, None
    return qty, time, dim


def price_unit_for_rate(rate_unit: str) -> str:
    """Return the DEFAULT price unit for a given rate unit dimension."""
    _, _, dim = parse_rate_unit(rate_unit)
    return _DIM_TO_DEFAULT_PRICE_UNIT.get(dim, "$/kg")


def valid_price_units_for_rate(rate_unit: str) -> list[str]:
    """Return all valid price units for the dimension of a given rate unit."""
    _, _, dim = parse_rate_unit(rate_unit)
    return _DIM_TO_PRICE_UNITS.get(dim, ["$/kg"])


def price_per_base_unit(price: float, price_unit: str) -> float:
    """
    Convert a user price (e.g. 5 $/lb) to price per base unit (e.g. $/kg).
    Returns price_per_base = price / qty_to_base_factor_of_price_unit.
    Example: 5 $/lb → 5 / 0.453592 = 11.02 $/kg
    """
    qty_label = _PRICE_UNIT_TO_QTY.get(price_unit)
    if qty_label is None:
        return price  # unknown unit, pass through
    factor = _QTY_TO_BASE.get(qty_label, 1.0)
    return price / factor if factor else price


def annual_quantity(rate: float, rate_unit: str, working_hours: float) -> tuple[float, str]:
    """
    Convert a rate + rate_unit to annual quantity in the base unit.
    Returns (annual_qty_in_base_unit, base_unit_label).

    Formula:
      - Power units:  rate [kW] × working_hours [h/year] = kWh/year
      - /year units:  rate × qty_factor  (time already annual)
      - other:        rate × qty_factor × time_factor × working_hours
    """
    qty_u, time_u, dim = parse_rate_unit(rate_unit)

    if qty_u is None:
        return 0.0, "?"

    qty_factor = _QTY_TO_BASE.get(qty_u, 1.0)

    if dim == "power":
        # rate is in W/kW/MW → convert to kW, then × hours → kWh/year
        kw_factor = _QTY_TO_BASE[qty_u]   # W→0.001, kW→1, MW→1000
        return rate * kw_factor * working_hours, "kWh/year"

    base_labels = {
        "mass": "kg/year", "volume": "m3/year",
        "energy": "kWh/year", "molar": "mol/year",
    }
    label = base_labels.get(dim, "?/year")

    if time_u == "year":
        return rate * qty_factor, label

    time_factor = _TIME_TO_PER_HOUR[time_u]  # per-hour factor
    return rate * qty_factor * time_factor * working_hours, label


def annual_capacity(rate: float, cap_unit: str, working_hours: float) -> tuple[float, str]:
    """
    Convert main product capacity rate to annual quantity in base unit.
    cap_unit format: "qty/time"  e.g. "kg/h", "t/year", "kg/year"

    Returns (annual_value_in_base_unit, clean_annual_label).
    The label is always "<base_unit>/year" with no double /year.
    """
    qty, time, dim = parse_rate_unit(cap_unit)
    if qty is None:
        return 0.0, "?/year"

    qty_factor = _QTY_TO_BASE.get(qty, 1.0)
    base_labels = {
        "mass": "kg/year", "volume": "m3/year",
        "energy": "kWh/year", "molar": "mol/year",
    }
    label = base_labels.get(dim, "?/year")

    if time == "year":
        return rate * qty_factor, label          # already annual — no × working_hours
    time_factor = _TIME_TO_PER_HOUR[time]
    return rate * qty_factor * time_factor * working_hours, label


# ══════════════════════════════════════════════════════════════════════════════
# HARDCODED SCENARIOS  — add your benchmark cases here
# ══════════════════════════════════════════════════════════════════════════════

HARDCODED_SCENARIOS: dict[str, dict] = {
    # Example (replace with real data):
    # "Benchmark Case A": {
    #     "Product Name": "Ethanol",
    #     "Unit": "kg/year",
    #     "Capacity": 10_000_000.0,
    #     ...  (all keys from the save payload)
    # },
}


def _load_hardcoded_scenario(name: str):
    """Load a hardcoded scenario into session state (same logic as load_scenario_data)."""
    st.session_state.scenarios[name] = HARDCODED_SCENARIOS[name]
    st.session_state.sn_input = name
    load_scenario_data()
    st.success(f"Hardcoded scenario '{name}' loaded.")


# ══════════════════════════════════════════════════════════════════════════════
# PROGRESSIVE DISCLOSURE — section visibility state
# ══════════════════════════════════════════════════════════════════════════════

_SECTIONS = [
    "dma",               # 4 — unlocked by any source selected
    "aspen_pea",         # 5 — unlocked by Aspen PEA source
    "capex_equipment",   # 6 — unlocked by eq_cost_src set
    "capex_installation",# 7 — unlocked by oth_cost_src set
    "capex_indirect",    # 8 — unlocked by oth_cost_src set
    "capex_nonfield",    # 9 — unlocked by oth_cost_src set
    "capex_calculations",# 10 — unlocked by any cost source set
    "additional_info",   # 11 — unlocked by any cost source set
    "variable_costs",    # 12 — unlocked by basic info complete
    "fixed_costs",       # 13 — unlocked by CAPEX section visible
    "working_capital",   # 14 — unlocked by fixed costs visible
    "startup_costs",     # 15 — unlocked by working capital visible
    "financial_assumptions", # 16 — unlocked by basic info complete
    "project_lifetime",  # 17 — unlocked by financial assumptions visible
    "process_variables", # 18 — unlocked by scenario name filled
]

if "section_visible" not in st.session_state:
    # ALL sections start hidden — progressively unlocked by user actions
    st.session_state.section_visible = {s: False for s in _SECTIONS}


def _gate(section: str) -> bool:
    """Return True if section is currently shown (auto-unlocked OR manually toggled)."""
    return st.session_state.section_visible.get(section, False)


def _update_gates():
    """
    Auto-unlock sections based on current input state.
    Rule: only ever set True (unlock) — never force-close a section the user opened.
    Unlocking is one-way; toggling is controlled exclusively by _section_toggle().
    """
    sv = st.session_state.section_visible
    sn     = st.session_state.get("sn_input", "").strip()
    mp     = st.session_state.get("mp_input", "").strip()
    pc     = st.session_state.get("pc_input", None)
    eq_src = st.session_state.get("eq_cost_src", "")
    oth_src= st.session_state.get("oth_cost_src", "")

    basic_ok  = bool(mp and pc)
    eq_set    = bool(eq_src)
    oth_set   = bool(oth_src)
    any_src   = eq_set or oth_set
    any_pea   = "Aspen PEA" in (eq_src, oth_src)
    capex_vis = sv.get("capex_calculations", False)
    fc_vis    = sv.get("fixed_costs", False)
    wc_vis    = sv.get("working_capital", False)
    fin_vis   = sv.get("financial_assumptions", False)

    # Sequential unlock — only set True, never override user manual toggle to False
    if any_src:
        sv["dma"] = True
    if any_pea:
        sv["aspen_pea"] = True
    if eq_set:
        sv["capex_equipment"] = True
    if oth_set:
        sv["capex_installation"] = True
        sv["capex_indirect"] = True
        sv["capex_nonfield"] = True
    if any_src:
        sv["capex_calculations"] = True
        sv["additional_info"] = True
    if basic_ok:
        sv["variable_costs"] = True
        sv["financial_assumptions"] = True
    # Fixed costs unlocked once CAPEX calculations section is visible
    if sv.get("capex_calculations", False) and basic_ok:
        sv["fixed_costs"] = True
    # Working capital unlocked once fixed costs visible
    if sv.get("fixed_costs", False):
        sv["working_capital"] = True
    # Startup unlocked once working capital visible
    if sv.get("working_capital", False):
        sv["startup_costs"] = True
    # Project lifetime unlocked once financial assumptions visible
    if sv.get("financial_assumptions", False):
        sv["project_lifetime"] = True
    # Process variables: unlocked when scenario name is filled
    if sn:
        sv["process_variables"] = True


def _section_toggle(section: str, label: str) -> bool:
    """
    Render a section header with a Show/Hide toggle button.
    Always available — user can open/close any section at any time.
    Returns True if the section content should be rendered.
    """
    sv = st.session_state.section_visible
    current = sv.get(section, False)
    col1, col2 = st.columns([8, 1])
    with col1:
        st.markdown(f"#### {label}")
    with col2:
        btn_label = "▲ Hide" if current else "▼ Show"
        if st.button(btn_label, key=f"toggle_{section}", use_container_width=True):
            sv[section] = not current
            st.rerun()
    return sv.get(section, False)


# ══════════════════════════════════════════════════════════════════════════════
# EXISTING HELPERS  (unchanged from original)
# ══════════════════════════════════════════════════════════════════════════════

def _result_row(label: str, value: float, key: str, formula: str = ""):
    c1, c2 = st.columns([3, 2])
    with c1:
        if formula:
            st.markdown(
                f'**{label}** <span title="{formula}" '
                f'style="cursor:help;color:#888;font-size:0.85rem;">ⓘ</span>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(f"**{label}**")
    with c2:
        st.text_input(key, value=fmt_curr(value), disabled=True,
                      label_visibility="collapsed")


def parse_aspen_pea(ipewb_bytes, reports_bytes, import_equip, import_other):
    result = {}
    if ipewb_bytes and import_equip:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(ipewb_bytes), data_only=True)
            sheets = wb.sheetnames
            if "Equipment Summary" in sheets:
                fmt = "new"
            elif "Equipment" in sheets:
                fmt = "legacy"
            else:
                fmt = None

            if fmt == "new":
                ws_eq = wb["Equipment Summary"]
                equip_acq = sum(r[7] for r in ws_eq.iter_rows(min_row=1, values_only=True) if isinstance(r[7], (int, float)))
                result["equip_acq"] = float(equip_acq)
                ws_ps = wb["Project Summary"]
                ops = ws_ps.cell(row=146, column=3).value
                sups = ws_ps.cell(row=155, column=3).value
                epc_weeks = ws_ps.cell(row=49, column=3).value
                if epc_weeks:
                    result["epc_years"] = max(1, round(epc_weeks * 7 / 365))
                if import_other:
                    eq_set = ws_ps.cell(row=102, column=3).value
                    if isinstance(eq_set, (int, float)):
                        result["equipment_setting"] = float(eq_set)
            elif fmt == "legacy":
                ws_eq = wb["Equipment"]
                equip_acq = sum(r[4] for r in ws_eq.iter_rows(min_row=1, values_only=True) if isinstance(r[4], (int, float)))
                result["equip_acq"] = float(equip_acq)
                ws_ps = wb["Project Summary"]
                ops  = ws_ps.cell(row=225, column=3).value
                sups = ws_ps.cell(row=234, column=3).value
                epc_weeks = ws_ps.cell(row=62, column=3).value
                if epc_weeks:
                    result["epc_years"] = max(1, round(epc_weeks * 7 / 365))
                if import_other:
                    eq_set = ws_ps.cell(row=173, column=3).value
                    if isinstance(eq_set, (int, float)):
                        result["equipment_setting"] = float(eq_set)

            if fmt in ("new", "legacy"):
                if isinstance(ops, (int, float)):
                    result["n_operators"] = max(1, int(ops))
                if isinstance(sups, (int, float)):
                    result["n_supervisors"] = max(1, int(sups))
        except Exception as e:
            result["_ipewb_error"] = str(e)

    if reports_bytes and import_other:
        try:
            wb_r = openpyxl.load_workbook(io.BytesIO(reports_bytes), data_only=True)
            ws = wb_r["Proj Cost Sumry"]
            def _v(row, col):
                v = ws.cell(row=row, column=col).value
                return float(v) if isinstance(v, (int, float)) else 0.0
            reports_total_equip = _v(8, 5)
            ipewb_equip_acq = result.get("equip_acq", 0.0)
            result["spare_parts"] = max(0.0, reports_total_equip - ipewb_equip_acq)
            result["piping"]           = _v(9,  6)
            result["civil"]            = _v(10, 6)
            result["steel"]            = _v(11, 6)
            result["instrumentals"]    = _v(12, 6)
            result["electrical"]       = _v(13, 6)
            result["insulation"]       = _v(14, 6)
            result["paint"]            = _v(15, 6)
            result["field_office_staff"]       = _v(7,  13)
            result["construction_indirects"]   = _v(18, 13)
            result["freight"]              = _v(6,  18)
            result["taxes_and_permits"]    = _v(10, 18)
            result["engineering_and_ho"]   = _v(16, 18) + _v(19, 18)
            result["ga_overheads"]         = _v(20, 18)
            result["contract_fee"]         = _v(21, 18)
        except Exception as e:
            result["_reports_error"] = str(e)

    return result


def reset_state(keys: dict = DEFAULTS):
    for k, v in keys.items():
        st.session_state[k] = v
    for k in list(st.session_state.keys()):
        if k.startswith("val_") or k.startswith("w_"):
            del st.session_state[k]
    st.session_state.table_key = st.session_state.get("table_key", 0) + 1


def _reseed_lang_fields(equip_acq: float):
    idx = 0 if st.session_state.lang_utility else 1
    for field, factors in LANG_FACTORS.items():
        key = f"lf_{field.lower().replace(' ', '_')}"
        st.session_state[key] = equip_acq * factors[idx]
    st.session_state.lang_seeded_acq = equip_acq


def lang_val(field: str, equip_acq: float) -> float:
    idx = 0 if st.session_state.lang_utility else 1
    return equip_acq * LANG_FACTORS[field][idx]


def _seed_override_values():
    acq = st.session_state.get("equip_acq", 0.0)
    idx = 0 if st.session_state.get("lang_utility", False) else 1
    for field, factors in LANG_FACTORS.items():
        vkey = f"val_{field.lower().replace(' ', '_')}"
        st.session_state[vkey] = acq * factors[idx]


def lang_or_manual(field: str, label: str, equip_acq: float, step: float = 100.0) -> float:
    ss_key  = field.lower().replace(" ", "_")
    val_key = f"val_{ss_key}"
    w_key   = f"w_{ss_key}"

    if st.session_state.oth_cost_src == "Lang Factors":
        lang_computed = lang_val(field, equip_acq)
        allow = st.session_state.get("allow_override", False)

        if not allow:
            st.text_input(f"{label}", value=fmt_curr(lang_computed), disabled=True)
            return lang_computed

        stored = st.session_state.get(val_key, lang_computed)
        is_overridden = abs(stored - lang_computed) > 0.005
        color = "#e67e00" if is_overridden else "#2e7d32"
        hint  = " ⚠ overridden" if is_overridden else " ✓ Lang"
        st.markdown(
            f'<p style="margin-bottom:0px; font-size:0.85rem; color:{color};">'
            f'<b>{label} ($)</b>{hint}</p>',
            unsafe_allow_html=True,
        )
        def _sync(vk=val_key, wk=w_key):
            st.session_state[vk] = st.session_state[wk]
        result = st.number_input(
            label, min_value=0.0, step=step, value=stored,
            key=w_key, label_visibility="collapsed", on_change=_sync,
        )
        return result

    return st.number_input(f"{label} ($)", min_value=0.0, step=step, key=ss_key)


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE INIT
# ══════════════════════════════════════════════════════════════════════════════

if "scenarios" not in st.session_state:
    st.session_state.scenarios = {}
if "table_key" not in st.session_state:
    st.session_state.table_key = 0

for key, default in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = default

# Seed hardcoded scenarios into the scenarios dict on first load
for hname, hdata in HARDCODED_SCENARIOS.items():
    if hname not in st.session_state.scenarios:
        st.session_state.scenarios[hname] = hdata

# Database overrides
def _db(name, default):
    return st.session_state.get(f"db_{name}", default)

UNSCHED_FACTORS     = _db("UNSCHED_FACTORS",     UNSCHED_FACTORS)
LANG_FACTORS        = _db("LANG_FACTORS",         LANG_FACTORS)
PROJECT_CONTINGENCY = _db("PROJECT_CONTINGENCY",  PROJECT_CONTINGENCY)
LAB_CHARGES         = _db("LAB_CHARGES",          LAB_CHARGES)
OFFICE_LABOR        = _db("OFFICE_LABOR",         OFFICE_LABOR)
MAINTENANCE_REPAIRS = _db("MAINTENANCE_REPAIRS",  MAINTENANCE_REPAIRS)
OPERATING_SUPPLIES  = _db("OPERATING_SUPPLIES",   OPERATING_SUPPLIES)
ADMIN_OVERHEAD      = _db("ADMIN_OVERHEAD",       ADMIN_OVERHEAD)
MFG_OVERHEAD        = _db("MFG_OVERHEAD",         MFG_OVERHEAD)
TAXES_INSURANCE     = _db("TAXES_INSURANCE",      TAXES_INSURANCE)
PATENTS_ROYALTIES   = _db("PATENTS_ROYALTIES",    PATENTS_ROYALTIES)
DIST_SELLING        = _db("DIST_SELLING",         DIST_SELLING)
R_AND_D             = _db("R_AND_D",              R_AND_D)
TIC_LOWER           = _db("TIC_LOWER",            TIC_LOWER)
TIC_UPPER           = _db("TIC_UPPER",            TIC_UPPER)
TAXES_BY_COUNTRY    = _db("TAXES_BY_COUNTRY",     TAXES_BY_COUNTRY)
COUNTRY_LIST        = sorted(TAXES_BY_COUNTRY.keys())

for key, default in DEFAULTS.items():
    if key not in st.session_state:
        st.session_state[key] = default

if st.session_state.pop("clear_on_next_run", False):
    reset_state()

if st.session_state.get("success_msg"):
    st.success(st.session_state.pop("success_msg"))


# ══════════════════════════════════════════════════════════════════════════════
# LOAD SCENARIO CALLBACK
# ══════════════════════════════════════════════════════════════════════════════

def load_scenario_data():
    sn = st.session_state.sn_input
    data = st.session_state.scenarios.get(sn, {})

    mapping = {
        "mp_input":        ("Product Name",            ""),
        "pu_input":        ("Unit",                    "kg/h"),
        "pc_input":        ("Capacity",                None),
        "eq_cost_src":     ("Equipment Costs Source",  ""),
        "oth_cost_src":    ("Other Costs Source",      ""),
        "lang_utility":    ("Contains Utility Systems",False),
        "dm_prod_type":    ("Product Type",            "Basic Chemical"),
        "dm_trl":          ("TRL",                     "Industrial (8 or 9)"),
        "dm_info_avail":   ("Info Availability",       "Medium"),
        "dm_severity":     ("Process Severity",        "Medium"),
        "dm_mat_type":     ("Material Handled",        "Fluids"),
        "dm_plant_size":   ("Plant Size",              "Medium"),
        "equip_acq":       ("Equipment Acquisition",   0.0),
        "spare_parts":     ("Spare Parts",             0.0),
        "equipment_setting": ("Equipment Setting",     0.0),
        "piping":          ("Piping",                  0.0),
        "civil":           ("Civil",                   0.0),
        "steel":           ("Steel",                   0.0),
        "instrumentals":   ("Instrumentals",           0.0),
        "electrical":      ("Electrical",              0.0),
        "insulation":      ("Insulation",              0.0),
        "paint":           ("Paint",                   0.0),
        "isbl_contrib":    ("ISBL Contribution (%)",   100.0),
        "field_office_staff":      ("Field Office Staff",      0.0),
        "construction_indirects":  ("Construction Indirects",  0.0),
        "freight":         ("Freight",                 0.0),
        "taxes_and_permits": ("Taxes and Permits",     0.0),
        "engineering_and_ho": ("Engineering and HO",  0.0),
        "ga_overheads":    ("GA Overheads",            0.0),
        "contract_fee":    ("Contract Fee",            0.0),
        "allow_override":  ("Allow Override",          False),
        "databank_year":   ("Databank Year",           2022),
        "analysis_year":   ("Year of Analysis",        PCI_YEARS[-1]),
        "proj_location":   ("Project Location",        "Brazil"),
        "location_factor": ("Location Factor",         0.97),
        "working_hours":   ("Working Hours per Year",  8000.0),
        "scaling_factor":  ("Scaling Factor",          0.6),
        "n_operators":               ("Num Operators",            2),
        "operator_salary":           ("Operator Salary",          1247.75),
        "n_supervisors":             ("Num Supervisors",          1),
        "supervisor_salary":         ("Supervisor Salary",        1660.155),
        "salary_charges":            ("Salary Charges",           2.2),
        "plant_daily_hours":         ("Plant Daily Hours",        24.0),
        "weekly_op_days":            ("Weekly Op Days",           7.0),
        "worker_hours_shift":        ("Worker Hours per Shift",   8.0),
        "worker_shifts_week":        ("Worker Shifts per Week",   5.0),
        "worker_vacation_weeks":     ("Worker Vacation Weeks",    4.0),
        "lab_charges_override":      ("Lab Charges Override",     None),
        "office_labor_override":     ("Office Labor Override",    None),
        "labor_working_hrs_override":("Labor Working Hrs Override", None),
        "maint_repair_override":     ("Maint Repair Override",    None),
        "op_supplies_override":      ("Op Supplies Override",     None),
        "admin_overhead_override":   ("Admin Overhead Override",  None),
        "mfg_overhead_override":     ("Mfg Overhead Override",    None),
        "taxes_ins_override":        ("Taxes Ins Override",       None),
        "patents_roy_override":      ("Patents Roy Override",     None),
        "admin_costs_override":      ("Admin Costs Override",     None),
        "mfg_costs_override":        ("Mfg Costs Override",       None),
        "dist_selling_override":     ("Dist Selling Override",    None),
        "r_and_d_override":          ("R And D Override",         None),
        "wc_method":                 ("WC Method",                "Percentage"),
        "wc_pct":                    ("WC Pct",                   5.0),
        "wc_equiv_cash_days":        ("WC Equiv Cash Days",       30.0),
        "wc_raw_mat_days":           ("WC Raw Mat Days",          15.0),
        "wc_accounts_rec_days":      ("WC Accounts Rec Days",     30.0),
        "wc_accrued_payroll_days":   ("WC Accrued Payroll Days",  30.0),
        "wc_accounts_pay_days":      ("WC Accounts Pay Days",     30.0),
        "startup_method":            ("Startup Method",           "Multiple Factors"),
        "startup_single_pct":        ("Startup Single Pct",       8.0),
        "startup_op_training_days":  ("Startup Op Training Days", 150.0),
        "startup_commerc_pct":       ("Startup Commerc Pct",      5.0),
        "startup_inefficiency_pct":  ("Startup Inefficiency Pct", 4.0),
        "additional_costs":          ("Additional Costs",         0.0),
        "tic_lower_override":        ("TIC Lower Override",       None),
        "tic_upper_override":        ("TIC Upper Override",       None),
        "land_option":               ("Land Option",              "Buy"),
        "land_buy_pct_override":     ("Land Buy Pct Override",    None),
        "land_rent_pct_override":    ("Land Rent Pct Override",   None),
        "depreciation_method":       ("Depreciation Method",      "Straight Line"),
        "depreciation_years":        ("Depreciation Years",       10),
        "residual_value_pct":        ("Residual Value Pct",       20.0),
        "tax_country":               ("Tax Country",              "Brazil"),
        "tax_rate_override":         ("Tax Rate Override",        None),
        "financing_type":            ("Financing Type",           "None"),
        "debt_ratio_pct":            ("Debt Ratio Pct",           50.0),
        "amortization_years":        ("Amortization Years",       13),
        "grace_period_years":        ("Grace Period Years",       5),
        "central_bank_rate":         ("Central Bank Rate",        5.45),
        "credit_spread":             ("Credit Spread",            2.94),
        "unlevered_beta":            ("Unlevered Beta",           1.0),
        "market_return":             ("Market Return",            8.63),
        "risk_free_rate":            ("Risk Free Rate",           1.94),
        "country_risk_premium":      ("Country Risk Premium",     3.63),
        "us_cpi":                    ("US CPI",                   2.46),
        "country_cpi":               ("Country CPI",              4.65),
        "marr_override":             ("MARR Override",            None),
        "epc_years":                 ("EPC Years",                3),
        "project_lifetime":          ("Project Lifetime",         20),
        "capacity_first_year":       ("Capacity First Year",      100.0),
        "capacity_intermediate":     ("Capacity Intermediate",    100.0),
        "capacity_last_year":        ("Capacity Last Year",       100.0),
        "fixed_costs_first_year":    ("Fixed Costs First Year",   100.0),
        "fixed_costs_intermediate":  ("Fixed Costs Intermediate", 100.0),
        "fixed_costs_last_year":     ("Fixed Costs Last Year",    100.0),
        "growth_main_price":         ("Growth Main Price",        0.0),
        "growth_byproduct_price":    ("Growth Byproduct Price",   0.0),
        "growth_raw_materials":      ("Growth Raw Materials",     0.0),
        "growth_chem_utilities":     ("Growth Chem Utilities",    0.0),
        "growth_fixed_costs":        ("Growth Fixed Costs",       0.0),
        "main_product_price":        ("_reset_", None),
    }

    for ss_key, (data_key, default) in mapping.items():
        if data_key == "_reset_":
            st.session_state[ss_key] = None
        else:
            st.session_state[ss_key] = data.get(data_key, default)

    st.session_state.table_key += 1
    st.session_state.lang_seeded_acq = None

    # Re-evaluate section visibility for the loaded scenario
    _update_gates()


# ══════════════════════════════════════════════════════════════════════════════
# DMA CALLBACK
# ══════════════════════════════════════════════════════════════════════════════

_DMA_DERIVED_OVERRIDES = [
    "op_supplies_override", "maint_repair_override",
    "admin_overhead_override", "mfg_overhead_override", "taxes_ins_override",
    "patents_roy_override", "lab_charges_override", "office_labor_override",
    "admin_costs_override", "mfg_costs_override",
    "dist_selling_override", "r_and_d_override",
    "tic_lower_override", "tic_upper_override",
]

def _on_dma_change():
    prod = st.session_state.get("dm_prod_type", "Basic Chemical")
    trl  = st.session_state.get("dm_trl", "Industrial (8 or 9)")
    info = st.session_state.get("dm_info_avail", "Medium")
    sev  = st.session_state.get("dm_severity", "Medium")
    mat  = st.session_state.get("dm_mat_type", "Fluids")
    office_pct_frac = OFFICE_LABOR.get(prod, 0.10)
    maint_pct_frac  = MAINTENANCE_REPAIRS.get((mat, prod), 0.01)

    _new_refs = {
        "lab_charges_override":      LAB_CHARGES.get(prod, 0.10) * 100.0,
        "office_labor_override":     OFFICE_LABOR.get(prod, 0.10) * 100.0,
        "maint_repair_override":     MAINTENANCE_REPAIRS.get((mat, prod), 0.01) * 100.0,
        "op_supplies_override":      OPERATING_SUPPLIES.get(sev, 0.0015) * 100.0,
        "admin_overhead_override":   ADMIN_OVERHEAD.get(prod, 0.50) * 100.0 * (1 + office_pct_frac),
        "mfg_overhead_override":     MFG_OVERHEAD.get(sev, 0.006) * 100.0,
        "taxes_ins_override":        TAXES_INSURANCE.get(sev, 0.032) * 100.0,
        "patents_roy_override":      (PATENTS_ROYALTIES.get((trl, prod)) or 0.0) * 100.0,
        "admin_costs_override":      (1.0 + office_pct_frac) * 0.15 * 100.0,
        "mfg_costs_override":        maint_pct_frac * 0.15 * 100.0,
        "dist_selling_override":     DIST_SELLING.get(prod, 0.08) * 100.0,
        "r_and_d_override":          R_AND_D.get((trl, prod), 0.02) * 100.0,
    }
    tic_lo = TIC_LOWER.get((trl, info))
    tic_hi = TIC_UPPER.get((trl, info))
    if tic_lo is not None:
        _new_refs["tic_lower_override"] = tic_lo
    if tic_hi is not None:
        _new_refs["tic_upper_override"] = tic_hi

    for ok in _DMA_DERIVED_OVERRIDES:
        st.session_state[ok] = None
        wk = f"w_{ok}"
        if ok in _new_refs:
            st.session_state[wk] = float(_new_refs[ok])
        else:
            st.session_state.pop(wk, None)


# ══════════════════════════════════════════════════════════════════════════════
# DEV SIDEBAR — Hardcoded scenarios
# ══════════════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown("### 🔧 Dev Tools")

    # ── Load hardcoded scenarios ──────────────────────────────────────────
    if HARDCODED_SCENARIOS:
        st.markdown("**Load hardcoded scenario:**")
        for hname in HARDCODED_SCENARIOS:
            if st.button(f"📥 {hname}", key=f"hc_{hname}", use_container_width=True):
                _load_hardcoded_scenario(hname)
                st.rerun()
    else:
        st.caption("No hardcoded scenarios yet. Use the button below to export the current scenario.")

    st.markdown("---")

    # ── Export current scenario as hardcoded ──────────────────────────────
    st.markdown("**Export scenario as hardcoded:**")
    _export_sn = st.session_state.get("sn_input", "").strip()
    if _export_sn and _export_sn in st.session_state.scenarios:
        _export_data = st.session_state.scenarios[_export_sn]

        # JSON download
        import json as _json
        _json_bytes = _json.dumps({_export_sn: _export_data}, indent=4, default=str).encode()
        st.download_button(
            label=f"💾 Download '{_export_sn}' as JSON",
            data=_json_bytes,
            file_name=f"{_export_sn.replace(' ', '_')}_hardcoded.json",
            mime="application/json",
            use_container_width=True,
        )

        # Python snippet
        if st.button("📋 Show Python snippet", key="show_snippet", use_container_width=True):
            st.session_state["show_hc_snippet"] = not st.session_state.get("show_hc_snippet", False)

        if st.session_state.get("show_hc_snippet", False):
            import pprint as _pprint
            _snippet = f"HARDCODED_SCENARIOS[{_export_sn!r}] = " + _pprint.pformat(_export_data, width=80)
            st.code(_snippet, language="python")
            st.caption("Copy the above and paste into HARDCODED_SCENARIOS in input_data.py")
    else:
        st.caption("Save a scenario first, then export it here.")

    st.markdown("---")

    # ── Section visibility reset ──────────────────────────────────────────
    st.markdown("**Section visibility:**")
    if st.button("Reset all sections to auto", use_container_width=True):
        st.session_state.section_visible = {s: False for s in _SECTIONS}
        _update_gates()
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PAGE HEADER
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("### Scenario Configuration")
st.caption("Define parameters for your chemical plant scenario. "
           "Sections unlock progressively as you fill in prerequisites — "
           "or expand any section manually using the Show/Hide buttons.")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Scenario Name  (always visible)
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("#### 1. Scenario Name")

existing_scenarios = list(st.session_state.scenarios.keys())
scenario_options = ["➕ New scenario"] + existing_scenarios

# Determine current selectbox index
_current_sn = st.session_state.get("sn_input", "")
if _current_sn and _current_sn in existing_scenarios:
    _sel_idx = existing_scenarios.index(_current_sn) + 1  # +1 for the "New" option
else:
    _sel_idx = 0

col_sel, col_txt = st.columns([2, 3])
with col_sel:
    _scenario_sel = st.selectbox(
        "Select existing scenario",
        options=scenario_options,
        index=_sel_idx,
        key="scenario_selector",
        label_visibility="collapsed",
    )

# Handle selection change
if _scenario_sel != "➕ New scenario" and _scenario_sel != st.session_state.get("_last_sel"):
    st.session_state["_last_sel"] = _scenario_sel
    st.session_state.sn_input = _scenario_sel
    load_scenario_data()
    st.rerun()
elif _scenario_sel == "➕ New scenario":
    st.session_state["_last_sel"] = _scenario_sel

with col_txt:
    scenario_name = st.text_input(
        "Or type new scenario name",
        key="sn_input",
        on_change=load_scenario_data,
        placeholder="Type a new scenario name...",
        label_visibility="collapsed",
    )

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — Basic Information  (always visible)
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("#### 2. Basic Information")
col1, col2 = st.columns(2)
with col1:
    st.text_input("Main Product Name", key="mp_input")
with col2:
    # Capacity unit selector — drives annual_capacity() extrapolation
    cap_unit = st.selectbox(
        "Capacity unit",
        options=CAPACITY_UNITS,
        index=CAPACITY_UNITS.index(st.session_state.get("pu_input", "kg/h"))
              if st.session_state.get("pu_input", "kg/h") in CAPACITY_UNITS else 0,
        key="pu_input",
    )
    cap_rate = st.number_input(
        f"Main product capacity ({cap_unit})",
        min_value=0.0, step=1.0, key="pc_input",
    )

# Show derived annual capacity
_wh_preview = st.session_state.get("working_hours", 8000.0)
if cap_rate and cap_rate > 0:
    _ann_cap, _ann_unit = annual_capacity(cap_rate, cap_unit, _wh_preview)
    st.caption(f"≈ **{_ann_cap:,.1f} {_ann_unit}** at {_wh_preview:,.0f} h/year operating hours")

# Update gates after basic info is filled
_update_gates()
st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — Investment Cost Sources  (always visible, starts empty)
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("#### 3. Investment Cost Sources")
st.caption("Choose data sources to unlock the corresponding CAPEX input sections below.")

col_eq, col_oth = st.columns(2)
with col_eq:
    _eq_options = ["", "Manual Input", "Aspen PEA"]
    _eq_idx = _eq_options.index(st.session_state.get("eq_cost_src", "")) \
              if st.session_state.get("eq_cost_src", "") in _eq_options else 0
    eq_cost_src = st.selectbox(
        "Equipment costs source",
        options=_eq_options,
        index=_eq_idx,
        key="eq_cost_src",
        format_func=lambda x: "— select —" if x == "" else x,
        on_change=_update_gates,
    )

with col_oth:
    _oth_options = ["", "Manual Input", "Aspen PEA", "Lang Factors"]
    _oth_idx = _oth_options.index(st.session_state.get("oth_cost_src", "")) \
               if st.session_state.get("oth_cost_src", "") in _oth_options else 0
    oth_cost_src = st.selectbox(
        "Other Costs source",
        options=_oth_options,
        index=_oth_idx,
        key="oth_cost_src",
        format_func=lambda x: "— select —" if x == "" else x,
        on_change=_update_gates,
    )
    if st.session_state.oth_cost_src == "Lang Factors":
        st.checkbox("Contains utility systems?", key="lang_utility")

_eq_is_pea  = st.session_state.eq_cost_src  == "Aspen PEA"
_oth_is_pea = st.session_state.oth_cost_src == "Aspen PEA"

_update_gates()
st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — Process Variables  (gated: scenario name filled)
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("process_variables", "18. Process Variables"):
    saved_vars = st.session_state.scenarios.get(scenario_name, {}).get("Process Variables", [])
    df_vars = pd.DataFrame(saved_vars) if saved_vars else pd.DataFrame(columns=["Variable Name", "Unit", "Value"])
    edited_process_vars = st.data_editor(
        df_vars, key=f"editor_{st.session_state.table_key}", num_rows="dynamic",
        use_container_width=True, hide_index=True,
        column_config={
            "Variable Name": st.column_config.TextColumn(required=True),
            "Unit":          st.column_config.TextColumn(required=True),
            "Value":         st.column_config.NumberColumn(required=True),
        },
    )
else:
    # Preserve last saved process vars when section is hidden
    _saved = st.session_state.scenarios.get(scenario_name, {}).get("Process Variables", [])
    edited_process_vars = pd.DataFrame(_saved) if _saved else pd.DataFrame(columns=["Variable Name", "Unit", "Value"])

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — Decision Making Assistant  (gated: basic info complete)
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("dma", "4. Decision Making Assistant"):
    col_dm1, col_dm2, col_dm3 = st.columns(3)
    with col_dm1:
        st.selectbox("Type of main product",        PRODUCT_TYPES,    key="dm_prod_type", on_change=_on_dma_change)
        st.selectbox("Availability of information", INFO_OPTIONS,     key="dm_info_avail", on_change=_on_dma_change)
    with col_dm2:
        st.selectbox("TRL",                         TRL_OPTIONS,      key="dm_trl",       on_change=_on_dma_change)
        st.selectbox("Process severity",            SEVERITY_OPTIONS, key="dm_severity",  on_change=_on_dma_change)
    with col_dm3:
        st.selectbox("Type of material handled",    MAT_OPTIONS,      key="dm_mat_type",  on_change=_on_dma_change)
        st.selectbox("Plant size",                  SIZE_OPTIONS,     key="dm_plant_size")

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — Aspen PEA Import  (gated: any source = Aspen PEA)
# ══════════════════════════════════════════════════════════════════════════════

if _gate("aspen_pea") and (_eq_is_pea or _oth_is_pea):
    st.markdown("---")
    if _section_toggle("aspen_pea", "5. Aspen PEA File Import"):
        st.caption(
            "Upload the IPEWB (.xlsx) and/or Reports (.xlsm) files. "
            "Auto-detects Aspen PEA version. Imported values become editable defaults."
        )
        col_up1, col_up2 = st.columns(2)
        with col_up1:
            ipewb_file = st.file_uploader("IPEWB file (.xlsx)", type=["xlsx"], key="ipewb_uploader") if _eq_is_pea else None
        with col_up2:
            reports_file = st.file_uploader("Reports file (.xlsm / .xlsx)", type=["xlsm", "xlsx"], key="reports_uploader") if _oth_is_pea else None

        _ipewb_fp   = f"{ipewb_file.name}:{ipewb_file.size}"   if ipewb_file   else ""
        _reports_fp = f"{reports_file.name}:{reports_file.size}" if reports_file else ""
        _current_fp = f"{_ipewb_fp}|{_reports_fp}"
        _last_fp    = st.session_state.get("pea_last_fingerprint", "")

        if _current_fp != _last_fp and (_ipewb_fp or _reports_fp):
            _ipewb_bytes   = ipewb_file.read()   if ipewb_file   else None
            _reports_bytes = reports_file.read() if reports_file else None
            with st.spinner("Reading Aspen PEA files…"):
                _parsed = parse_aspen_pea(_ipewb_bytes, _reports_bytes, import_equip=_eq_is_pea, import_other=_oth_is_pea)
            _errs = {k: v for k, v in _parsed.items() if k.startswith("_") and k.endswith("_error")}
            if _errs:
                for e in _errs.values():
                    st.error(f"Import error: {e}")
            else:
                _skip = {"fmt", "_ipewb_error", "_reports_error"}
                for k, v in _parsed.items():
                    if k not in _skip:
                        st.session_state[k] = v
                st.session_state["pea_last_fingerprint"] = _current_fp
                st.session_state["pea_last_parsed"] = {k: v for k, v in _parsed.items() if k not in _skip}
                st.session_state["pea_fmt"] = _parsed.get("fmt", "unknown")

        if st.session_state.get("pea_last_parsed") and _current_fp == st.session_state.get("pea_last_fingerprint", ""):
            _p = st.session_state["pea_last_parsed"]
            _fmt_label = {"legacy": "Legacy (Equipment sheet)", "new": "New (Equipment Summary sheet)"}.get(
                st.session_state.get("pea_fmt", ""), "Unknown")
            st.success(f"✓ Aspen PEA import complete — format: **{_fmt_label}**")

    st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — Project CAPEX  (equipment gated on eq_cost_src, rest on oth_cost_src)
# ══════════════════════════════════════════════════════════════════════════════

_overridable_number_defined = False  # flag — defined below once

def _overridable_number(label, ref_val, override_key, step=0.01, min_value=None):
    stored = st.session_state.get(override_key)
    wk = f"w_{override_key}"
    current = stored if stored is not None else ref_val
    is_overridden = stored is not None and abs(stored - ref_val) > 1e-9
    if is_overridden:
        st.markdown(f'<p style="margin-bottom:0px;font-size:0.85rem;color:#e67e00;"><b>{label}</b> ⚠ overridden</p>', unsafe_allow_html=True)
    else:
        st.markdown(f'<p style="margin-bottom:0px;font-size:0.85rem;"><b>{label}</b></p>', unsafe_allow_html=True)
    _ref_key = f"_ref_{override_key}"
    st.session_state[_ref_key] = ref_val
    def _on_change(ok=override_key, rk_key=_ref_key, wk=wk):
        rk = st.session_state.get(rk_key, 0.0)
        v = st.session_state.get(wk, rk)
        st.session_state[ok] = v if abs(v - rk) > 1e-9 else None
    return st.number_input(label, min_value=min_value, step=step, value=float(current),
                           key=wk, label_visibility="collapsed", on_change=_on_change)


is_lang = st.session_state.oth_cost_src == "Lang Factors"

# ── 7.1 Equipment Costs ───────────────────────────────────────────────────────
if _section_toggle("capex_equipment", "6. Equipment Costs"):
    if is_lang:
        st.checkbox("Allow manual override of Lang factor fields?", key="allow_override",
                    on_change=_seed_override_values,
                    help="Fields that differ from Lang value are highlighted in amber.")
    else:
        st.session_state.allow_override = False

    col1, col2, col3 = st.columns(3)
    with col1:
        equip_acq = st.number_input("Equipment Acquisition ($)", min_value=0.0, step=1000.0, key="equip_acq")
    with col2:
        spare_parts = lang_or_manual("Spare Parts", "Spare Parts", equip_acq)
    with col3:
        equip_setting = lang_or_manual("Equipment Setting", "Equipment Setting", equip_acq)

    base_equip_costs  = equip_acq + spare_parts + equip_setting
    unsched_pct       = UNSCHED_FACTORS.get((st.session_state.dm_trl, st.session_state.dm_info_avail), 0.0)
    total_equip_costs = base_equip_costs * (1 + unsched_pct)

    st.markdown("##### Total Equipment Assembly")
    col4, col5, col6 = st.columns(3)
    with col4:
        st.text_input("Base Equipment Costs",      value=fmt_curr(base_equip_costs),  disabled=True)
    with col5:
        st.text_input("Unscheduled Equipment (%)", value=f"{unsched_pct * 100:.2f}%", disabled=True)
    with col6:
        st.text_input("Total Equipment Costs",     value=fmt_curr(total_equip_costs), disabled=True)
else:
    # Read from saved scenario to keep calculations flowing
    _d = st.session_state.scenarios.get(scenario_name, {})
    equip_acq         = st.session_state.get("equip_acq", 0.0)
    spare_parts       = _d.get("Spare Parts", 0.0)
    equip_setting     = _d.get("Equipment Setting", 0.0)
    base_equip_costs  = equip_acq + spare_parts + equip_setting
    unsched_pct       = UNSCHED_FACTORS.get((st.session_state.dm_trl, st.session_state.dm_info_avail), 0.0)
    total_equip_costs = base_equip_costs * (1 + unsched_pct)

st.markdown("---")

# ── 7.2 Installation Costs ────────────────────────────────────────────────────
if _section_toggle("capex_installation", "7. Installation Costs"):
    inst_fields = ["Piping", "Civil", "Steel", "Instrumentals", "Electrical", "Insulation"]
    inst_cols   = st.columns(3)
    inst_values = []
    for i, field in enumerate(inst_fields):
        with inst_cols[i % 3]:
            inst_values.append(lang_or_manual(field, field, equip_acq))
    col_p, col_t = st.columns(2)
    with col_p:
        paint = lang_or_manual("Paint", "Paint", equip_acq)
    total_inst_costs = sum(inst_values) + paint
    with col_t:
        st.text_input("Total Installation Costs", value=fmt_curr(total_inst_costs), disabled=True)
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    inst_fields = ["Piping", "Civil", "Steel", "Instrumentals", "Electrical", "Insulation"]
    inst_values = [_d.get(f, st.session_state.get(f.lower(), 0.0)) for f in inst_fields]
    paint = _d.get("Paint", st.session_state.get("paint", 0.0))
    total_inst_costs = sum(inst_values) + paint

st.markdown("---")

# ── Direct Field Costs (computed, always shown when equipment visible)
total_direct_field_costs = total_equip_costs + total_inst_costs
if _gate("capex_equipment") or _gate("capex_installation"):
    st.markdown("**Direct Field Costs**")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.text_input("Total Direct Field Costs", value=fmt_curr(total_direct_field_costs), disabled=True)
    with col2:
        isbl_contrib = st.number_input("ISBL Contribution (%)", min_value=0.0, max_value=100.0, step=1.0, key="isbl_contrib")
    with col3:
        st.text_input("OSBL Contribution (%)", value=f"{100.0 - isbl_contrib:.2f}%", disabled=True)
else:
    isbl_contrib = st.session_state.get("isbl_contrib", 100.0)

st.markdown("---")

# ── 7.3 Indirect Field Costs ──────────────────────────────────────────────────
if _section_toggle("capex_indirect", "8. Indirect Field Costs"):
    col1, col2, col3 = st.columns(3)
    with col1:
        field_office = lang_or_manual("Field Office Staff", "Field Office Staff", equip_acq)
    with col2:
        const_indirects = lang_or_manual("Construction Indirects", "Construction Indirects", equip_acq)
    total_indirect_field_costs = field_office + const_indirects
    with col3:
        st.text_input("Total Indirect Field Costs", value=fmt_curr(total_indirect_field_costs), disabled=True)
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    field_office    = _d.get("Field Office Staff",     st.session_state.get("field_office_staff", 0.0))
    const_indirects = _d.get("Construction Indirects", st.session_state.get("construction_indirects", 0.0))
    total_indirect_field_costs = field_office + const_indirects

st.markdown("---")

# ── 7.4 Non-Field Costs ───────────────────────────────────────────────────────
if _section_toggle("capex_nonfield", "9. Non-Field Costs"):
    nf_fields = ["Freight", "Taxes and Permits", "Engineering and HO", "GA Overheads", "Contract Fee"]
    nf_cols   = st.columns(3)
    nf_values = []
    for i, field in enumerate(nf_fields):
        with nf_cols[i % 3]:
            nf_values.append(lang_or_manual(field, field, equip_acq))
    total_non_field_costs = sum(nf_values)
    with nf_cols[(len(nf_fields)) % 3]:
        st.text_input("Total Non-Field Costs", value=fmt_curr(total_non_field_costs), disabled=True)
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    nf_fields = ["Freight", "Taxes and Permits", "Engineering and HO", "GA Overheads", "Contract Fee"]
    nf_values = [_d.get(f, st.session_state.get(f.lower().replace(" ", "_").replace("&", "and"), 0.0)) for f in nf_fields]
    total_non_field_costs = sum(nf_values)

st.markdown("---")

# ── 7.5 CAPEX Calculations ────────────────────────────────────────────────────
if _section_toggle("capex_calculations", "10. CAPEX Calculations"):
    project_costs_isbl_osbl = total_direct_field_costs + total_indirect_field_costs + total_non_field_costs
    st.metric("Project Costs for ISBL + OSBL", fmt_curr(project_costs_isbl_osbl))
    st.markdown("---")

    col_cx1, col_cx2, col_cx3 = st.columns(3)
    with col_cx1:
        contingency_pct = PROJECT_CONTINGENCY.get((st.session_state.dm_trl, st.session_state.dm_severity), 0.0)
        st.text_input("Project Contingency (auto)",
                      value=f"{contingency_pct * 100:.1f}%  [TRL: {st.session_state.dm_trl} / Severity: {st.session_state.dm_severity}]",
                      disabled=True)
    with col_cx2:
        databank_year_input = st.number_input("Databank Year", min_value=PCI_YEARS[0], max_value=PCI_YEARS[-1], step=1, key="databank_year")
        pci_databank = PLANT_COST_INDEX.get(int(databank_year_input))
        st.text_input("PCI (Databank Year)", value=f"{pci_databank:.2f}" if pci_databank else "—", disabled=True)
    with col_cx3:
        analysis_year_input = st.number_input("Year of Analysis", min_value=1900, max_value=2100, step=1, key="analysis_year")
        pci_analysis = PLANT_COST_INDEX.get(int(analysis_year_input))
        if pci_analysis is None:
            st.warning(f"Index not available. Defaulting to {PCI_YEARS[-1]}.")
            pci_analysis = PLANT_COST_INDEX[PCI_YEARS[-1]]
            effective_analysis_year = PCI_YEARS[-1]
        else:
            effective_analysis_year = int(analysis_year_input)
        st.text_input("PCI (Year of Analysis)", value=f"{pci_analysis:.2f}", disabled=True)

    st.markdown("---")
    col_cx4, col_cx5, col_cx6, col_cx7 = st.columns(4)
    time_update_factor = pci_analysis / pci_databank if pci_databank else 0.0
    with col_cx4:
        st.text_input("Time Update Factor", value=f"{time_update_factor:.4f}", disabled=True)
    with col_cx5:
        st.text_input("Project Location", key="proj_location")
    with col_cx6:
        location_factor = st.number_input("Location Factor", min_value=0.0, step=0.01, key="location_factor")
    with col_cx7:
        project_capex = project_costs_isbl_osbl * (1 + contingency_pct) * time_update_factor * location_factor
        st.metric("Project CAPEX", fmt_curr(project_capex))
else:
    # Read computed values from saved scenario
    _d = st.session_state.scenarios.get(scenario_name, {})
    project_costs_isbl_osbl = total_direct_field_costs + total_indirect_field_costs + total_non_field_costs
    contingency_pct      = PROJECT_CONTINGENCY.get((st.session_state.dm_trl, st.session_state.dm_severity), 0.0)
    databank_year_input  = st.session_state.get("databank_year", 2022)
    pci_databank         = PLANT_COST_INDEX.get(int(databank_year_input))
    analysis_year_input  = st.session_state.get("analysis_year", PCI_YEARS[-1])
    pci_analysis         = PLANT_COST_INDEX.get(int(analysis_year_input)) or PLANT_COST_INDEX[PCI_YEARS[-1]]
    effective_analysis_year = int(analysis_year_input)
    time_update_factor   = pci_analysis / pci_databank if pci_databank else 0.0
    location_factor      = st.session_state.get("location_factor", 0.97)
    project_capex        = _d.get("Project CAPEX", project_costs_isbl_osbl * (1 + contingency_pct) * time_update_factor * location_factor)

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — Additional Information  (gated: any cost source set)
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("additional_info", "11. Additional Information"):
    col_ai1, col_ai2 = st.columns(2)
    with col_ai1:
        working_hours = st.number_input("Working Hours per Year (h/y)", min_value=1.0, max_value=8760.0, step=10.0, key="working_hours")
    with col_ai2:
        scaling_factor = st.number_input("Scaling Factor", min_value=0.001, step=0.01, key="scaling_factor")
else:
    working_hours  = st.session_state.get("working_hours", 8000.0)
    scaling_factor = st.session_state.get("scaling_factor", 0.6)

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 9 — Variable Costs  (gated: basic info complete)
# ══════════════════════════════════════════════════════════════════════════════

def _cost_table_v2(section_key: str, editor_key: str, working_hours: float) -> tuple[pd.DataFrame, float]:
    """
    Variable cost table with full dimensional unit conversion.

    Rate unit sets the dimension (mass/volume/energy/molar/power).
    Price unit must be within the same dimension — user picks freely.
    Cost = annual_qty_in_base × price_per_base_unit
         = rate × qty_factor × time_factor × wh × (price / price_qty_factor)

    Returns (active_rows_df, total_annual_cost_usd).
    """
    existing = st.session_state.scenarios.get(st.session_state.sn_input, {}).get(section_key, [])
    _blank = {"Name": None, "Rate": 0.0, "Rate Unit": "kg/h", "Price": 0.0, "Price Unit": "$/kg"}

    if existing:
        df = pd.DataFrame(existing)
        for col in ["Name","Rate","Rate Unit","Price","Price Unit"]:
            if col not in df.columns:
                df[col] = _blank[col]
        df = df[["Name","Rate","Rate Unit","Price","Price Unit"]]
    else:
        df = pd.DataFrame([_blank])

    # ── Render editor ────────────────────────────────────────────────────────
    # Price Unit column uses PRICE_UNITS_FULL (all options) — per-row
    # filtering is done post-edit with a warning if mismatch detected.
    edited = st.data_editor(
        df,
        key=editor_key,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "Name":       st.column_config.TextColumn("Name", width="medium"),
            "Rate":       st.column_config.NumberColumn(
                               "Rate", min_value=0.0, step=0.000001, format="%.6f", width="small"),
            "Rate Unit":  st.column_config.SelectboxColumn(
                               "Rate Unit", options=RATE_UNITS_FULL, width="small"),
            "Price":      st.column_config.NumberColumn(
                               "Price", min_value=0.0, step=0.000001, format="%.6f", width="small"),
            "Price Unit": st.column_config.SelectboxColumn(
                               "Price Unit", options=PRICE_UNITS_FULL, width="small"),
        },
    )

    edited = edited.copy()

    # ── Auto-correct Price Unit when Rate Unit dimension changes ─────────────
    # If the stored Price Unit is not valid for the current Rate Unit dimension,
    # reset it to the dimension default. If it is valid, leave it alone.
    def _fix_price_unit(row):
        ru = row["Rate Unit"] if pd.notna(row["Rate Unit"]) else "kg/h"
        pu = row["Price Unit"] if pd.notna(row["Price Unit"]) else ""
        valid = valid_price_units_for_rate(ru)
        if pu in valid:
            return pu   # user's choice is dimensionally consistent — keep it
        return price_unit_for_rate(ru)  # reset to dimension default

    edited["Price Unit"] = edited.apply(_fix_price_unit, axis=1)

    active = edited[edited["Name"].notna() & (edited["Name"].str.strip() != "")].copy()

    # ── Compute annual cost per row ──────────────────────────────────────────
    ann_qtys  = []
    ann_costs = []
    dim_warns = []

    for _, row in active.iterrows():
        rate  = float(row["Rate"])  if pd.notna(row["Rate"])  else 0.0
        price = float(row["Price"]) if pd.notna(row["Price"]) else 0.0
        ru    = row["Rate Unit"]  if pd.notna(row["Rate Unit"])  else "kg/h"
        pu    = row["Price Unit"] if pd.notna(row["Price Unit"]) else price_unit_for_rate(ru)

        # Annual quantity in base unit (kg, m3, kWh, mol)
        ann_qty, ann_label = annual_quantity(rate, ru, working_hours)

        # Price converted to $/base_unit
        price_base = price_per_base_unit(price, pu)

        ann_cost = ann_qty * price_base
        ann_qtys.append(f"{ann_qty:,.4f} {ann_label}")
        ann_costs.append(ann_cost)

        # Dimensional consistency check (should never fire after _fix_price_unit)
        valid = valid_price_units_for_rate(ru)
        if pu not in valid:
            dim_warns.append(f"{row['Name']}: {ru} rate with {pu} price — incompatible dimensions.")

    if dim_warns:
        for w in dim_warns:
            st.warning(f"⚠️ Unit mismatch: {w}")

    active["Annual Quantity"] = ann_qtys
    active["_ann_cost"]       = ann_costs

    # ── Display results ──────────────────────────────────────────────────────
    display_df = active[["Name","Rate","Rate Unit","Price","Price Unit","Annual Quantity"]].copy()
    display_df["Cost (USD/year)"] = [f"${c:,.2f}" for c in ann_costs]
    if not display_df.empty:
        st.dataframe(display_df, use_container_width=True, hide_index=True)

    return active, sum(ann_costs)


def _total_row(label, value, key):
    st.markdown("---")
    c1, c2 = st.columns([3, 2])
    with c1:
        st.markdown(f"**{label}**")
    with c2:
        st.text_input(key, value=fmt_curr(value), disabled=True, label_visibility="collapsed")


if _section_toggle("variable_costs", "12. Variable Costs"):
    st.subheader("Raw Materials")
    rm_active, total_raw_material_cost = _cost_table_v2("Raw Materials", f"rm_editor_{st.session_state.table_key}", working_hours)
    _total_row("Total raw materials cost", total_raw_material_cost, "total_rm")
    st.space("medium")

    st.subheader("Chemical Inputs and Utilities")
    cu_active, total_chemical_utilities = _cost_table_v2("Chemical Inputs and Utilities", f"cu_editor_{st.session_state.table_key}", working_hours)
    _total_row("Total chemical inputs and utilities", total_chemical_utilities, "total_cu")
    st.space("medium")

    st.subheader("Credits and Byproducts")
    cb_active, total_revenue = _cost_table_v2("Credits and Byproducts", f"cb_editor_{st.session_state.table_key}", working_hours)
    _total_row("Total Revenue (Credits)", total_revenue, "total_cb")
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    total_raw_material_cost = _d.get("Total Raw Material Cost", 0.0)
    total_chemical_utilities = _d.get("Total Chemical Inputs Utilities", 0.0)
    total_revenue = _d.get("Total Revenue", 0.0)
    rm_active = pd.DataFrame(_d.get("Raw Materials", []))
    cu_active = pd.DataFrame(_d.get("Chemical Inputs and Utilities", []))
    cb_active = pd.DataFrame(_d.get("Credits and Byproducts", []))

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 10 — Fixed Costs  (gated: basic info + CAPEX)
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("fixed_costs", "13. Fixed Costs"):
    st.subheader("Labor Costs")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        n_operators = st.number_input("Number of operators per shift", min_value=1, step=1, key="n_operators")
    with col2:
        operator_salary = st.number_input("Operator salary (USD/month)", min_value=MIN_SALARY, step=10.0, key="operator_salary")
        if operator_salary < MIN_SALARY:
            st.error(f"Cannot be lower than {fmt_curr(MIN_SALARY)}.")
    with col3:
        n_supervisors = st.number_input("Number of supervisors per shift", min_value=1, step=1, key="n_supervisors")
        if n_supervisors >= n_operators:
            st.error("Must be less than number of operators.")
    with col4:
        supervisor_salary = st.number_input("Supervisor salary (USD/month)", min_value=MIN_SALARY, step=10.0, key="supervisor_salary")
        if supervisor_salary <= operator_salary:
            st.error("Must exceed operator salary.")

    col1, col2 = st.columns([1, 3])
    with col1:
        salary_charges = st.number_input("Salary charges (multiplier)", min_value=1.0, step=0.05, key="salary_charges")

    st.markdown("---")
    st.markdown("##### Schedule")
    labor_wh_ref = st.session_state.get("working_hours", 8000.0)
    col1, col2, col3 = st.columns(3)
    with col1:
        labor_wh = _overridable_number("Plant working hours per year", labor_wh_ref, "labor_working_hrs_override", step=10.0)
    with col2:
        plant_daily_hours = st.number_input("Plant daily operation hours (h/day)", min_value=0.1, max_value=24.0, step=0.5, key="plant_daily_hours")
    with col3:
        weekly_op_days = st.number_input("Weekly operation days (days/week)", min_value=1.0, max_value=7.0, step=1.0, key="weekly_op_days")

    operating_weeks = labor_wh / plant_daily_hours / weekly_op_days

    col4, col5, col6 = st.columns(3)
    with col4:
        st.text_input("Operating weeks per year", value=f"{operating_weeks:.2f} weeks", disabled=True)
    with col5:
        worker_hours_shift = st.number_input("Worker hours per shift (h/shift)", min_value=0.1, step=0.5, key="worker_hours_shift")
    with col6:
        worker_shifts_week = st.number_input("Worker shifts per week (shifts/week)", min_value=0.1, step=0.5, key="worker_shifts_week")

    col7, col8, col9 = st.columns(3)
    with col7:
        worker_vacation_weeks = st.number_input("Worker vacation weeks per year", min_value=0.0, step=1.0, key="worker_vacation_weeks")

    worker_weeks_per_year = math.floor(operating_weeks - worker_vacation_weeks)
    with col8:
        st.text_input("Worker weeks per year", value=f"{worker_weeks_per_year} weeks", disabled=True)

    denom = worker_shifts_week * worker_weeks_per_year
    op_team_factor = math.ceil((labor_wh / worker_hours_shift) / denom) if denom > 0 else 0
    with col9:
        st.text_input("Operating team factor", value=f"{op_team_factor}", disabled=True)

    st.markdown("---")
    olc = (n_operators * operator_salary + n_supervisors * supervisor_salary) * salary_charges * op_team_factor * 12.0

    st.markdown("##### Operating Labor Costs")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.text_input("OLC – Operating Labor Costs (USD/year)", value=fmt_curr(olc), disabled=True)

    lab_ref_pct    = LAB_CHARGES.get(st.session_state.dm_prod_type, 0.10) * 100.0
    office_ref_pct = OFFICE_LABOR.get(st.session_state.dm_prod_type, 0.10) * 100.0
    with col2:
        lab_pct_input = _overridable_number(f"Laboratory charges (% of OLC)  [ref: {lab_ref_pct:.2f}%]", lab_ref_pct, "lab_charges_override", step=0.1)
        lab_pct = max(0.0, lab_pct_input) / 100.0
    with col3:
        office_pct_input = _overridable_number(f"Office labor (% of OLC)  [ref: {office_ref_pct:.2f}%]", office_ref_pct, "office_labor_override", step=0.1)
        office_pct = max(0.0, office_pct_input) / 100.0

    total_labor_costs = olc * (1.0 + lab_pct + office_pct)
    st.markdown("---")
    col_lc1, col_lc2 = st.columns([3, 2])
    with col_lc1:
        st.markdown("**Total labor costs (USD/year)**")
    with col_lc2:
        st.text_input("total_labor", value=fmt_curr(total_labor_costs), disabled=True, label_visibility="collapsed")

    st.space("medium")
    st.subheader("Supply and Maintenance Costs")

    maint_ref_pct  = MAINTENANCE_REPAIRS.get((st.session_state.dm_mat_type, st.session_state.dm_prod_type), 0.01) * 100.0
    op_sup_ref_pct = OPERATING_SUPPLIES.get(st.session_state.dm_severity, 0.15) * 100.0

    col1, col2, col3 = st.columns(3)
    with col1:
        maint_pct_input = _overridable_number(f"Maintenance and repairs (% of CAPEX)  [ref: {maint_ref_pct:.2f}%]", maint_ref_pct, "maint_repair_override", step=0.1)
    maint_pct = maint_pct_input / 100.0
    with col2:
        op_sup_pct_input = _overridable_number(f"Operating supplies (% of Maintenance)  [ref: {op_sup_ref_pct:.2f}%]", op_sup_ref_pct, "op_supplies_override", step=0.1)
    op_sup_pct = op_sup_pct_input / 100.0
    supply_maint_costs = (maint_pct + op_sup_pct) * project_capex
    with col3:
        st.text_input("Supply and maintenance costs (USD/year)", value=fmt_curr(supply_maint_costs), disabled=True)

    st.space("medium")
    st.subheader("Additional Fixed Costs")

    admin_ov_ref_pct  = ADMIN_OVERHEAD.get(st.session_state.dm_prod_type, 0.50) * 100.0 * (1 + office_pct)
    mfg_ov_ref_pct    = MFG_OVERHEAD.get(st.session_state.dm_severity, 0.006) * 100.0
    taxes_ins_ref_pct = TAXES_INSURANCE.get(st.session_state.dm_severity, 0.032) * 100.0
    patents_ref       = PATENTS_ROYALTIES.get((st.session_state.dm_trl, st.session_state.dm_prod_type))
    patents_ref_pct   = (patents_ref * 100.0) if patents_ref is not None else 0.0
    patents_na        = patents_ref is None

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        admin_ov_pct_input = _overridable_number(f"Administrative overhead (% of OLC)  [ref: {admin_ov_ref_pct:.2f}%]", admin_ov_ref_pct, "admin_overhead_override", step=0.1)
    admin_ov_pct = admin_ov_pct_input / 100.0
    with col2:
        mfg_ov_pct_input = _overridable_number(f"Manufacturing overhead (% of CAPEX)  [ref: {mfg_ov_ref_pct:.2f}%]", mfg_ov_ref_pct, "mfg_overhead_override", step=0.01)
    mfg_ov_pct = mfg_ov_pct_input / 100.0
    with col3:
        taxes_ins_pct_input = _overridable_number(f"Taxes and insurance (% of CAPEX)  [ref: {taxes_ins_ref_pct:.2f}%]", taxes_ins_ref_pct, "taxes_ins_override", step=0.01)
    taxes_ins_pct = taxes_ins_pct_input / 100.0
    with col4:
        if patents_na:
            st.info(f"Patents N/A for {st.session_state.dm_trl} / {st.session_state.dm_prod_type}")
            patents_pct = 0.0
        else:
            patents_pct_input = _overridable_number(f"Patents and royalties (% of OPEX)  [ref: {patents_ref_pct:.2f}%]", patents_ref_pct, "patents_roy_override", step=0.1)
            patents_pct = patents_pct_input / 100.0

    st.space("medium")
    st.subheader("Indirect Fixed Costs")

    admin_costs_ref_pct = (1.0 + office_pct) * 0.15 * 100.0
    mfg_costs_ref_pct   = maint_pct * 0.15 * 100.0
    dist_sell_ref_pct   = DIST_SELLING.get(st.session_state.dm_prod_type, 0.08) * 100.0
    r_d_ref_pct         = R_AND_D.get((st.session_state.dm_trl, st.session_state.dm_prod_type), 0.02) * 100.0

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        admin_costs_pct_input = _overridable_number(f"Administrative costs (% of OLC)  [ref: {admin_costs_ref_pct:.2f}%]", admin_costs_ref_pct, "admin_costs_override", step=0.1)
    admin_costs_pct = admin_costs_pct_input / 100.0
    with col2:
        mfg_costs_pct_input = _overridable_number(f"Manufacturing costs (% of CAPEX)  [ref: {mfg_costs_ref_pct:.2f}%]", mfg_costs_ref_pct, "mfg_costs_override", step=0.01)
    mfg_costs_pct = mfg_costs_pct_input / 100.0
    with col3:
        dist_sell_pct_input = _overridable_number(f"Distribution and selling (% of OPEX)  [ref: {dist_sell_ref_pct:.2f}%]", dist_sell_ref_pct, "dist_selling_override", step=0.1)
    dist_sell_pct = dist_sell_pct_input / 100.0
    with col4:
        r_d_pct_input = _overridable_number(f"R&D (% of OPEX)  [ref: {r_d_ref_pct:.2f}%]", r_d_ref_pct, "r_and_d_override", step=0.1)
    r_d_pct = r_d_pct_input / 100.0

    # Analytical OPEX solve
    # OPEX solve uses gross variable costs (RM + CU only).
    # Credits/byproducts appear on the revenue side of the cash flow, not as cost deductions.
    direct_var_costs = total_raw_material_cost + total_chemical_utilities
    _olc_coeff   = admin_ov_pct + admin_costs_pct
    _capex_coeff = mfg_ov_pct + taxes_ins_pct + mfg_costs_pct
    _numerator   = (direct_var_costs + total_labor_costs + supply_maint_costs
                    + _olc_coeff * olc + _capex_coeff * project_capex)
    _denominator = 1.0 - patents_pct - dist_sell_pct - r_d_pct
    opex = _numerator / _denominator if _denominator > 0 else 0.0

    afc = (admin_ov_pct * olc + (mfg_ov_pct + taxes_ins_pct) * project_capex + patents_pct * opex)
    indirect_fixed_costs = (admin_costs_pct * olc + mfg_costs_pct * project_capex + (dist_sell_pct + r_d_pct) * opex)
    direct_fixed_costs   = total_labor_costs + supply_maint_costs + afc
    total_fixed_costs    = direct_fixed_costs + indirect_fixed_costs
    total_opex           = opex

    st.markdown("---")
    _result_row("Additional fixed costs (USD/year)", afc, "afc_display",
        "AFC = admin_overhead×OLC + (mfg_overhead + taxes_insurance)×CAPEX + patents×OPEX")
    st.markdown("---")
    _result_row("Direct fixed costs (USD/year)", direct_fixed_costs, "dfc_display",
        "Direct fixed costs = Labor + Supply & maintenance + AFC")
    st.markdown("---")
    _result_row("Indirect fixed costs (USD/year)", indirect_fixed_costs, "ifc_display",
        "Indirect fixed costs = adm_costs×OLC + mfg_costs×CAPEX + (dist_sell + R&D)×OPEX")
    st.markdown("---")
    _result_row("Total fixed costs (USD/year)", total_fixed_costs, "tfc_display", "")
    st.space("medium")
    _result_row("Total OPEX (USD/year)", total_opex, "total_opex_display",
        "OPEX = (TVC + Labor + Supply + coeffs×OLC + coeffs×CAPEX) / (1 − patents − dist_sell − R&D)")
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    olc                  = _d.get("OLC", 0.0)
    office_pct           = _d.get("Office Labor Pct", 0.0)
    lab_pct              = _d.get("Lab Charges Pct", 0.0)
    maint_pct            = _d.get("Maint Pct", 0.0)
    op_sup_pct           = _d.get("Op Sup Pct", 0.0)
    admin_ov_pct         = _d.get("Admin Ov Pct", 0.0)
    mfg_ov_pct           = _d.get("Mfg Ov Pct", 0.0)
    taxes_ins_pct        = _d.get("Taxes Ins Pct", 0.0)
    patents_pct          = _d.get("Patents Pct", 0.0)
    admin_costs_pct      = _d.get("Admin Costs Pct", 0.0)
    mfg_costs_pct        = _d.get("Mfg Costs Pct", 0.0)
    dist_sell_pct        = _d.get("Dist Sell Pct", 0.0)
    r_d_pct              = _d.get("R D Pct", 0.0)
    total_labor_costs    = _d.get("Total Labor Costs", 0.0)
    supply_maint_costs   = _d.get("Supply Maint Costs", 0.0)
    afc                  = _d.get("AFC Pre Patents", 0.0)
    direct_fixed_costs   = _d.get("Direct Fixed Costs", 0.0)
    indirect_fixed_costs = _d.get("Indirect Fixed Costs", 0.0)
    total_fixed_costs    = _d.get("Total Fixed Costs", 0.0)
    total_opex           = _d.get("Total OPEX", 0.0)
    opex                 = total_opex
    n_operators          = st.session_state.get("n_operators", 2)
    n_supervisors        = st.session_state.get("n_supervisors", 1)
    salary_charges       = st.session_state.get("salary_charges", 2.2)
    op_team_factor       = _d.get("Operating Team Factor", 1)
    worker_weeks_per_year = _d.get("Worker Weeks per Year", 48)
    operating_weeks      = _d.get("Operating Weeks", 48.0)

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 11 — Working Capital
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("working_capital", "14. Working Capital"):
    wc_method = st.radio("Working capital method", options=["Percentage", "Operating Cycle"],
                         index=0 if st.session_state.wc_method == "Percentage" else 1,
                         horizontal=True, key="wc_method")
    st.markdown("---")

    if wc_method == "Percentage":
        col1, col2 = st.columns(2)
        with col1:
            wc_pct_input = st.number_input("Working capital percentage (% of CAPEX)", min_value=0.0, step=0.5, value=st.session_state.wc_pct, key="wc_pct")
        working_capital = project_capex * (wc_pct_input / 100.0)
        with col2:
            _result_row("Working capital (USD)", working_capital, "wc_display", "WC = CAPEX × %")
    else:
        st.markdown("##### Current Assets")
        col1, col2, col3 = st.columns(3)
        with col1:
            equiv_cash_days = st.number_input("Equivalent cash (days of OPEX)", min_value=0.0, step=1.0, value=st.session_state.wc_equiv_cash_days, key="wc_equiv_cash_days")
        with col2:
            raw_mat_days = st.number_input("Raw material inventory (days of variable costs)", min_value=0.0, step=1.0, value=st.session_state.wc_raw_mat_days, key="wc_raw_mat_days")
        with col3:
            acc_rec_days = st.number_input("Accounts receivable (days of OPEX + ROI)", min_value=0.0, step=1.0, value=st.session_state.wc_accounts_rec_days, key="wc_accounts_rec_days")

        _r, _n = 0.15, 10
        simple_roi = (_r * project_capex) / (1.0 - (1.0 + _r) ** (-_n)) if project_capex > 0 else 0.0
        # Raw material inventory uses gross variable costs (RM + CU).
        # Credits are revenue, not a cost reduction for inventory purposes.
        direct_var_costs_wc = total_raw_material_cost + total_chemical_utilities
        current_assets = ((equiv_cash_days * opex + direct_var_costs_wc * raw_mat_days
                           + acc_rec_days * (opex + simple_roi)) * (24.0 / working_hours))
        st.markdown("---")
        _result_row("Current assets (USD)", current_assets, "current_assets_display", "")

        st.markdown("---")
        st.markdown("##### Current Liabilities")
        col1, col2 = st.columns(2)
        with col1:
            acc_payroll_days = st.number_input("Accrued payroll (days of OLC)", min_value=0.0, step=1.0, value=st.session_state.wc_accrued_payroll_days, key="wc_accrued_payroll_days")
        with col2:
            acc_pay_days = st.number_input("Accounts payable (days of OPEX)", min_value=0.0, step=1.0, value=st.session_state.wc_accounts_pay_days, key="wc_accounts_pay_days")

        current_liabilities = ((acc_payroll_days * olc + acc_pay_days * opex) * (24.0 / working_hours))
        st.markdown("---")
        _result_row("Current liabilities (USD)", current_liabilities, "current_liab_display", "")
        working_capital = current_assets - current_liabilities
        st.markdown("---")
        _result_row("Working capital (USD)", working_capital, "wc_display", "WC = Current assets − Current liabilities")
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    working_capital = _d.get("Working Capital", 0.0)

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 12 — Startup Costs
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("startup_costs", "15. Startup Costs"):
    startup_method = st.radio("Startup costs method", options=["Single Factor", "Multiple Factors"],
                              index=0 if st.session_state.startup_method == "Single Factor" else 1,
                              horizontal=True, key="startup_method")
    st.markdown("---")

    if startup_method == "Single Factor":
        col1, col2 = st.columns(2)
        with col1:
            startup_single_pct = st.number_input("Startup cost single factor (% of CAPEX)", min_value=0.0, step=0.5,
                                                  value=st.session_state.startup_single_pct, key="startup_single_pct")
        startup_costs = project_capex * (startup_single_pct / 100.0)
        with col2:
            _result_row("Startup costs (USD)", startup_costs, "startup_display", "")
    else:
        col1, col2, col3 = st.columns(3)
        with col1:
            startup_op_training_days = st.number_input("Operational training (days of OLC)", min_value=0.0, step=1.0,
                                                       value=st.session_state.startup_op_training_days, key="startup_op_training_days")
        with col2:
            startup_commerc_pct = st.number_input("Commercialization costs (% of direct field costs)", min_value=0.0, step=0.5,
                                                   value=st.session_state.startup_commerc_pct, key="startup_commerc_pct")
        with col3:
            startup_inefficiency_pct = st.number_input("Startup inefficiency (% of OPEX)", min_value=0.0, step=0.5,
                                                       value=st.session_state.startup_inefficiency_pct, key="startup_inefficiency_pct")

        startup_costs = (startup_op_training_days * olc * (24.0 / working_hours)
                         + total_direct_field_costs * (startup_commerc_pct / 100.0)
                         + (startup_inefficiency_pct / 100.0) * total_opex)
        st.markdown("---")
        _result_row("Startup costs (USD)", startup_costs, "startup_costs_display", "")
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    startup_costs = _d.get("Startup Costs", 0.0)

st.space("medium")

# Total Investment Summary (always shown if capex is computed)
if _gate("capex_calculations") or _gate("startup_costs") or _gate("working_capital"):
    st.subheader("Total Investment Summary")
    col1, col2 = st.columns([3, 2])
    with col1:
        additional_costs = st.number_input("Additional costs (USD)", min_value=0.0, step=1000.0, key="additional_costs")

    total_investment = project_capex + working_capital + startup_costs + additional_costs

    st.markdown("---")
    _result_row("Project CAPEX (USD)", project_capex, "ti_capex", "")
    st.markdown("---")
    _result_row("Working capital (USD)", working_capital, "ti_wc", "")
    st.markdown("---")
    _result_row("Startup costs (USD)", startup_costs, "ti_startup", "")
    st.markdown("---")
    _result_row("Additional costs (USD)", additional_costs, "ti_additional", "")
    st.markdown("---")
    _result_row("Total investment costs (USD)", total_investment, "ti_total",
        "TIC = Project CAPEX + Working capital + Startup costs + Additional costs")

    st.space("medium")

    # TIC Accuracy
    st.subheader("Total Investment Cost (TIC) Accuracy")
    st.caption(f"TRL = {st.session_state.dm_trl} | Info availability = {st.session_state.dm_info_avail}")
    _tic_key = (st.session_state.dm_trl, st.session_state.dm_info_avail)
    tic_lower_ref = TIC_LOWER.get(_tic_key)
    tic_upper_ref = TIC_UPPER.get(_tic_key)
    col1, col2 = st.columns(2)
    with col1:
        if tic_lower_ref is None:
            st.info("TIC lower bound: N/A")
            tic_lower_input = None
        else:
            tic_lower_input = _overridable_number(f"Lower bound (%)  [ref: {tic_lower_ref:.1f}%]", tic_lower_ref, "tic_lower_override", step=0.5, min_value=None)
    with col2:
        if tic_upper_ref is None:
            st.info("TIC upper bound: N/A")
            tic_upper_input = None
        else:
            tic_upper_input = _overridable_number(f"Upper bound (%)  [ref: {tic_upper_ref:.1f}%]", tic_upper_ref, "tic_upper_override", step=0.5, min_value=None)

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if tic_lower_input is not None:
            _result_row("TIC lower bound (USD)", total_investment * (1.0 + tic_lower_input / 100.0), "tic_lower_usd", "")
        else:
            st.info("TIC lower bound: N/A")
    with col2:
        if tic_upper_input is not None:
            _result_row("TIC upper bound (USD)", total_investment * (1.0 + tic_upper_input / 100.0), "tic_upper_usd", "")
        else:
            st.info("TIC upper bound: N/A")
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    additional_costs  = st.session_state.get("additional_costs", 0.0)
    total_investment  = _d.get("Total Investment", 0.0)
    tic_lower_input   = _d.get("TIC Lower Pct")
    tic_upper_input   = _d.get("TIC Upper Pct")

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 13 — Financial Assumptions
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("financial_assumptions", "16. Financial Assumptions"):
    # Land
    st.subheader("Land")
    land_option = st.radio("Land option", options=["Buy", "Rent"],
                           index=0 if st.session_state.land_option == "Buy" else 1,
                           horizontal=True, key="land_option")
    st.markdown("---")
    col1, col2 = st.columns(2)
    if land_option == "Buy":
        with col1:
            land_buy_pct = _overridable_number("Land purchase factor (% of ISBL+OSBL)  [ref: 2.00%]", 2.0, "land_buy_pct_override", step=0.1)
        land_rent_pct = 0.0
        with col2:
            _result_row("Land cost (USD)", project_costs_isbl_osbl * (land_buy_pct / 100.0), "land_cost_display", "")
    else:
        with col1:
            land_rent_pct = _overridable_number("Land rent factor (% of ISBL+OSBL per year)  [ref: 0.20%]", 0.2, "land_rent_pct_override", step=0.01)
        land_buy_pct = 0.0
        with col2:
            _result_row("Annual land rent (USD/year)", project_costs_isbl_osbl * (land_rent_pct / 100.0), "land_rent_display", "")

    st.space("medium")

    # Depreciation
    st.subheader("Depreciation")
    depreciation_method = st.radio("Depreciation method", options=["Straight Line", "MACRS"],
                                   index=0 if st.session_state.depreciation_method == "Straight Line" else 1,
                                   horizontal=True, key="depreciation_method")
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        depreciation_years = st.number_input("Depreciation period (years)", min_value=1, step=1, key="depreciation_years")
    with col2:
        residual_value_pct = st.number_input("Residual value (% of CAPEX)", min_value=0.0, max_value=99.9, step=1.0, key="residual_value_pct")

    residual_value_usd = project_capex * (residual_value_pct / 100.0)
    depreciable_base   = project_capex - residual_value_usd
    col3, col4 = st.columns(2)
    with col3:
        _result_row("Residual value (USD)", residual_value_usd, "residual_value_display", "")
    with col4:
        _result_row("Depreciable base (USD)", depreciable_base, "depreciable_base_display", "")

    st.space("medium")

    # Taxes
    st.subheader("Taxes")
    def _on_country_change():
        st.session_state.tax_rate_override = None
        st.session_state.pop("w_tax_rate_override", None)

    col1, col2 = st.columns(2)
    with col1:
        tax_country = st.selectbox("Country", options=COUNTRY_LIST,
                                   index=COUNTRY_LIST.index(st.session_state.get("tax_country", "Brazil")),
                                   key="tax_country", on_change=_on_country_change)
    tax_ref_pct = TAXES_BY_COUNTRY.get(tax_country, 0.34) * 100.0
    with col2:
        tax_rate_input = _overridable_number(f"Corporate tax rate (%)  [ref: {tax_ref_pct:.1f}% — {tax_country}]", tax_ref_pct, "tax_rate_override", step=0.1)
    tax_rate = tax_rate_input / 100.0

    st.space("medium")

    # Financial Leverage
    st.subheader("Financial Leverage")
    financing_type = st.radio("Financing type", options=["None", "Straight Line"],
                              index=0 if st.session_state.financing_type == "None" else 1,
                              horizontal=True, key="financing_type")
    st.markdown("---")
    is_leveraged = financing_type == "Straight Line"
    col1, col2, col3 = st.columns(3)
    with col1:
        debt_ratio_pct = st.number_input("Debt ratio (% of CAPEX)", min_value=0.0, max_value=99.9, step=1.0, key="debt_ratio_pct", disabled=not is_leveraged)
    with col2:
        amortization_years = st.number_input("Amortization period (years)", min_value=1, step=1, key="amortization_years", disabled=not is_leveraged)
    with col3:
        grace_period_years = st.number_input("Grace period (years)", min_value=0, step=1, key="grace_period_years", disabled=not is_leveraged)
    debt_ratio = (debt_ratio_pct / 100.0) if is_leveraged else 0.0

    st.space("medium")

    # MARR
    st.subheader("MARR")
    col1, col2 = st.columns(2)
    with col1:
        central_bank_rate = st.number_input("Central bank interest rate (%)", min_value=0.0, step=0.01, key="central_bank_rate")
    with col2:
        credit_spread = st.number_input("Credit spread (%)", min_value=0.0, step=0.01, key="credit_spread")

    cod = (central_bank_rate + credit_spread) / 100.0 * (1.0 - tax_rate)
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        _result_row("Cost of Debt — COD (%)", cod * 100.0, "cod_display", "COD = (CBR + CS) × (1 − tax)")

    col1, col2 = st.columns(2)
    with col1:
        unlevered_beta = st.number_input("Unlevered / Asset beta", min_value=0.0, step=0.01, key="unlevered_beta")
    _leverage_factor = (1.0 - tax_rate) * debt_ratio / (1.0 - debt_ratio) if debt_ratio < 1.0 else 0.0
    levered_beta = unlevered_beta * (1.0 + _leverage_factor)
    with col2:
        _result_row("Levered / Equity beta", levered_beta, "levered_beta_display", "")

    st.markdown("---")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        market_return = st.number_input("Market return — rM (%)", min_value=0.0, step=0.01, key="market_return")
    with col2:
        risk_free_rate = st.number_input("Risk-free rate (%)", min_value=0.0, step=0.01, key="risk_free_rate")
    market_risk_premium = levered_beta * (market_return - risk_free_rate) / 100.0
    with col3:
        _result_row("Market risk premium — PRM (%)", market_risk_premium * 100.0, "mrp_display", "PRM = beta × (rM − rf)")
    with col4:
        country_risk_premium = st.number_input("Country risk premium — PRP (%)", min_value=0.0, step=0.01, key="country_risk_premium")

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        us_cpi = st.number_input("U.S. CPI (%)", min_value=0.0, step=0.01, key="us_cpi")
    with col2:
        country_cpi = st.number_input("Country CPI (%)", min_value=0.0, step=0.01, key="country_cpi")

    _rfr   = risk_free_rate  / 100.0
    _ccpi  = country_cpi     / 100.0
    _uscpi = us_cpi          / 100.0
    coe = ((1.0 + _rfr) * (1.0 + _ccpi) / (1.0 + _uscpi) - 1.0) + market_risk_premium + country_risk_premium / 100.0

    st.markdown("---")
    _result_row("Cost of Equity — COE (%)", coe * 100.0, "coe_display", "COE = CAPM + PRP adjusted for inflation differential")

    _marr_formula = (debt_ratio * cod + (1.0 - debt_ratio) * coe) if is_leveraged else coe
    _marr_ref_pct = _marr_formula * 100.0
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        marr_input = _overridable_number("MARR — Minimum Acceptable Rate of Return (%)", _marr_ref_pct, "marr_override", step=0.01, min_value=None)
    marr = marr_input / 100.0
    with col2:
        st.markdown(
            '<p style="font-size:0.82rem;color:#555;margin-top:28px;">'
            + ("COE  (no leverage)" if not is_leveraged
               else f"debt×COD + equity×COE | {debt_ratio_pct:.1f}%×{cod*100:.2f}% + {100-debt_ratio_pct:.1f}%×{coe*100:.2f}%")
            + "</p>", unsafe_allow_html=True)
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    land_option         = _d.get("Land Option", "Buy")
    land_buy_pct        = _d.get("Land Buy Pct", 2.0)
    land_rent_pct       = _d.get("Land Rent Pct", 0.0)
    depreciation_method = _d.get("Depreciation Method", "Straight Line")
    depreciation_years  = _d.get("Depreciation Years", 10)
    residual_value_pct  = _d.get("Residual Value Pct", 20.0)
    residual_value_usd  = _d.get("Residual Value USD", 0.0)
    depreciable_base    = _d.get("Depreciable Base", 0.0)
    tax_country         = _d.get("Tax Country", "Brazil")
    tax_rate            = _d.get("Tax Rate", 0.34)
    financing_type      = _d.get("Financing Type", "None")
    is_leveraged        = financing_type == "Straight Line"
    debt_ratio_pct      = _d.get("Debt Ratio Pct", 50.0)
    debt_ratio          = _d.get("Debt Ratio", 0.0)
    amortization_years  = _d.get("Amortization Years", 13)
    grace_period_years  = _d.get("Grace Period Years", 5)
    central_bank_rate   = _d.get("Central Bank Rate", 5.45)
    credit_spread       = _d.get("Credit Spread", 2.94)
    cod                 = _d.get("COD", 0.0)
    unlevered_beta      = _d.get("Unlevered Beta", 1.0)
    levered_beta        = _d.get("Levered Beta", 1.0)
    market_return       = _d.get("Market Return", 8.63)
    risk_free_rate      = _d.get("Risk Free Rate", 1.94)
    market_risk_premium = _d.get("Market Risk Premium", 0.0)
    country_risk_premium = _d.get("Country Risk Premium", 3.63)
    us_cpi              = _d.get("US CPI", 2.46)
    country_cpi         = _d.get("Country CPI", 4.65)
    coe                 = _d.get("COE", 0.0)
    marr                = _d.get("MARR", 0.1)
    tax_rate_input      = tax_rate * 100.0

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 14 — Project Lifetime & CAPEX Distribution
# ══════════════════════════════════════════════════════════════════════════════

if _section_toggle("project_lifetime", "17. Project Lifetime & CAPEX Distribution"):
    col1, col2, col3 = st.columns(3)
    with col1:
        epc_years = st.number_input("EPC time (years)", min_value=1, max_value=10, step=1, key="epc_years")
    with col2:
        project_lifetime = st.number_input("Project lifetime (years)", min_value=1, max_value=40, step=1, key="project_lifetime")
    with col3:
        _result_row("Total project lifetime (years)", float(epc_years + project_lifetime), "total_lifetime_display", "")

    st.markdown("---")
    _capex_col  = str(epc_years)
    _capex_dist = CAPEX_DISTRIBUTION[_capex_col]
    _year_labels = ["1st year (%)","2nd year (%)","3rd year (%)","4th year (%)","5th year (%)",
                    "6th year (%)","7th year (%)","8th year (%)","9th year (%)","10th year (%)"]
    _capex_df = pd.DataFrame({"Year": _year_labels, "CAPEX Distribution": [f"{v*100:.4f}%" for v in _capex_dist.values]})
    st.dataframe(_capex_df, use_container_width=False, hide_index=True)

    st.space("medium")
    st.subheader("Percentage of Total Capacity per Year")
    col1, col2, col3 = st.columns(3)
    with col1:
        capacity_first_year    = st.number_input("First year (%)", min_value=0.0, max_value=100.0, step=1.0, key="capacity_first_year")
    with col2:
        capacity_intermediate  = st.number_input("Intermediate years (%)", min_value=0.0, max_value=100.0, step=1.0, key="capacity_intermediate")
    with col3:
        capacity_last_year     = st.number_input("Last year (%)", min_value=0.0, max_value=100.0, step=1.0, key="capacity_last_year")

    st.space("medium")
    st.subheader("Percentage of Fixed Costs per Year")
    col1, col2, col3 = st.columns(3)
    with col1:
        fixed_costs_first_year   = st.number_input("First year (%)", min_value=0.0, max_value=100.0, step=1.0, key="fixed_costs_first_year")
    with col2:
        fixed_costs_intermediate = st.number_input("Intermediate years (%)", min_value=0.0, max_value=100.0, step=1.0, key="fixed_costs_intermediate")
    with col3:
        fixed_costs_last_year    = st.number_input("Last year (%)", min_value=0.0, max_value=100.0, step=1.0, key="fixed_costs_last_year")

    st.space("medium")
    st.subheader("Market Assumptions — Annual Growth Rates")
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        growth_main_price      = st.number_input("Main product price (% / year)", step=0.1, key="growth_main_price")
    with col2:
        growth_byproduct_price = st.number_input("Byproduct prices (% / year)", step=0.1, key="growth_byproduct_price")
    with col3:
        growth_raw_materials   = st.number_input("Raw materials (% / year)", step=0.1, key="growth_raw_materials")
    with col4:
        growth_chem_utilities  = st.number_input("Chem inputs & utilities (% / year)", step=0.1, key="growth_chem_utilities")
    with col5:
        growth_fixed_costs     = st.number_input("Fixed costs (% / year)", step=0.1, key="growth_fixed_costs")

    st.space("medium")
    st.subheader("Other Premises")
    col1, col2 = st.columns(2)
    with col1:
        main_product_price = st.number_input(
            f"Main product selling price (USD / base unit)",
            min_value=0.0, step=0.01,
            value=st.session_state.main_product_price if st.session_state.main_product_price is not None else 0.0,
            key="main_product_price",
        )
else:
    _d = st.session_state.scenarios.get(scenario_name, {})
    epc_years              = _d.get("EPC Years", st.session_state.get("epc_years", 3))
    project_lifetime       = _d.get("Project Lifetime", st.session_state.get("project_lifetime", 20))
    capacity_first_year    = _d.get("Capacity First Year", 100.0)
    capacity_intermediate  = _d.get("Capacity Intermediate", 100.0)
    capacity_last_year     = _d.get("Capacity Last Year", 100.0)
    fixed_costs_first_year   = _d.get("Fixed Costs First Year", 100.0)
    fixed_costs_intermediate = _d.get("Fixed Costs Intermediate", 100.0)
    fixed_costs_last_year    = _d.get("Fixed Costs Last Year", 100.0)
    growth_main_price      = _d.get("Growth Main Price", 0.0)
    growth_byproduct_price = _d.get("Growth Byproduct Price", 0.0)
    growth_raw_materials   = _d.get("Growth Raw Materials", 0.0)
    growth_chem_utilities  = _d.get("Growth Chem Utilities", 0.0)
    growth_fixed_costs     = _d.get("Growth Fixed Costs", 0.0)
    main_product_price     = st.session_state.get("main_product_price", 0.0) or 0.0

st.space("medium")

# ══════════════════════════════════════════════════════════════════════════════
# SAVE — VALIDATION & WARNINGS
# ══════════════════════════════════════════════════════════════════════════════

def _run_validations() -> tuple[list[str], list[str]]:
    """
    Returns (critical_errors, soft_warnings).
    critical_errors  → block save entirely, must be fixed first.
    soft_warnings    → shown with proceed checkbox, save allowed after acknowledgement.
    """
    crits = []
    warns = []

    # ── Critical checks ──────────────────────────────────────────────────────
    if not st.session_state.get("sn_input", "").strip():
        crits.append("Scenario name is empty.")
    if not st.session_state.get("mp_input", "").strip():
        crits.append("Main product name is not defined.")
    if not st.session_state.get("pc_input"):
        crits.append("Main product capacity is zero or not set.")
    if project_capex <= 0:
        crits.append(f"Project CAPEX is {fmt_curr(project_capex)} — must be greater than zero for cash flow to be valid.")
    if total_opex <= 0:
        crits.append(f"Total OPEX is {fmt_curr(total_opex)} — must be greater than zero. Check that fixed and variable costs are entered.")

    # ── Soft warnings ─────────────────────────────────────────────────────────
    if equip_acq <= 0:
        warns.append("Equipment Acquisition is zero — is this intentional?")

    # Installation costs — check each individually
    _inst_names = ["Piping", "Civil", "Steel", "Instrumentals", "Electrical", "Insulation", "Paint"]
    _inst_vals  = list(inst_values) + [paint]
    for _n, _v in zip(_inst_names, _inst_vals):
        if _v <= 0:
            warns.append(f"{_n} installation cost is zero — proceed without it?")

    # Variable cost tables
    _rm_rows  = rm_active if not rm_active.empty else pd.DataFrame()
    _cu_rows  = cu_active if not cu_active.empty else pd.DataFrame()
    _cb_rows  = cb_active if not cb_active.empty else pd.DataFrame()

    if _rm_rows.empty or len(_rm_rows) == 0:
        warns.append("No raw materials defined — is this a process with no material inputs?")
    if _cu_rows.empty or len(_cu_rows) == 0:
        warns.append("No chemical inputs or utilities defined — proceed without utilities?")
    if _cb_rows.empty or len(_cb_rows) == 0:
        warns.append("No byproducts or credits defined — proceed without any revenue credits?")

    # Working hours
    if st.session_state.get("working_hours", 0.0) <= 0:
        warns.append("Working hours per year is zero or not set.")

    return crits, warns


# Compute validation state once (before rendering save UI)
_crits, _warns = _run_validations()

st.markdown("---")
st.markdown("#### Save Scenario")

# Show critical errors — always visible, block save
if _crits:
    for _msg in _crits:
        st.error(f"🚫 **{_msg}**")

# Show soft warnings with proceed checkbox
_proceed = True  # default: no warnings to acknowledge
if _warns:
    st.warning("⚠️ **The following issues were detected. Review before saving:**")
    for _msg in _warns:
        st.markdown(f"- {_msg}")
    _proceed = st.checkbox(
        "I have reviewed the above warnings and wish to proceed with saving.",
        key="save_proceed_checkbox",
    )

_save_blocked = bool(_crits) or not _proceed

if st.button(
    "Save / Update Scenario",
    type="primary",
    disabled=_save_blocked,
):
    util_bool = st.session_state.lang_utility if st.session_state.oth_cost_src == "Lang Factors" else False

    freight, taxes_permits, eng_ho, ga_overheads, contract_fee = (nf_values + [0.0]*5)[:5]
    piping, civil, steel, instrumentals, electrical, insulation = (inst_values + [0.0]*6)[:6]

    # Compute annual capacity for saving
    _ann_cap_val, _ann_cap_unit = annual_capacity(
        st.session_state.pc_input or 0.0,
        st.session_state.pu_input,
        working_hours,
    )

    st.session_state.scenarios[st.session_state.sn_input] = {
            "Product Name":              st.session_state.mp_input,
            "Unit":                      st.session_state.pu_input,
            "Capacity":                  st.session_state.pc_input,
            "Annual Capacity":           _ann_cap_val,
            "Annual Capacity Unit":      _ann_cap_unit,
            "Capacity Label":            f"{st.session_state.pc_input} {st.session_state.pu_input} → {_ann_cap_val:,.1f} {_ann_cap_unit}",
            "Process Variables":         edited_process_vars.dropna(how="all").to_dict(orient="records"),
            "Equipment Costs Source":    st.session_state.eq_cost_src,
            "Other Costs Source":        st.session_state.oth_cost_src,
            "Contains Utility Systems":  util_bool,
            "Product Type":              st.session_state.dm_prod_type,
            "TRL":                       st.session_state.dm_trl,
            "Info Availability":         st.session_state.dm_info_avail,
            "Process Severity":          st.session_state.dm_severity,
            "Material Handled":          st.session_state.dm_mat_type,
            "Plant Size":                st.session_state.dm_plant_size,
            # CAPEX
            "Equipment Acquisition":     equip_acq,
            "Spare Parts":               spare_parts,
            "Equipment Setting":         equip_setting,
            "Base Equipment Costs":      base_equip_costs,
            "Unscheduled Equip Pct":     unsched_pct,
            "Total Equipment Costs":     total_equip_costs,
            "Piping":                    piping,
            "Civil":                     civil,
            "Steel":                     steel,
            "Instrumentals":             instrumentals,
            "Electrical":                electrical,
            "Insulation":                insulation,
            "Paint":                     paint,
            "Total Installation Costs":  total_inst_costs,
            "Total Direct Field Costs":  total_direct_field_costs,
            "ISBL Contribution (%)":     isbl_contrib,
            "OSBL Contribution (%)":     100.0 - isbl_contrib,
            "Field Office Staff":        field_office,
            "Construction Indirects":    const_indirects,
            "Total Indirect Field Costs":total_indirect_field_costs,
            "Freight":                   freight,
            "Taxes and Permits":         taxes_permits,
            "Engineering and HO":        eng_ho,
            "GA Overheads":              ga_overheads,
            "Contract Fee":              contract_fee,
            "Total Non-Field Costs":     total_non_field_costs,
            "Project Costs ISBL+OSBL":   project_costs_isbl_osbl,
            "Allow Override":            st.session_state.allow_override,
            "Contingency Pct":           contingency_pct,
            "Databank Year":             int(databank_year_input),
            "Year of Analysis":          effective_analysis_year,
            "PCI Databank":              pci_databank,
            "PCI Analysis":              pci_analysis,
            "Time Update Factor":        time_update_factor,
            "Project Location":          st.session_state.proj_location,
            "Location Factor":           location_factor,
            "Project CAPEX":             project_capex,
            "Working Hours per Year":    working_hours,
            "Scaling Factor":            scaling_factor,
            # Variable costs
            "Raw Materials":                    rm_active.drop(columns=[c for c in ["_ann_cost","Annual Quantity","Annual Cost (USD/year)"] if c in rm_active.columns], errors="ignore").to_dict(orient="records") if not rm_active.empty else [],
            "Total Raw Material Cost":          total_raw_material_cost,
            "Chemical Inputs and Utilities":    cu_active.drop(columns=[c for c in ["_ann_cost","Annual Quantity","Annual Cost (USD/year)"] if c in cu_active.columns], errors="ignore").to_dict(orient="records") if not cu_active.empty else [],
            "Total Chemical Inputs Utilities":  total_chemical_utilities,
            "Credits and Byproducts":           cb_active.drop(columns=[c for c in ["_ann_cost","Annual Quantity","Annual Cost (USD/year)"] if c in cb_active.columns], errors="ignore").to_dict(orient="records") if not cb_active.empty else [],
            "Total Revenue":                    total_revenue,
            # Fixed costs
            "Num Operators":             st.session_state.get("n_operators", 2),
            "Operator Salary":           st.session_state.get("operator_salary", 0.0),
            "Num Supervisors":           st.session_state.get("n_supervisors", 1),
            "Supervisor Salary":         st.session_state.get("supervisor_salary", 0.0),
            "Salary Charges":            st.session_state.get("salary_charges", 2.2),
            "Plant Daily Hours":         st.session_state.get("plant_daily_hours", 24.0),
            "Weekly Op Days":            st.session_state.get("weekly_op_days", 7.0),
            "Operating Weeks":           operating_weeks,
            "Worker Hours per Shift":    st.session_state.get("worker_hours_shift", 8.0),
            "Worker Shifts per Week":    st.session_state.get("worker_shifts_week", 5.0),
            "Worker Vacation Weeks":     st.session_state.get("worker_vacation_weeks", 4.0),
            "Worker Weeks per Year":     worker_weeks_per_year,
            "Operating Team Factor":     op_team_factor,
            "OLC":                       olc,
            "Lab Charges Override":      st.session_state.get("lab_charges_override"),
            "Office Labor Override":     st.session_state.get("office_labor_override"),
            "Labor Working Hrs Override":st.session_state.get("labor_working_hrs_override"),
            "Lab Charges Pct":           lab_pct,
            "Office Labor Pct":          office_pct,
            "Total Labor Costs":         total_labor_costs,
            "Maint Repair Override":     st.session_state.get("maint_repair_override"),
            "Op Supplies Override":      st.session_state.get("op_supplies_override"),
            "Maint Pct":                 maint_pct,
            "Op Sup Pct":                op_sup_pct,
            "Supply Maint Costs":        supply_maint_costs,
            "Admin Overhead Override":   st.session_state.get("admin_overhead_override"),
            "Mfg Overhead Override":     st.session_state.get("mfg_overhead_override"),
            "Taxes Ins Override":        st.session_state.get("taxes_ins_override"),
            "Patents Roy Override":      st.session_state.get("patents_roy_override"),
            "Admin Ov Pct":              admin_ov_pct,
            "Mfg Ov Pct":               mfg_ov_pct,
            "Taxes Ins Pct":             taxes_ins_pct,
            "Patents Pct":               patents_pct,
            "AFC Pre Patents":           afc,
            "Direct Fixed Costs":        direct_fixed_costs,
            "Admin Costs Override":      st.session_state.get("admin_costs_override"),
            "Mfg Costs Override":        st.session_state.get("mfg_costs_override"),
            "Dist Selling Override":     st.session_state.get("dist_selling_override"),
            "R And D Override":          st.session_state.get("r_and_d_override"),
            "Admin Costs Pct":           admin_costs_pct,
            "Mfg Costs Pct":             mfg_costs_pct,
            "Dist Sell Pct":             dist_sell_pct,
            "R D Pct":                   r_d_pct,
            "OPEX":                      opex,
            "Indirect Fixed Costs":      indirect_fixed_costs,
            "Total Fixed Costs":         total_fixed_costs,
            "Total OPEX":                total_opex,
            # Working capital
            "WC Method":                 st.session_state.get("wc_method", "Percentage"),
            "WC Pct":                    st.session_state.get("wc_pct", 5.0),
            "WC Equiv Cash Days":        st.session_state.get("wc_equiv_cash_days", 30.0),
            "WC Raw Mat Days":           st.session_state.get("wc_raw_mat_days", 15.0),
            "WC Accounts Rec Days":      st.session_state.get("wc_accounts_rec_days", 30.0),
            "WC Accrued Payroll Days":   st.session_state.get("wc_accrued_payroll_days", 30.0),
            "WC Accounts Pay Days":      st.session_state.get("wc_accounts_pay_days", 30.0),
            "Working Capital":           working_capital,
            # Startup
            "Startup Method":            st.session_state.get("startup_method", "Multiple Factors"),
            "Startup Single Pct":        st.session_state.get("startup_single_pct", 8.0),
            "Startup Op Training Days":  st.session_state.get("startup_op_training_days", 150.0),
            "Startup Commerc Pct":       st.session_state.get("startup_commerc_pct", 5.0),
            "Startup Inefficiency Pct":  st.session_state.get("startup_inefficiency_pct", 4.0),
            "Startup Costs":             startup_costs,
            # Investment totals
            "Additional Costs":          additional_costs,
            "Total Investment":          total_investment,
            "TIC Lower Override":        st.session_state.get("tic_lower_override"),
            "TIC Upper Override":        st.session_state.get("tic_upper_override"),
            "TIC Lower Pct":             tic_lower_input,
            "TIC Upper Pct":             tic_upper_input,
            # Financial assumptions
            "Land Option":               land_option,
            "Land Buy Pct":              land_buy_pct,
            "Land Rent Pct":             land_rent_pct,
            "Land Buy Pct Override":     st.session_state.get("land_buy_pct_override"),
            "Land Rent Pct Override":    st.session_state.get("land_rent_pct_override"),
            "Depreciation Method":       depreciation_method,
            "Depreciation Years":        depreciation_years,
            "Residual Value Pct":        residual_value_pct,
            "Residual Value USD":        residual_value_usd,
            "Depreciable Base":          depreciable_base,
            "Tax Country":               tax_country,
            "Tax Rate Override":         st.session_state.get("tax_rate_override"),
            "Tax Rate":                  tax_rate,
            "Financing Type":            financing_type,
            "Debt Ratio Pct":            debt_ratio_pct,
            "Debt Ratio":                debt_ratio,
            "Amortization Years":        amortization_years,
            "Grace Period Years":        grace_period_years,
            "Central Bank Rate":         central_bank_rate,
            "Credit Spread":             credit_spread,
            "COD":                       cod,
            "Unlevered Beta":            unlevered_beta,
            "Levered Beta":              levered_beta,
            "Market Return":             market_return,
            "Risk Free Rate":            risk_free_rate,
            "Market Risk Premium":       market_risk_premium,
            "Country Risk Premium":      country_risk_premium,
            "US CPI":                    us_cpi,
            "Country CPI":              country_cpi,
            "COE":                       coe,
            "MARR Override":             st.session_state.get("marr_override"),
            "MARR":                      marr,
            # Project lifetime
            "EPC Years":                 epc_years,
            "Project Lifetime":          project_lifetime,
            "Total Project Lifetime":    epc_years + project_lifetime,
            "Capacity First Year":       capacity_first_year,
            "Capacity Intermediate":     capacity_intermediate,
            "Capacity Last Year":        capacity_last_year,
            "Fixed Costs First Year":    fixed_costs_first_year,
            "Fixed Costs Intermediate":  fixed_costs_intermediate,
            "Fixed Costs Last Year":     fixed_costs_last_year,
            "Growth Main Price":         growth_main_price,
            "Growth Byproduct Price":    growth_byproduct_price,
            "Growth Raw Materials":      growth_raw_materials,
            "Growth Chem Utilities":     growth_chem_utilities,
            "Growth Fixed Costs":        growth_fixed_costs,
            "Main Product Price":        main_product_price,
        }

    st.session_state.success_msg       = f"Scenario '{st.session_state.sn_input}' saved successfully!"
    st.session_state.clear_on_next_run = True
    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
# COMPILED SCENARIOS TABLE
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("#### Compiled Scenarios")
if st.session_state.scenarios:
    summary_rows = []
    for name, d in st.session_state.scenarios.items():
        is_hardcoded = name in HARDCODED_SCENARIOS
        summary_rows.append({
            "Scenario":         name,
            "Type":             "🔒 Hardcoded" if is_hardcoded else "✏️ User",
            "Product":          d.get("Product Name", ""),
            "Capacity":         d.get("Capacity Label", ""),
            "Eq. Cost Source":  d.get("Equipment Costs Source", ""),
            "Total Equip.":     fmt_curr(d.get("Total Equipment Costs", 0.0)),
            "Project CAPEX":    fmt_curr(d.get("Project CAPEX", 0.0)),
            "Total OPEX":       fmt_curr(d.get("Total OPEX", 0.0)),
            "TIC":              fmt_curr(d.get("Total Investment", 0.0)),
            "MARR":             f"{d.get('MARR', 0.0)*100:.2f}%",
        })
    st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

    col_clr, col_clr_usr = st.columns(2)
    with col_clr:
        if st.button("Clear user-defined scenarios", type="secondary"):
            for name in list(st.session_state.scenarios.keys()):
                if name not in HARDCODED_SCENARIOS:
                    del st.session_state.scenarios[name]
            st.rerun()
    with col_clr_usr:
        if st.button("Clear ALL scenarios (incl. hardcoded)", type="secondary"):
            st.session_state.scenarios = {}
            st.rerun()
else:
    st.info("No scenarios defined yet.")
